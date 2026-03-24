#!/usr/bin/env python3
"""
Generic financial model builder.

Reads  {RESEARCH_DIR}/{TICKER}/{TICKER}_tikr_financials.csv
       {RESEARCH_DIR}/{TICKER}/{TICKER}_tikr_forward_estimates.csv

Writes {RESEARCH_DIR}/{TICKER}/{TICKER}_Financial_Model.xlsx

Usage:
    python build_model.py TICKER [--company NAME] [--currency GBP] [--units M]
    python build_model.py GS --company "Goldman Sachs" --currency $ --units M
"""

import argparse
import csv
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

sys.path.insert(0, str(Path(__file__).resolve().parent))
from config import DEFAULT_RESEARCH_DIR

# ── COLOUR & FORMAT CONSTANTS (identical to build_dge_model.py) ───────────────
NAVY="1F3864"; BLUE_H="2F5496"; GREEN_H="375623"; SECT="D6E4F7"
GREEN_TOT="E2EFDA"; RED_BG="C00000"; YELLOW="FFFF00"; LT_BLUE="D9E1F2"
WHITE="FFFFFF"; BLF="0000FF"; BKF="000000"; GRF="008000"; GYF="888888"; WHF="FFFFFF"
SECT_TXT="1F3864"; BROWN="843C0C"; PURPLE="7030A0"
FMT_N='#,##0'; FMT_NN='#,##0;(#,##0);-'; FMT_P='0.0%'; FMT_EP='$#,##0.00'; FMT_MX='0.0x'

def fl(h): return PatternFill("solid", fgColor=h)
def fn(c=BKF,bold=False,italic=False,sz=10):
    return Font(color=c,bold=bold,italic=italic,size=sz,name="Calibri")
def bdr(c=BLUE_H):
    s=Side(style='medium',color=c); return Border(top=s)
CTR=Alignment(horizontal='center',vertical='center')
LFT=Alignment(horizontal='left',vertical='center')
IND=Alignment(horizontal='left',vertical='center',indent=2)

def w(ws,r,c,val=None,bg=WHITE,fc=BKF,bold=False,italic=False,fmt=None,
      align=CTR,mc=None,bc=None,sz=10):
    cell=ws.cell(row=r,column=c,value=val)
    cell.fill=fl(bg); cell.font=fn(fc,bold,italic,sz); cell.alignment=align
    if fmt: cell.number_format=fmt
    if mc: ws.merge_cells(start_row=r,start_column=c,end_row=r,end_column=mc)
    if bc: cell.border=bdr(bc)
    return cell

def spacer(ws,r,h=5):
    ws.row_dimensions[r].height=h
    for c in range(1,24): ws.cell(row=r,column=c).fill=fl(WHITE)


# ── CSV PARSING ───────────────────────────────────────────────────────────────
def clean_num(s):
    """Parse TIKR numeric string (handles parentheses negatives, % suffix, commas)."""
    if s is None: return None
    s=str(s).strip()
    if s in ('','-','N/A','NM','n/a','--','nan'): return None
    neg=s.startswith('(') and s.endswith(')')
    if neg: s=s[1:-1]
    is_pct=s.endswith('%')
    if is_pct: s=s[:-1]
    try:
        v=float(s.replace(',','').replace(' ',''))
    except ValueError:
        return None
    if neg: v=-v
    if is_pct: v=v/100
    return v


def parse_tikr_financials(path: Path) -> dict:
    """Parse multi-section TIKR financials CSV → {section: {headers, rows}}."""
    sections={}; current=None; year_hdrs=[]
    with open(path, encoding='utf-8-sig', newline='') as f:
        for row in csv.reader(f):
            if not row: continue
            first=row[0].strip()
            m=re.match(r'^===\s*(.+?)\s*===', first)
            if m:
                current=m.group(1)
                sections[current]={'headers':[],'rows':{}}
                year_hdrs=[]; continue
            if current is None: continue
            if '| TIKR' in first.upper() or '| tikr' in first:
                year_hdrs=[h.strip() for h in row[1:] if h.strip()
                           and 'LTM' not in h and 'TTM' not in h]
                sections[current]['headers']=year_hdrs; continue
            if year_hdrs and first:
                vals=[clean_num(row[i+1] if i+1<len(row) else '') for i in range(len(year_hdrs))]
                sections[current]['rows'][first]=vals
    return sections


def parse_tikr_estimates(path: Path) -> dict:
    """Parse TIKR estimates CSV → {headers, rows}."""
    headers=None; rows={}
    with open(path, encoding='utf-8-sig', newline='') as f:
        for row in csv.reader(f):
            if not row: continue
            first=row[0].strip()
            if not first: continue
            if '| TIKR' in first.upper() or '| tikr' in first:
                headers=[h.strip() for h in row[1:] if h.strip() and 'CAGR' not in h.upper()]
                continue
            if headers is None: continue
            if first and first not in rows:
                vals=[clean_num(row[i+1] if i+1<len(row) else '') for i in range(len(headers))]
                rows[first]=vals
    return {'headers':headers or [],'rows':rows}


def parse_sec_combined(ticker_dir: Path) -> dict:
    """
    Parse SEC EDGAR combined CSVs (from extract_sec.py / combine_financials.py).

    Looks in <ticker_dir>/combined/ for:
        {TICKER}_10K_income_statement_combined.csv
        {TICKER}_10K_balance_sheet_combined.csv
        {TICKER}_10K_cash_flow_statement_combined.csv

    Returns same structure as parse_tikr_financials():
        {section: {'headers': [yr_labels], 'rows': {label: [vals…]}}}

    Values are divided by 1,000,000 to convert absolute dollars → millions.
    """
    ticker = ticker_dir.name
    combined_dir = ticker_dir / "combined"
    sections_map = [
        ('Income Statement',      f'{ticker}_10K_income_statement_combined.csv'),
        ('Balance Sheet',         f'{ticker}_10K_balance_sheet_combined.csv'),
        ('Cash Flow Statement',   f'{ticker}_10K_cash_flow_statement_combined.csv'),
    ]
    result = {}
    for section_name, fname in sections_map:
        path = combined_dir / fname
        if not path.exists():
            result[section_name] = {'headers': [], 'rows': {}}
            continue
        with open(path, encoding='utf-8-sig', newline='') as f:
            reader = csv.reader(f)
            header = next(reader)       # concept, label, date1, date2, …
        raw_dates = header[2:]
        # Format dates as M/D/YY to match TIKR style
        def _fmt(d):
            try:
                from datetime import datetime
                dt = datetime.strptime(d[:10], '%Y-%m-%d')
                return f"{dt.month}/{dt.day}/{str(dt.year)[2:]}"
            except Exception:
                return d[:10]
        years = [_fmt(d) for d in raw_dates]
        rows = {}
        with open(path, encoding='utf-8-sig', newline='') as f:
            reader = csv.reader(f)
            next(reader)  # skip header
            for data_row in reader:
                if len(data_row) < 2:
                    continue
                label = data_row[1].strip()
                if not label:
                    continue
                vals = []
                for v in data_row[2:]:
                    try:
                        num = float(v)
                        vals.append(num / 1_000_000)   # absolute → millions
                    except (ValueError, TypeError):
                        vals.append(None)
                rows[label] = vals
        result[section_name] = {'headers': years, 'rows': rows}
    return result


def find_val(rows: dict, *keys):
    """Case-insensitive substring search; returns first matching value list."""
    for key in keys:
        kl=key.lower()
        for label,vals in rows.items():
            if kl in label.lower(): return vals
    return None


def last_n(vals, n, default=0.0):
    """Last non-None value in the first n elements of vals."""
    if not vals: return default
    for i in range(min(n,len(vals))-1,-1,-1):
        if i<len(vals) and vals[i] is not None: return vals[i]
    return default


def gv(vals, n, default=0.0):
    """Get exactly n values, padding/truncating with default."""
    if not vals: return [default]*n
    return [(vals[i] if i<len(vals) and vals[i] is not None else default) for i in range(n)]


def fv_act(rows, n_act, *keys, negate=False, default=0.0):
    """Get n_act actuals for given keys; return (values_list, found_bool)."""
    vals=find_val(rows,*keys)
    if vals is not None:
        act=gv(vals,n_act,default)
        if negate: act=[-abs(v) for v in act]
        return act, True
    return [default]*n_act, False


# ── ROW CONSTANTS (same numbers as DGE — keeps cross-sheet formulas stable) ────
IS_R = {
    'rev':6,'tot_rev':7,'cogs':11,'gp':14,'sga':18,'ooe':19,'topex':20,
    'oi':22,'int_exp':26,'int_inc':27,'eq_inc':28,'oth_no':29,
    'ebt_ex':30,'unusual':32,'ebt':33,'tax':34,'econt':35,
    'mi':37,'ni':38,'dshares':42,'deps':43,'bshares':44,'beps':45,'dps':46,
    'rd':49,'da':50,'ebitda':51,'ebitdam':52
}
BS_R = {
    'cash':7,'st_inv':8,'tot_cash':9,'ar':10,'oth_recv':11,'tot_recv':12,
    'inv':13,'prepaid':14,'oth_ca':15,'tot_ca':16,
    'ppe':18,'lt_inv':19,'goodwill':20,'intang':21,'dta_lt':22,'oth_lta':23,
    'tot_nca':24,'tot_assets':26,
    'ap':30,'accrued':31,'st_borrow':32,'curr_ltd':33,'curr_lease':34,
    'tax_pay':35,'unearned':36,'oth_cl':37,'tot_cl':38,
    'lt_debt':40,'lt_lease':41,'pension':42,'dtl_nc':43,'oth_ncl':44,
    'tot_ncl':45,'tot_liab':47,
    'comm_stk':50,'apic':51,'ret_earn':52,'treasury':53,'aoci':54,
    'tot_ceq':55,'mi_bs':56,'tot_equity':57,'tot_le':59,'bal_chk':60
}
CFS_R = {
    'ni':6,'da':7,'amort_gi':8,'amort_def':9,'writedown':10,'eq_rev':11,
    'oth_ops':12,'d_ar':13,'d_inv':14,'d_ap':15,'cfo':16,
    'capex':19,'pp_sale':20,'acquisitions':21,'divestitures':22,
    'inv_sec':23,'net_loans':24,'cfi':25,
    'debt_issued':28,'debt_repaid':29,'stk_issued':30,'buybacks':31,
    'dividends':32,'oth_fin':33,'cff':34,
    'fx':36,'misc':37,'net_chg':38,'beg_cash':39,'end_cash':41,
    'fcf':44,'fcf_margin':45
}


# ── MODEL DATA EXTRACTION ─────────────────────────────────────────────────────
def build_model_data(ticker, company, currency, units, fin_path, est_path):
    if fin_path and fin_path.exists():
        fin = parse_tikr_financials(fin_path)
    else:
        # Fall back to SEC combined CSVs (US tickers via extract_sec.py)
        ticker_dir = fin_path.parent if fin_path else est_path.parent
        fin = parse_sec_combined(ticker_dir)
    est=parse_tikr_estimates(est_path)

    # Section lookup (case-insensitive)
    def get_sec(d, *names):
        for name in names:
            nl=name.lower()
            for k in d:
                if nl in k.lower(): return d[k]
        return {}

    is_sec=get_sec(fin,'income statement')
    bs_sec=get_sec(fin,'balance sheet')
    cfs_sec=get_sec(fin,'cash flow statement','cash flow')

    is_rows=is_sec.get('rows',{}); bs_rows=bs_sec.get('rows',{}); cfs_rows=cfs_sec.get('rows',{})
    est_rows=est['rows']

    act_yrs=is_sec.get('headers',[]); est_yrs=est['headers']
    n_act=len(act_yrs); n_est=len(est_yrs)

    if n_act==0: raise ValueError(f"No actual years found in {fin_path}")
    if n_est==0: raise ValueError(f"No estimate years found in {est_path}")

    # ── Key last-actual values for ratio computation ──────────────────────────
    def la_is(*keys): return last_n(find_val(is_rows,*keys), n_act, 0.0) or 0.0
    def la_bs_v(*keys): return last_n(find_val(bs_rows,*keys), n_act, 0.0) or 0.0
    def la_cfs(*keys): return last_n(find_val(cfs_rows,*keys), n_act, 0.0) or 0.0

    la_rev=abs(la_is('revenues','revenue','net revenues','net sales','total revenue'))
    la_cogs=abs(la_is('cost of goods sold','cost of revenue','cost of products sold','cost of sales'))
    la_gp=abs(la_is('gross profit','gross income'))
    la_sga=abs(la_is('selling, general and admin','selling, general & admin','sg&a','selling general'))
    la_oi=la_is('operating income','income from operations','operating profit')
    la_da=abs(la_is('depreciation','d&a','depreciation and amortization','depreciation & amortization'))
    la_dil=abs(la_is('diluted shares','weighted average diluted','diluted weighted average','diluted (in shares)','(in shares)'))
    la_bas=abs(la_is('basic shares','weighted average basic','basic weighted average'))
    la_dps=abs(la_is('dividends per share','dividend per share'))
    la_rd=abs(la_is('research and development','r&d','research & development'))

    la_ar=abs(la_bs_v('accounts receivable','trade receivable','net receivable'))
    la_inv=abs(la_bs_v('inventor'))
    la_ap=abs(la_bs_v('accounts payable','trade payable','trade and other payable'))
    la_cash=abs(la_bs_v('cash and cash equiv','cash & cash equiv','cash, cash equiv'))
    la_st_inv=abs(la_bs_v('short-term invest','marketable securities','available-for-sale'))
    la_lt_debt=abs(la_bs_v('long-term debt','long term debt','long-term borrowing'))

    la_capex=abs(la_cfs('capital expenditure','capital expenditures','purchase of property',
                        'purchases of property','purchase of property, plant'))

    # Working capital ratios from last actual
    dso=int(la_ar/(la_rev/365)) if la_rev and la_ar else 45
    dio=int(la_inv/(la_cogs/365)) if la_cogs and la_inv else 60
    dpo=int(la_ap/(la_cogs/365)) if la_cogs and la_ap else 45
    dso=min(max(dso,10),365); dio=min(max(dio,0),1825); dpo=min(max(dpo,10),365)

    # Derived % drivers from actuals
    gm_pct=la_gp/la_rev if la_rev else 0.50
    sga_pct=la_sga/la_rev if la_rev else 0.20
    ooe_pct=abs(la_gp-la_sga-la_oi)/la_rev if la_rev else 0.10
    capex_pct=la_capex/la_rev if la_rev else 0.05

    # ── Pull estimates from estimates CSV ────────────────────────────────────
    def est_n(default_list, *keys, raw_pct=False):
        """Get n_est values for given keys; auto-scale if raw_pct and >1."""
        vals=find_val(est_rows,*keys)
        result=[]
        for i in range(n_est):
            v=(vals[i] if vals and i<len(vals) else None)
            if v is None:
                v=default_list[i] if i<len(default_list) else default_list[-1]
            if raw_pct and v is not None and abs(v)>1.5: v=v/100
            result.append(v)
        return result

    a_rev_growth=est_n([0.03]*n_est,'% change yoy','revenue growth','net revenue growth')
    a_gross_margin=est_n([gm_pct]*n_est,'gross margin',raw_pct=True)
    a_sga_pct=[sga_pct]*n_est
    a_ooe_pct=[ooe_pct]*n_est
    a_int_exp=est_n([la_is('interest expense')]*n_est,'interest expense')
    a_int_inc=[abs(la_is('interest income','interest & investment income','interest and investment income'))]*n_est
    a_eq_inc=[la_is('equity income','income on equity invest','equity in earnings')]*n_est
    a_oth_no=[la_is('other non-operating','other income (expense)','other non-operating income')]*n_est
    a_unusual=[0.0]*n_est
    a_tax_rate=est_n([0.25]*n_est,'effective tax rate','tax rate',raw_pct=True)
    a_mi=[la_is('minority interest','non-controlling interest')]*n_est
    a_dil_shares=[max(1.0, la_dil*(1-0.005*(i+1))) for i in range(n_est)]
    a_bas_shares=[max(1.0, la_bas*(1-0.005*(i+1))) for i in range(n_est)]
    a_dps=[la_dps*(1+0.03)**i for i in range(n_est)]
    a_rd=[la_rd]*n_est
    a_da_raw=est_n([la_da*(1+0.02*i) for i in range(n_est)],
                  'depreciation & amortization','depreciation and amortization')
    a_da=[abs(v) for v in a_da_raw]

    def la_bsf(*keys): return last_n(find_val(bs_rows,*keys), n_act, 0.0) or 0.0

    asmp={
        'rev_growth':a_rev_growth,'gross_margin':a_gross_margin,
        'sga_pct':a_sga_pct,'ooe_pct':a_ooe_pct,
        'int_exp':a_int_exp,'int_inc':a_int_inc,'eq_inc':a_eq_inc,
        'oth_no':a_oth_no,'unusual':a_unusual,'tax_rate':a_tax_rate,
        'mi':a_mi,'dil_shares':a_dil_shares,'bas_shares':a_bas_shares,
        'dps':a_dps,'rd':a_rd,'da':a_da,
        'dso':[float(dso)]*n_est,'dio':[float(dio)]*n_est,'dpo':[float(dpo)]*n_est,
        'oth_ca':[abs(la_bsf('other current assets'))]*n_est,
        'prepaid':[abs(la_bsf('prepaid'))]*n_est,
        'ppe':[abs(la_bsf('net property','property plant and equipment','property, plant'))]*n_est,
        'lt_inv':[abs(la_bsf('long-term invest','long term invest'))]*n_est,
        'goodwill':[abs(la_bsf('goodwill'))]*n_est,
        'intang':[abs(la_bsf('intangible','other intangible'))]*n_est,
        'dta_lt':[abs(la_bsf('deferred tax asset'))]*n_est,
        'oth_lta':[abs(la_bsf('other long-term asset','other non-current asset'))]*n_est,
        'accrued':[abs(la_bsf('accrued expense','accrued liab'))]*n_est,
        'st_borrow':[abs(la_bsf('short-term borrow'))]*n_est,
        'curr_ltd':[abs(la_bsf('current portion of long-term','current maturities'))]*n_est,
        'curr_lease':[abs(la_bsf('current portion of capital','current lease'))]*n_est,
        'tax_pay':[abs(la_bsf('income taxes payable','taxes payable'))]*n_est,
        'unearned':[abs(la_bsf('unearned revenue','deferred revenue'))]*n_est,
        'oth_cl':[abs(la_bsf('other current liab'))]*n_est,
        'lt_debt':[la_lt_debt]*n_est,
        'lt_lease':[abs(la_bsf('capital lease','finance lease long','operating lease long'))]*n_est,
        'pension':[abs(la_bsf('pension','post-retirement'))]*n_est,
        'dtl_nc':[abs(la_bsf('deferred tax liab'))]*n_est,
        'oth_ncl':[abs(la_bsf('other non-current liab','other long-term liab'))]*n_est,
        'capex_pct':[capex_pct]*n_est,
        'pp_sale':[la_cfs('sale of property','proceeds from sale of property')]*n_est,
        'acquisitions':[la_cfs('cash acquisition','acquisition','business acquisition')]*n_est,
        'divestitures':[la_cfs('divestiture','proceeds from divestiture','disposal of business')]*n_est,
        'oth_inv':[la_cfs('other investing activities','other investing')]*n_est,
        'debt_issued':[la_cfs('total debt issued','proceeds from debt','issuance of long-term debt',
                              'proceeds from issuance of long-term debt')]*n_est,
        'debt_repaid':[la_cfs('total debt repaid','repayment of debt','repayment of long-term debt')]*n_est,
        'stk_issued':[la_cfs('issuance of common stock','proceeds from issuance of common')]*n_est,
        'buybacks':[la_cfs('repurchase of common stock','purchase of common stock','buyback')]*n_est,
        'oth_fin':[la_cfs('other financing activities','other financing')]*n_est,
        'fx':[la_cfs('effect of','fx adjustment','foreign exchange effect')]*n_est,
        'misc':[la_cfs('miscellaneous','other cash flow adjustment')]*n_est,
    }

    la_total_debt=la_lt_debt+abs(la_bsf('short-term borrow'))+abs(la_bsf('current portion of long-term debt','current maturities'))

    return {
        'ticker':ticker,'company':company,'currency':currency,'units':units,
        'act_yrs':act_yrs,'est_yrs':est_yrs,'n_act':n_act,'n_est':n_est,
        'is_rows':is_rows,'bs_rows':bs_rows,'cfs_rows':cfs_rows,'est_rows':est_rows,
        'asmp':asmp,
        'dcf':{'wacc':0.09,'tgr':0.025,'cash':la_cash,'st_inv':la_st_inv,
               'total_debt':la_total_debt,'dil_shares':la_dil,'share_price':0.0,'ntm_ev_ebitda':12.0},
    }


# ═══════════════════════════════════════════════════════════════════════════════
# ASSUMPTIONS TAB
# ═══════════════════════════════════════════════════════════════════════════════
def build_assumptions(wb, md):
    n_est=md['n_est']; asmp=md['asmp']
    ticker=md['ticker']; company=md['company']
    curr=md['currency']; u=md['units']; est_yrs=md['est_yrs']
    nc=3+n_est  # last column index for merge

    ws=wb.create_sheet("Assumptions")
    ws.sheet_properties.tabColor=PURPLE
    ws.column_dimensions['A'].width=1
    ws.column_dimensions['B'].width=38
    for i in range(n_est): ws.column_dimensions[get_column_letter(3+i)].width=13
    ws.column_dimensions[get_column_letter(3+n_est)].width=45

    ws.row_dimensions[1].height=3.75
    w(ws,2,2,f"{company} ({ticker}) — Assumptions ({curr}{u})",NAVY,WHF,True,sz=13,align=LFT,mc=nc)
    ws.row_dimensions[2].height=22
    hdr_labels=["Assumption"]+est_yrs+["Notes"]
    for i,lbl in enumerate(hdr_labels):
        is_e=isinstance(lbl,str) and lbl.strip().endswith('E')
        w(ws,3,2+i,lbl,GREEN_H if is_e else BLUE_H,WHF,True,align=CTR)
    ws.row_dimensions[3].height=18

    def ar(row, lbl, vals, fmt, note):
        w(ws,row,2,lbl,WHITE,BKF,align=LFT)
        for i,v in enumerate(vals): w(ws,row,3+i,v,WHITE,BLF,fmt=fmt)
        w(ws,row,3+n_est,note,WHITE,GYF,italic=True,align=LFT)

    spacer(ws,4)
    w(ws,5,2,"INCOME STATEMENT DRIVERS",SECT,SECT_TXT,True,align=LFT,mc=nc)
    ar(6,"Revenue Growth %",asmp['rev_growth'],FMT_P,"TIKR consensus estimates")
    ar(7,"Gross Margin %",asmp['gross_margin'],FMT_P,"TIKR consensus estimates")
    ar(8,"SG&A % of Revenue",asmp['sga_pct'],FMT_P,"Average from actuals")
    ar(9,"Other OpEx % of Revenue",asmp['ooe_pct'],FMT_P,"Implied from EBIT margin")
    ar(10,f"Interest Expense ({curr}{u})",asmp['int_exp'],FMT_NN,"TIKR consensus / last actual")
    ar(11,f"Interest & Investment Income ({curr}{u})",asmp['int_inc'],FMT_NN,"Last actual level")
    ar(12,f"Equity Income ({curr}{u})",asmp['eq_inc'],FMT_NN,"Last actual level")
    ar(13,f"Other Non-Operating Income ({curr}{u})",asmp['oth_no'],FMT_NN,"Last actual level")
    ar(14,f"Unusual / Non-Recurring Items ({curr}{u})",asmp['unusual'],FMT_NN,"Normalized to zero")
    ar(15,"Tax Rate %",asmp['tax_rate'],FMT_P,"TIKR consensus / effective rate")
    ar(16,f"Minority Interest ({curr}{u})",asmp['mi'],FMT_NN,"Last actual level")
    ar(17,f"Diluted Shares ({u})",asmp['dil_shares'],FMT_N,"Last actual; slight decline assumed")
    ar(18,f"Basic Shares ({u})",asmp['bas_shares'],FMT_N,"Consistent with diluted")
    ar(19,f"Dividends Per Share ({curr})",asmp['dps'],FMT_EP,"Last actual; 3% annual growth")
    ar(20,f"R&D Expense ({curr}{u})",asmp['rd'],FMT_NN,"Flat at last actual level")
    ar(21,f"Depreciation & Amortisation ({curr}{u})",asmp['da'],FMT_NN,"TIKR consensus / last actual")

    spacer(ws,22)
    w(ws,23,2,"BALANCE SHEET DRIVERS",SECT,SECT_TXT,True,align=LFT,mc=nc)
    bs_rows_asmp=[
        (24,"DSO — Accounts Receivable (days)",asmp['dso'],FMT_N,"Computed from last actual"),
        (25,"DIO — Inventory (days)",asmp['dio'],FMT_N,"Computed from last actual"),
        (26,"DPO — Accounts Payable (days)",asmp['dpo'],FMT_N,"Computed from last actual"),
        (27,f"Other Current Assets ({curr}{u})",asmp['oth_ca'],FMT_NN,"Flat at last actual"),
        (28,f"Prepaid Expenses ({curr}{u})",asmp['prepaid'],FMT_NN,"Flat at last actual"),
        (29,f"Net PP&E ({curr}{u})",asmp['ppe'],FMT_NN,"Flat at last actual"),
        (30,f"Long-Term Investments ({curr}{u})",asmp['lt_inv'],FMT_NN,"Flat at last actual"),
        (31,f"Goodwill ({curr}{u})",asmp['goodwill'],FMT_NN,"Flat at last actual"),
        (32,f"Other Intangibles ({curr}{u})",asmp['intang'],FMT_NN,"Flat at last actual"),
        (33,f"Deferred Tax Assets LT ({curr}{u})",asmp['dta_lt'],FMT_NN,"Flat at last actual"),
        (34,f"Other Long-Term Assets ({curr}{u})",asmp['oth_lta'],FMT_NN,"Flat at last actual"),
        (35,f"Accrued Expenses ({curr}{u})",asmp['accrued'],FMT_NN,"Flat at last actual"),
        (36,f"Short-Term Borrowings ({curr}{u})",asmp['st_borrow'],FMT_NN,"Flat at last actual"),
        (37,f"Current Portion of LT Debt ({curr}{u})",asmp['curr_ltd'],FMT_NN,"Flat at last actual"),
        (38,f"Current Capital Lease Obligations ({curr}{u})",asmp['curr_lease'],FMT_NN,"Flat at last actual"),
        (39,f"Current Income Taxes Payable ({curr}{u})",asmp['tax_pay'],FMT_NN,"Flat at last actual"),
        (40,f"Unearned Revenue — Current ({curr}{u})",asmp['unearned'],FMT_NN,"Flat at last actual"),
        (41,f"Other Current Liabilities ({curr}{u})",asmp['oth_cl'],FMT_NN,"Flat at last actual"),
        (42,f"Long-Term Debt ({curr}{u})",asmp['lt_debt'],FMT_NN,"Flat at last actual"),
        (43,f"Capital Leases — Long-Term ({curr}{u})",asmp['lt_lease'],FMT_NN,"Flat at last actual"),
        (44,f"Pension & Post-Retirement ({curr}{u})",asmp['pension'],FMT_NN,"Flat at last actual"),
        (45,f"Deferred Tax Liability NC ({curr}{u})",asmp['dtl_nc'],FMT_NN,"Flat at last actual"),
        (46,f"Other Non-Current Liabilities ({curr}{u})",asmp['oth_ncl'],FMT_NN,"Flat at last actual"),
    ]
    for row,lbl,vals,fmt,note in bs_rows_asmp: ar(row,lbl,vals,fmt,note)

    spacer(ws,47)
    w(ws,48,2,"CASH FLOW DRIVERS",SECT,SECT_TXT,True,align=LFT,mc=nc)
    cfs_rows_asmp=[
        (49,"Capex % of Revenue",asmp['capex_pct'],FMT_P,"Computed from last actual"),
        (50,f"Sale of PP&E ({curr}{u})",asmp['pp_sale'],FMT_NN,"Last actual level"),
        (51,f"Cash Acquisitions ({curr}{u})",asmp['acquisitions'],FMT_NN,"Last actual level"),
        (52,f"Divestitures ({curr}{u})",asmp['divestitures'],FMT_NN,"Last actual level"),
        (53,f"Investment in Securities ({curr}{u})",asmp['oth_inv'],FMT_NN,"Last actual level"),
        (54,f"Net Loans Originated ({curr}{u})",[0.0]*n_est,FMT_NN,"Zero assumed"),
        (55,f"Total Debt Issued ({curr}{u})",asmp['debt_issued'],FMT_NN,"Last actual level"),
        (56,f"Total Debt Repaid ({curr}{u})",asmp['debt_repaid'],FMT_NN,"Last actual level"),
        (57,f"Issuance of Common Stock ({curr}{u})",asmp['stk_issued'],FMT_NN,"Last actual level"),
        (58,f"Repurchase of Common Stock ({curr}{u})",asmp['buybacks'],FMT_NN,"Last actual level"),
        (59,f"Other Financing Activities ({curr}{u})",asmp['oth_fin'],FMT_NN,"Last actual level"),
        (60,f"FX Adjustments on Cash ({curr}{u})",asmp['fx'],FMT_NN,"Last actual level"),
        (61,f"Miscellaneous Cash Adjustments ({curr}{u})",asmp['misc'],FMT_NN,"Last actual level"),
    ]
    for row,lbl,vals,fmt,note in cfs_rows_asmp: ar(row,lbl,vals,fmt,note)

    spacer(ws,62)
    w(ws,63,2,"MODEL NOTES",SECT,SECT_TXT,True,align=LFT,mc=nc)
    notes=[
        (64,"COLOUR CODING",True),
        (65,"Blue (#0000FF) — hardcoded inputs (actuals & user-editable assumptions)",False),
        (66,"Black (#000000) — calculated formulas",False),
        (67,"Green (#008000) — cross-sheet links pulling from another tab",False),
        (68,"Grey italic (#888888) — helper / ratio rows (margins, growth %)",False),
        (69,f"UNITS: All monetary values in {curr} {u}.",False),
    ]
    for row,txt,bold in notes:
        w(ws,row,2,txt,WHITE,GYF,bold=bold,italic=not bold,align=LFT,mc=nc)
    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# INCOME STATEMENT TAB
# ═══════════════════════════════════════════════════════════════════════════════
def build_is(wb, md):
    n_act=md['n_act']; n_est=md['n_est']
    act_yrs=md['act_yrs']; est_yrs=md['est_yrs']
    is_rows=md['is_rows']; asmp=md['asmp']
    ticker=md['ticker']; company=md['company']; curr=md['currency']; u=md['units']
    nc=2+n_act+n_est

    ws=wb.create_sheet("IS")
    ws.sheet_properties.tabColor=BLUE_H
    ws.column_dimensions['A'].width=1; ws.column_dimensions['B'].width=35
    for i in range(n_act+n_est): ws.column_dimensions[get_column_letter(3+i)].width=13
    ws.row_dimensions[1].height=3.75
    w(ws,2,2,f"{company} ({ticker}) — Income Statement ({curr}{u})",NAVY,WHF,True,sz=13,align=LFT,mc=nc)
    ws.row_dimensions[2].height=22
    for i,lbl in enumerate([""]+act_yrs+est_yrs):
        is_e=isinstance(lbl,str) and lbl.strip().endswith('E')
        w(ws,3,2+i,lbl,GREEN_H if is_e else BLUE_H,WHF,True,align=CTR)
    ws.row_dimensions[3].height=18

    # Column helpers
    n=n_act; m=n_est
    def ac(i): return get_column_letter(3+i)          # actual col i
    def ec(i): return get_column_letter(3+n+i)         # estimate col i
    def asmpC(i): return get_column_letter(3+i)        # assumptions col i
    lac=get_column_letter(3+n-1)                       # last actual col
    ec_all=[ec(i) for i in range(m)]
    asmpC_all=[asmpC(i) for i in range(m)]

    def blue_act(row, vals):
        for i,v in enumerate(vals): w(ws,row,3+i,v,WHITE,BLF,fmt=FMT_NN)

    spacer(ws,4)
    w(ws,5,2,"REVENUE",SECT,SECT_TXT,True,align=LFT,mc=nc)

    # Row 6: Revenue
    rev_act,_=fv_act(is_rows,n,'revenues','revenue','net revenues','net sales','total revenues')
    w(ws,6,2,"Net Revenue / Revenues",WHITE,BKF,align=LFT)
    blue_act(6,rev_act)
    for i in range(m):
        prev=lac if i==0 else ec(i-1)
        w(ws,6,3+n+i,f"={prev}6*(1+Assumptions!{asmpC(i)}6)",WHITE,BKF,fmt=FMT_NN)

    # Row 7: Total Revenue (formula)
    w(ws,7,2,"Total Revenue",WHITE,BKF,True,align=LFT)
    for c in range(3,3+n+m):
        cl=get_column_letter(c)
        w(ws,7,c,f"={cl}6",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)

    # Row 8: YoY Growth %
    w(ws,8,2,"  % Change YoY",WHITE,GYF,italic=True,align=IND)
    w(ws,8,3,"",WHITE,GYF)
    for c in range(4,3+n+m):
        cl=get_column_letter(c); pc=get_column_letter(c-1)
        w(ws,8,c,f"={cl}7/{pc}7-1",WHITE,GYF,italic=True,fmt=FMT_P)

    spacer(ws,9)
    w(ws,10,2,"COST OF REVENUE",SECT,SECT_TXT,True,align=LFT,mc=nc)

    # Row 11: COGS (negative)
    cogs_act,_=fv_act(is_rows,n,'cost of goods sold','cost of revenue','cost of products sold','cost of sales',negate=True)
    cogs_act=[-abs(v) for v in cogs_act]
    w(ws,11,2,"Cost of Goods Sold / Cost of Revenue",WHITE,BKF,align=LFT)
    blue_act(11,cogs_act)
    for i in range(m):
        w(ws,11,3+n+i,f"=-{ec(i)}7*(1-Assumptions!{asmpC(i)}7)",WHITE,BKF,fmt=FMT_NN)

    # Row 12: Total Cost (formula)
    w(ws,12,2,"Total Cost of Revenue",WHITE,BKF,True,align=LFT)
    for c in range(3,3+n+m):
        cl=get_column_letter(c)
        w(ws,12,c,f"={cl}11",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)

    spacer(ws,13)
    # Row 14: Gross Profit
    gp_act,gp_found=fv_act(is_rows,n,'gross profit','gross income')
    w(ws,14,2,"Gross Profit",WHITE,BKF,True,align=LFT)
    if gp_found:
        for i,v in enumerate(gp_act): w(ws,14,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    else:
        for c in range(3,3+n):
            cl=get_column_letter(c)
            w(ws,14,c,f"={cl}7+{cl}12",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(m):
        w(ws,14,3+n+i,f"={ec(i)}7+{ec(i)}12",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)

    # Row 15: Gross Margin %
    w(ws,15,2,"  % Gross Margin",WHITE,GYF,italic=True,align=IND)
    for c in range(3,3+n+m):
        cl=get_column_letter(c)
        w(ws,15,c,f"={cl}14/{cl}7",WHITE,GYF,italic=True,fmt=FMT_P)

    spacer(ws,16)
    w(ws,17,2,"OPERATING EXPENSES",SECT,SECT_TXT,True,align=LFT,mc=nc)

    # Row 18: SG&A
    sga_act,_=fv_act(is_rows,n,'selling, general and admin','selling, general & admin','sg&a','selling general',negate=True)
    sga_act=[-abs(v) for v in sga_act]
    w(ws,18,2,"Selling, General & Administrative Expenses",WHITE,BKF,align=LFT)
    blue_act(18,sga_act)
    for i in range(m):
        w(ws,18,3+n+i,f"=-{ec(i)}7*Assumptions!{asmpC(i)}8",WHITE,BKF,fmt=FMT_NN)

    # Row 19: Other OpEx (derive from GP - SGA - OI if not in CSV)
    ooe_act,ooe_found=fv_act(is_rows,n,'other operating expense','other operating activities')
    oi_act_raw,_=fv_act(is_rows,n,'operating income','income from operations','operating profit','ebit')
    if not ooe_found:
        if gp_found:
            ooe_act=[-(abs(gp)-abs(sga)-abs(oi)) for gp,sga,oi in zip(gp_act,sga_act,oi_act_raw)]
        else:
            ooe_act=[0.0]*n
    ooe_act=[-abs(v) for v in ooe_act]
    w(ws,19,2,"Other Operating Expenses",WHITE,BKF,align=LFT)
    blue_act(19,ooe_act)
    for i in range(m):
        w(ws,19,3+n+i,f"=-{ec(i)}7*Assumptions!{asmpC(i)}9",WHITE,BKF,fmt=FMT_NN)

    # Row 20: Total OpEx
    w(ws,20,2,"Total Operating Expenses",WHITE,BKF,True,align=LFT)
    topex_act=[a+b for a,b in zip(sga_act,ooe_act)]
    for i,v in enumerate(topex_act): w(ws,20,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(m):
        w(ws,20,3+n+i,f"={ec(i)}18+{ec(i)}19",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)

    spacer(ws,21)
    # Row 22: EBIT
    oi_act,oi_found=fv_act(is_rows,n,'operating income','income from operations','operating profit')
    if not oi_found:
        oi_act=[gp+topex for gp,topex in zip([a+b for a,b in zip(rev_act,cogs_act)],topex_act)]
    w(ws,22,2,"Operating Income (EBIT)",WHITE,BKF,True,align=LFT)
    for i,v in enumerate(oi_act): w(ws,22,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(m):
        w(ws,22,3+n+i,f"={ec(i)}14+{ec(i)}20",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)

    # Row 23: EBIT Margin %
    w(ws,23,2,"  % EBIT Margin",WHITE,GYF,italic=True,align=IND)
    for c in range(3,3+n+m):
        cl=get_column_letter(c)
        w(ws,23,c,f"={cl}22/{cl}7",WHITE,GYF,italic=True,fmt=FMT_P)

    spacer(ws,24)
    w(ws,25,2,"OTHER INCOME / EXPENSE",SECT,SECT_TXT,True,align=LFT,mc=nc)

    int_exp_act=gv(find_val(is_rows,'interest expense'),n,0)
    int_inc_act=gv(find_val(is_rows,'interest income','interest & investment income','interest and investment income'),n,0)
    eq_inc_act=gv(find_val(is_rows,'equity income','income on equity invest','equity in earnings'),n,0)
    oth_no_act=gv(find_val(is_rows,'other non-operating','other income (expense)','other non-op'),n,0)

    for row,lbl,vals,arow in [
        (26,"Interest Expense",int_exp_act,10),
        (27,"Interest & Investment Income",int_inc_act,11),
        (28,"Income on Equity Investments",eq_inc_act,12),
        (29,"Other Non-Operating Income (Expenses)",oth_no_act,13),
    ]:
        w(ws,row,2,lbl,WHITE,BKF,align=LFT)
        for i,v in enumerate(vals): w(ws,row,3+i,v,WHITE,BLF,fmt=FMT_NN)
        for i in range(m):
            w(ws,row,3+n+i,f"=Assumptions!{asmpC(i)}{arow}",WHITE,BKF,fmt=FMT_NN)

    # Row 30: EBT excl unusual
    w(ws,30,2,"EBT Excl. Unusual Items",WHITE,BKF,True,align=LFT)
    ebt_ex_act,ebt_ex_found=fv_act(is_rows,n,'pretax income excluding','ebt excl')
    for c in range(3,3+n):
        cl=get_column_letter(c)
        if ebt_ex_found:
            w(ws,30,c,ebt_ex_act[c-3],WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
        else:
            w(ws,30,c,f"={cl}22+{cl}26+{cl}27+{cl}28+{cl}29",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(m):
        w(ws,30,3+n+i,f"={ec(i)}22+{ec(i)}26+{ec(i)}27+{ec(i)}28+{ec(i)}29",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)

    spacer(ws,31)
    # Row 32: Unusual items
    unusual_act=gv(find_val(is_rows,'unusual','non-recurring','impairment','restructuring'),n,0)
    w(ws,32,2,"Unusual / Non-Recurring Items",WHITE,BKF,align=LFT)
    for i,v in enumerate(unusual_act): w(ws,32,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(m):
        w(ws,32,3+n+i,f"=Assumptions!{asmpC(i)}14",WHITE,BKF,fmt=FMT_NN)

    # Row 33: EBT incl unusual
    w(ws,33,2,"EBT Incl. Unusual Items",WHITE,BKF,True,align=LFT)
    ebt_act,ebt_found=fv_act(is_rows,n,'pretax income','pre-tax income','income before tax')
    for c in range(3,3+n):
        cl=get_column_letter(c)
        if ebt_found:
            w(ws,33,c,ebt_act[c-3],WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
        else:
            w(ws,33,c,f"={cl}30+{cl}32",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(m):
        w(ws,33,3+n+i,f"={ec(i)}30+{ec(i)}32",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)

    # Row 34: Tax (negative)
    tax_act=gv(find_val(is_rows,'income tax expense','tax provision','provision for income tax'),n,0)
    tax_act=[-abs(v) for v in tax_act]
    w(ws,34,2,"Income Tax Expense",WHITE,BKF,align=LFT)
    for i,v in enumerate(tax_act): w(ws,34,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(m):
        w(ws,34,3+n+i,f"=IF({ec(i)}33>0,-{ec(i)}33*Assumptions!{asmpC(i)}15,0)",WHITE,BKF,fmt=FMT_NN)

    # Row 35: Earnings from Cont Ops
    w(ws,35,2,"Earnings from Continuing Operations",WHITE,BKF,True,align=LFT)
    econt_act,econt_found=fv_act(is_rows,n,'earnings from continuing','income from continuing')
    for c in range(3,3+n):
        cl=get_column_letter(c)
        if econt_found:
            w(ws,35,c,econt_act[c-3],WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
        else:
            w(ws,35,c,f"={cl}33+{cl}34",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(m):
        w(ws,35,3+n+i,f"={ec(i)}33+{ec(i)}34",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)

    spacer(ws,36)
    # Row 37: Minority Interest
    mi_act=gv(find_val(is_rows,'minority interest','non-controlling interest'),n,0)
    w(ws,37,2,"Minority Interest",WHITE,BKF,align=LFT)
    for i,v in enumerate(mi_act): w(ws,37,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(m):
        w(ws,37,3+n+i,f"=Assumptions!{asmpC(i)}16",WHITE,BKF,fmt=FMT_NN)

    # Row 38: Net Income
    ni_act,ni_found=fv_act(is_rows,n,'net income','net earnings','net loss')
    w(ws,38,2,"Net Income",WHITE,BKF,True,align=LFT)
    for c in range(3,3+n):
        cl=get_column_letter(c)
        if ni_found:
            w(ws,38,c,ni_act[c-3],WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
        else:
            w(ws,38,c,f"={cl}35+{cl}37",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(m):
        w(ws,38,3+n+i,f"={ec(i)}35+{ec(i)}37",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)

    # Row 39: Net Income Margin %
    w(ws,39,2,"  % Net Income Margin",WHITE,GYF,italic=True,align=IND)
    for c in range(3,3+n+m):
        cl=get_column_letter(c)
        w(ws,39,c,f"={cl}38/{cl}7",WHITE,GYF,italic=True,fmt=FMT_P)

    spacer(ws,40)
    w(ws,41,2,"PER SHARE DATA",SECT,SECT_TXT,True,align=LFT,mc=nc)

    # Row 42: Diluted Shares
    dshares_act=gv(find_val(is_rows,'diluted shares','weighted average diluted','diluted weighted average','diluted (in shares)','(in shares)'),n,0)
    w(ws,42,2,f"Weighted Average Diluted Shares ({u})",WHITE,BKF,align=LFT)
    for i,v in enumerate(dshares_act): w(ws,42,3+i,v,WHITE,BLF,fmt=FMT_N)
    for i in range(m):
        w(ws,42,3+n+i,f"=Assumptions!{asmpC(i)}17",WHITE,BKF,fmt=FMT_N)

    # Row 43: Diluted EPS
    w(ws,43,2,"  Diluted EPS",WHITE,BKF,align=IND)
    deps_act=gv(find_val(is_rows,'diluted eps','diluted earnings per share'),n,0)
    if any(deps_act):
        for i,v in enumerate(deps_act): w(ws,43,3+i,v,WHITE,BLF,fmt=FMT_EP)
    else:
        for c in range(3,3+n):
            cl=get_column_letter(c); w(ws,43,c,f"={cl}38/{cl}42",WHITE,BKF,fmt=FMT_EP)
    for i in range(m):
        w(ws,43,3+n+i,f"={ec(i)}38/{ec(i)}42",WHITE,BKF,fmt=FMT_EP)

    # Row 44: Basic Shares
    bshares_act=gv(find_val(is_rows,'basic shares','weighted average basic','basic weighted average'),n,0)
    w(ws,44,2,f"Weighted Average Basic Shares ({u})",WHITE,BKF,align=LFT)
    for i,v in enumerate(bshares_act): w(ws,44,3+i,v,WHITE,BLF,fmt=FMT_N)
    for i in range(m):
        w(ws,44,3+n+i,f"=Assumptions!{asmpC(i)}18",WHITE,BKF,fmt=FMT_N)

    # Row 45: Basic EPS
    w(ws,45,2,"  Basic EPS",WHITE,BKF,align=IND)
    beps_act=gv(find_val(is_rows,'basic eps','basic earnings per share'),n,0)
    if any(beps_act):
        for i,v in enumerate(beps_act): w(ws,45,3+i,v,WHITE,BLF,fmt=FMT_EP)
    else:
        for c in range(3,3+n):
            cl=get_column_letter(c); w(ws,45,c,f"={cl}38/{cl}44",WHITE,BKF,fmt=FMT_EP)
    for i in range(m):
        w(ws,45,3+n+i,f"={ec(i)}38/{ec(i)}44",WHITE,BKF,fmt=FMT_EP)

    # Row 46: DPS
    dps_act=gv(find_val(is_rows,'dividends per share','dividend per share'),n,0)
    w(ws,46,2,f"  Dividends Per Share ({curr})",WHITE,BKF,align=IND)
    for i,v in enumerate(dps_act): w(ws,46,3+i,v,WHITE,BLF,fmt=FMT_EP)
    for i in range(m):
        w(ws,46,3+n+i,f"=Assumptions!{asmpC(i)}19",WHITE,BKF,fmt=FMT_EP)

    spacer(ws,47)
    w(ws,48,2,"SUPPLEMENTAL METRICS",SECT,SECT_TXT,True,align=LFT,mc=nc)

    # Row 49: R&D
    rd_act=gv(find_val(is_rows,'research and development','r&d','research & development'),n,0)
    w(ws,49,2,f"R&D Expense",WHITE,BKF,align=LFT)
    for i,v in enumerate(rd_act): w(ws,49,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(m):
        w(ws,49,3+n+i,f"=Assumptions!{asmpC(i)}20",WHITE,BKF,fmt=FMT_NN)

    # Row 50: D&A
    da_act=gv(find_val(is_rows,'depreciation and amortization','depreciation & amortization','depreciation'),n,0)
    w(ws,50,2,"Depreciation & Amortisation",WHITE,BKF,align=LFT)
    for i,v in enumerate(da_act): w(ws,50,3+i,abs(v),WHITE,BLF,fmt=FMT_NN)
    for i in range(m):
        w(ws,50,3+n+i,f"=Assumptions!{asmpC(i)}21",WHITE,BKF,fmt=FMT_NN)

    # Row 51: EBITDA
    w(ws,51,2,"EBITDA",WHITE,BKF,True,align=LFT)
    ebitda_act,ebitda_found=fv_act(is_rows,n,'ebitda')
    for c in range(3,3+n):
        cl=get_column_letter(c)
        if ebitda_found:
            w(ws,51,c,ebitda_act[c-3],WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
        else:
            w(ws,51,c,f"={cl}22+{cl}50",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(m):
        w(ws,51,3+n+i,f"={ec(i)}22+{ec(i)}50",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)

    # Row 52: EBITDA Margin %
    w(ws,52,2,"  % EBITDA Margin",WHITE,GYF,italic=True,align=IND)
    for c in range(3,3+n+m):
        cl=get_column_letter(c)
        w(ws,52,c,f"={cl}51/{cl}7",WHITE,GYF,italic=True,fmt=FMT_P)

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# BALANCE SHEET TAB
# ═══════════════════════════════════════════════════════════════════════════════
def build_bs(wb, md):
    n_act=md['n_act']; n_est=md['n_est']
    act_yrs=md['act_yrs']; est_yrs=md['est_yrs']
    bs_rows=md['bs_rows']; asmp=md['asmp']
    ticker=md['ticker']; company=md['company']; curr=md['currency']; u=md['units']
    nc=2+n_act+n_est; n=n_act; m=n_est

    ws=wb.create_sheet("BS")
    ws.sheet_properties.tabColor=GREEN_H
    ws.column_dimensions['A'].width=1; ws.column_dimensions['B'].width=35
    for i in range(n+m): ws.column_dimensions[get_column_letter(3+i)].width=13
    ws.row_dimensions[1].height=3.75
    w(ws,2,2,f"{company} ({ticker}) — Balance Sheet ({curr}{u})",NAVY,WHF,True,sz=13,align=LFT,mc=nc)
    ws.row_dimensions[2].height=22
    for i,lbl in enumerate([""]+act_yrs+est_yrs):
        is_e=isinstance(lbl,str) and lbl.strip().endswith('E')
        w(ws,3,2+i,lbl,GREEN_H if is_e else BLUE_H,WHF,True,align=CTR)
    ws.row_dimensions[3].height=18

    def ac(i): return get_column_letter(3+i)
    def ec(i): return get_column_letter(3+n+i)
    def asmpC(i): return get_column_letter(3+i)
    lac=get_column_letter(3+n-1)
    ec_all=[ec(i) for i in range(m)]
    asmpC_all=[asmpC(i) for i in range(m)]

    def brow(row, lbl, keys, bold=False, green=False, negate=False, default=0.0):
        bg=GREEN_TOT if green else WHITE
        vals,_=fv_act(bs_rows,n,*keys,default=default)
        if negate: vals=[abs(v) for v in vals]
        w(ws,row,2,lbl,bg,BKF,bold,align=LFT)
        for i,v in enumerate(vals):
            w(ws,row,3+i,v,bg,BLF,bold,fmt=FMT_NN)
        return vals

    def est_flat(row, asmp_key, bold=False, green=False, bc=None):
        bg=GREEN_TOT if green else WHITE
        for i,ac2 in enumerate(asmpC_all):
            w(ws,row,3+n+i,f"=Assumptions!{ac2}{ASMP_ROW[asmp_key]}",bg,BKF,bold,fmt=FMT_NN,bc=bc)

    # Assumptions row mapping for BS items
    ASMP_ROW={'oth_ca':27,'prepaid':28,'ppe':29,'lt_inv':30,'goodwill':31,'intang':32,
              'dta_lt':33,'oth_lta':34,'accrued':35,'st_borrow':36,'curr_ltd':37,'curr_lease':38,
              'tax_pay':39,'unearned':40,'oth_cl':41,'lt_debt':42,'lt_lease':43,'pension':44,
              'dtl_nc':45,'oth_ncl':46}

    spacer(ws,4)
    w(ws,5,2,"ASSETS",SECT,SECT_TXT,True,align=LFT,mc=nc)
    w(ws,6,2,"CURRENT ASSETS",SECT,SECT_TXT,True,align=LFT,mc=nc)

    # Row 7: Cash — estimates = green link to CFS!{col}41
    cash_act,_=fv_act(bs_rows,n,'cash and cash equiv','cash & cash equiv','cash, cash equiv')
    w(ws,7,2,"Cash & Cash Equivalents",WHITE,BKF,align=LFT)
    for i,v in enumerate(cash_act): w(ws,7,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(m): w(ws,7,3+n+i,f"=CFS!{ec(i)}41",WHITE,GRF,fmt=FMT_NN)

    # Row 8: ST Investments
    sti_act,_=fv_act(bs_rows,n,'short-term invest','marketable securities','available-for-sale')
    w(ws,8,2,"Short-Term Investments",WHITE,BKF,align=LFT)
    for i,v in enumerate(sti_act): w(ws,8,3+i,v,WHITE,BLF,fmt=FMT_NN)
    last_sti=sti_act[-1] if sti_act else 0
    for i in range(m): w(ws,8,3+n+i,last_sti,WHITE,BKF,fmt=FMT_NN)

    # Row 9: Total Cash & ST Inv
    w(ws,9,2,"Total Cash & Short-Term Investments",WHITE,BKF,True,align=LFT)
    for c in range(3,3+n):
        cl=get_column_letter(c); w(ws,9,c,f"={cl}7+{cl}8",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(m):
        w(ws,9,3+n+i,f"={ec(i)}7+{ec(i)}8",WHITE,BKF,True,fmt=FMT_NN,bc=GREEN_H)

    # Row 10: AR — DSO-driven
    ar_act,_=fv_act(bs_rows,n,'accounts receivable','trade receivable','net receivable')
    w(ws,10,2,"Accounts Receivable",WHITE,BKF,align=LFT)
    for i,v in enumerate(ar_act): w(ws,10,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(m):
        w(ws,10,3+n+i,f"=IS!{ec(i)}7*Assumptions!{asmpC(i)}24/365",WHITE,BKF,fmt=FMT_NN)

    # Row 11: Other Receivables (flat at last actual)
    oth_recv_act,_=fv_act(bs_rows,n,'other receivable','notes receivable')
    last_or=oth_recv_act[-1] if oth_recv_act else 0
    w(ws,11,2,"Other Receivables",WHITE,BKF,align=LFT)
    for i,v in enumerate(oth_recv_act): w(ws,11,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(m): w(ws,11,3+n+i,last_or,WHITE,BKF,fmt=FMT_NN)

    # Row 12: Total Receivables
    w(ws,12,2,"Total Receivables",WHITE,BKF,True,align=LFT)
    for c in range(3,3+n):
        cl=get_column_letter(c); w(ws,12,c,f"={cl}10+{cl}11",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(m):
        w(ws,12,3+n+i,f"={ec(i)}10+{ec(i)}11",WHITE,BKF,True,fmt=FMT_NN,bc=GREEN_H)

    # Row 13: Inventory — DIO-driven
    inv_act,_=fv_act(bs_rows,n,'inventor')
    w(ws,13,2,"Inventory",WHITE,BKF,align=LFT)
    for i,v in enumerate(inv_act): w(ws,13,3+i,abs(v),WHITE,BLF,fmt=FMT_NN)
    for i in range(m):
        w(ws,13,3+n+i,f"=-IS!{ec(i)}11*Assumptions!{asmpC(i)}25/365",WHITE,BKF,fmt=FMT_NN)

    # Row 14: Prepaid
    brow(14,"Prepaid Expenses",['prepaid'])
    est_flat(14,'prepaid')

    # Row 15: Other CA
    brow(15,"Other Current Assets",['other current assets'])
    est_flat(15,'oth_ca')

    # Row 16: Total CA
    w(ws,16,2,"Total Current Assets",WHITE,BKF,True,align=LFT)
    for c in range(3,3+n):
        cl=get_column_letter(c)
        w(ws,16,c,f"=SUM({cl}7:{cl}15)",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(m):
        w(ws,16,3+n+i,f"=SUM({ec(i)}7:{ec(i)}15)",WHITE,BKF,True,fmt=FMT_NN,bc=GREEN_H)

    spacer(ws,17)
    w(ws,17,2,"NON-CURRENT ASSETS",SECT,SECT_TXT,True,align=LFT,mc=nc)

    nca_items=[
        (18,"Net Property, Plant & Equipment",['net property','property plant and equipment','property, plant'],'ppe'),
        (19,"Long-Term Investments",['long-term invest','equity method invest'],'lt_inv'),
        (20,"Goodwill",['goodwill'],'goodwill'),
        (21,"Other Intangibles",['intangible','other intangible'],'intang'),
        (22,"Deferred Tax Assets — Long-Term",['deferred tax asset'],'dta_lt'),
        (23,"Other Long-Term Assets",['other long-term asset','other non-current asset'],'oth_lta'),
    ]
    for row,lbl,keys,akey in nca_items:
        brow(row,lbl,keys)
        est_flat(row,akey)

    # Row 24: Total NCA
    w(ws,24,2,"Total Non-Current Assets",WHITE,BKF,True,align=LFT)
    for c in range(3,3+n):
        cl=get_column_letter(c)
        w(ws,24,c,f"=SUM({cl}18:{cl}23)",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(m):
        w(ws,24,3+n+i,f"=SUM({ec(i)}18:{ec(i)}23)",WHITE,BKF,True,fmt=FMT_NN,bc=GREEN_H)

    spacer(ws,25)
    # Row 26: TOTAL ASSETS
    w(ws,26,2,"TOTAL ASSETS",GREEN_TOT,BKF,True,align=LFT)
    for c in range(3,3+n):
        cl=get_column_letter(c)
        w(ws,26,c,f"={cl}16+{cl}24",GREEN_TOT,BKF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(m):
        w(ws,26,3+n+i,f"={ec(i)}16+{ec(i)}24",GREEN_TOT,BKF,True,fmt=FMT_NN,bc=GREEN_H)

    spacer(ws,27)
    w(ws,28,2,"LIABILITIES & STOCKHOLDERS' EQUITY",SECT,SECT_TXT,True,align=LFT,mc=nc)
    w(ws,29,2,"CURRENT LIABILITIES",SECT,SECT_TXT,True,align=LFT,mc=nc)

    # Row 30: AP — DPO-driven
    ap_act,_=fv_act(bs_rows,n,'accounts payable','trade payable','trade and other payable')
    w(ws,30,2,"Accounts Payable",WHITE,BKF,align=LFT)
    for i,v in enumerate(ap_act): w(ws,30,3+i,abs(v),WHITE,BLF,fmt=FMT_NN)
    for i in range(m):
        w(ws,30,3+n+i,f"=-IS!{ec(i)}11*Assumptions!{asmpC(i)}26/365",WHITE,BKF,fmt=FMT_NN)

    cl_items=[
        (31,"Accrued Expenses",['accrued expense','accrued liab'],'accrued'),
        (32,"Short-Term Borrowings",['short-term borrow'],'st_borrow'),
        (33,"Current Portion of Long-Term Debt",['current portion of long-term','current maturities'],'curr_ltd'),
        (34,"Current Portion of Capital Lease Obligations",['current portion of capital','current lease'],'curr_lease'),
        (35,"Current Income Taxes Payable",['income taxes payable','taxes payable'],'tax_pay'),
        (36,"Unearned Revenue — Current",['unearned revenue','deferred revenue'],'unearned'),
        (37,"Other Current Liabilities",['other current liab'],'oth_cl'),
    ]
    for row,lbl,keys,akey in cl_items:
        brow(row,lbl,keys)
        est_flat(row,akey)

    # Row 38: Total CL
    w(ws,38,2,"Total Current Liabilities",WHITE,BKF,True,align=LFT)
    for c in range(3,3+n):
        cl=get_column_letter(c)
        w(ws,38,c,f"=SUM({cl}30:{cl}37)",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(m):
        w(ws,38,3+n+i,f"=SUM({ec(i)}30:{ec(i)}37)",WHITE,BKF,True,fmt=FMT_NN,bc=GREEN_H)

    spacer(ws,39)
    w(ws,39,2,"NON-CURRENT LIABILITIES",SECT,SECT_TXT,True,align=LFT,mc=nc)

    ncl_items=[
        (40,"Long-Term Debt",['long-term debt','long term debt','long-term borrowing'],'lt_debt'),
        (41,"Capital Leases — Long-Term",['capital lease','finance lease long','operating lease long'],'lt_lease'),
        (42,"Pension & Other Post-Retirement Benefits",['pension','post-retirement'],'pension'),
        (43,"Deferred Tax Liability — Non-Current",['deferred tax liab'],'dtl_nc'),
        (44,"Other Non-Current Liabilities",['other non-current liab','other long-term liab'],'oth_ncl'),
    ]
    for row,lbl,keys,akey in ncl_items:
        brow(row,lbl,keys)
        est_flat(row,akey)

    # Row 45: Total NCL
    w(ws,45,2,"Total Non-Current Liabilities",WHITE,BKF,True,align=LFT)
    for c in range(3,3+n):
        cl=get_column_letter(c)
        w(ws,45,c,f"=SUM({cl}40:{cl}44)",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(m):
        w(ws,45,3+n+i,f"=SUM({ec(i)}40:{ec(i)}44)",WHITE,BKF,True,fmt=FMT_NN,bc=GREEN_H)

    spacer(ws,46)
    # Row 47: Total Liabilities
    w(ws,47,2,"Total Liabilities",WHITE,BKF,True,align=LFT)
    for c in range(3,3+n):
        cl=get_column_letter(c)
        w(ws,47,c,f"={cl}38+{cl}45",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(m):
        w(ws,47,3+n+i,f"={ec(i)}38+{ec(i)}45",WHITE,BKF,True,fmt=FMT_NN,bc=GREEN_H)

    spacer(ws,48)
    w(ws,49,2,"STOCKHOLDERS' EQUITY",SECT,SECT_TXT,True,align=LFT,mc=nc)

    # Row 50: Common Stock (flat)
    cs_act,_=fv_act(bs_rows,n,'common stock','ordinary shares')
    last_cs=cs_act[-1] if cs_act else 0
    w(ws,50,2,"Common Stock / Ordinary Shares",WHITE,BKF,align=LFT)
    for i,v in enumerate(cs_act): w(ws,50,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(m): w(ws,50,3+n+i,last_cs,WHITE,BKF,fmt=FMT_NN)

    # Row 51: APIC — rolls with stock issuances
    apic_act,_=fv_act(bs_rows,n,'additional paid-in','paid-in capital','apic')
    last_apic=apic_act[-1] if apic_act else 0
    w(ws,51,2,"Additional Paid-In Capital",WHITE,BKF,align=LFT)
    for i,v in enumerate(apic_act): w(ws,51,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(m):
        prev=f"{lac}51" if i==0 else f"{ec(i-1)}51"
        w(ws,51,3+n+i,f"={prev}+Assumptions!{asmpC(i)}57",WHITE,BKF,fmt=FMT_NN)

    # Row 52: Retained Earnings — rolls with NI - Dividends
    re_act,_=fv_act(bs_rows,n,'retained earnings','accumulated deficit','accumulated earnings')
    w(ws,52,2,"Retained Earnings / Accumulated Deficit",WHITE,BKF,align=LFT)
    for i,v in enumerate(re_act): w(ws,52,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(m):
        prev=f"{lac}52" if i==0 else f"{ec(i-1)}52"
        ni_ref=f"IS!{ec(i)}38"
        div_ref=f"IS!{ec(i)}44*Assumptions!{asmpC(i)}19"
        w(ws,52,3+n+i,f"={prev}+{ni_ref}-{div_ref}",WHITE,BKF,fmt=FMT_NN)

    # Row 53: Treasury Stock (flat at last actual)
    ts_act,_=fv_act(bs_rows,n,'treasury stock','treasury shares')
    last_ts=ts_act[-1] if ts_act else 0
    w(ws,53,2,"Treasury Stock",WHITE,BKF,align=LFT)
    for i,v in enumerate(ts_act): w(ws,53,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(m): w(ws,53,3+n+i,last_ts,WHITE,BKF,fmt=FMT_NN)

    # Row 54: AOCI (flat at last actual)
    aoci_act,_=fv_act(bs_rows,n,'comprehensive income & other','accumulated other comprehensive','aoci')
    last_aoci=aoci_act[-1] if aoci_act else 0
    w(ws,54,2,"Accumulated Other Comprehensive Income (AOCI)",WHITE,BKF,align=LFT)
    for i,v in enumerate(aoci_act): w(ws,54,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(m): w(ws,54,3+n+i,last_aoci,WHITE,BKF,fmt=FMT_NN)

    # Row 55: Total Common Equity
    w(ws,55,2,"Total Common Equity",WHITE,BKF,True,align=LFT)
    for c in range(3,3+n):
        cl=get_column_letter(c)
        w(ws,55,c,f"=SUM({cl}50:{cl}54)",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(m):
        w(ws,55,3+n+i,f"=SUM({ec(i)}50:{ec(i)}54)",WHITE,BKF,True,fmt=FMT_NN,bc=GREEN_H)

    # Row 56: Minority Interest in BS (flat)
    mi_bs_act,_=fv_act(bs_rows,n,'minority interest','non-controlling interest')
    last_mi=mi_bs_act[-1] if mi_bs_act else 0
    w(ws,56,2,"Minority Interest",WHITE,BKF,align=LFT)
    for i,v in enumerate(mi_bs_act): w(ws,56,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(m): w(ws,56,3+n+i,last_mi,WHITE,BKF,fmt=FMT_NN)

    # Row 57: Total Equity
    w(ws,57,2,"Total Equity",WHITE,BKF,True,align=LFT)
    for c in range(3,3+n):
        cl=get_column_letter(c)
        w(ws,57,c,f"={cl}55+{cl}56",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(m):
        w(ws,57,3+n+i,f"={ec(i)}55+{ec(i)}56",WHITE,BKF,True,fmt=FMT_NN,bc=GREEN_H)

    spacer(ws,58)
    # Row 59: TOTAL LIABILITIES & EQUITY
    w(ws,59,2,"TOTAL LIABILITIES & EQUITY",GREEN_TOT,BKF,True,align=LFT)
    for c in range(3,3+n):
        cl=get_column_letter(c)
        w(ws,59,c,f"={cl}47+{cl}57",GREEN_TOT,BKF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(m):
        w(ws,59,3+n+i,f"={ec(i)}47+{ec(i)}57",GREEN_TOT,BKF,True,fmt=FMT_NN,bc=GREEN_H)

    # Row 60: Balance Check
    w(ws,60,2,"  Balance Check (Assets − L&E)",WHITE,GYF,italic=True,align=IND)
    for c in range(3,3+n+m):
        cl=get_column_letter(c)
        w(ws,60,c,f"={cl}26-{cl}59",WHITE,GYF,italic=True,fmt=FMT_NN)

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# CASH FLOW STATEMENT TAB
# ═══════════════════════════════════════════════════════════════════════════════
def build_cfs(wb, md):
    n_act=md['n_act']; n_est=md['n_est']
    act_yrs=md['act_yrs']; est_yrs=md['est_yrs']
    cfs_rows=md['cfs_rows']; asmp=md['asmp']
    ticker=md['ticker']; company=md['company']; curr=md['currency']; u=md['units']
    nc=2+n_act+n_est; n=n_act; m=n_est

    ws=wb.create_sheet("CFS")
    ws.sheet_properties.tabColor=BROWN
    ws.column_dimensions['A'].width=1; ws.column_dimensions['B'].width=35
    for i in range(n+m): ws.column_dimensions[get_column_letter(3+i)].width=13
    ws.row_dimensions[1].height=3.75
    w(ws,2,2,f"{company} ({ticker}) — Cash Flow Statement ({curr}{u})",NAVY,WHF,True,sz=13,align=LFT,mc=nc)
    ws.row_dimensions[2].height=22
    for i,lbl in enumerate([""]+act_yrs+est_yrs):
        is_e=isinstance(lbl,str) and lbl.strip().endswith('E')
        w(ws,3,2+i,lbl,GREEN_H if is_e else BLUE_H,WHF,True,align=CTR)
    ws.row_dimensions[3].height=18

    def ac(i): return get_column_letter(3+i)
    def ec(i): return get_column_letter(3+n+i)
    def asmpC(i): return get_column_letter(3+i)
    lac=get_column_letter(3+n-1)
    ec_all=[ec(i) for i in range(m)]
    asmpC_all=[asmpC(i) for i in range(m)]

    def crow(row, lbl, keys, est_formula=None, negate=False, indent=False, bold=False, bc=None):
        alg=IND if indent else LFT
        vals=gv(find_val(cfs_rows,*keys) if keys else None,n,0)
        if negate: vals=[-abs(v) for v in vals]
        w(ws,row,2,lbl,WHITE,BKF,bold,align=alg)
        for i,v in enumerate(vals): w(ws,row,3+i,v,WHITE,BLF,bold,fmt=FMT_NN,bc=bc)
        if est_formula:
            for i in range(m):
                formula=est_formula(ec(i),asmpC(i),i)
                fc=GRF if formula.startswith('=IS!') or formula.startswith('=-IS!') else BKF
                w(ws,row,3+n+i,formula,WHITE,fc,bold,fmt=FMT_NN,bc=bc)

    spacer(ws,4)
    w(ws,5,2,"OPERATING ACTIVITIES",SECT,SECT_TXT,True,align=LFT,mc=nc)

    # Row 6: Net Income — green link from IS
    crow(6,"Net Income",['net income','net earnings','net loss'],
         lambda ec2,ac2,i: f"=IS!{ec2}38",bc=None)
    # Override font for estimates to green
    for i in range(m): ws.cell(row=6,column=3+n+i).font=fn(GRF)

    # Row 7: D&A
    crow(7,"  Depreciation & Amortisation",['depreciation and amortization','depreciation & amortization','depreciation'],
         lambda ec2,ac2,i: f"=IS!{ec2}50",indent=True)
    for i in range(m): ws.cell(row=7,column=3+n+i).font=fn(GRF)

    # Row 8: Amort of GI
    crow(8,"  Amortisation of Intangibles",['amortization of intangible','amortization of goodwill'],
         lambda ec2,ac2,i: "=0",indent=True)

    # Row 9: Amort Deferred
    crow(9,"  Amortisation of Deferred Charges",['amortization of deferred'],
         lambda ec2,ac2,i: "=0",indent=True)

    # Row 10: Asset Writedown
    crow(10,"  Asset Writedown & Restructuring",['asset writedown','restructuring charges','impairment'],
         lambda ec2,ac2,i: "=0",indent=True)

    # Row 11: Equity Income reversal
    crow(11,"  (Income) Loss on Equity Investments",['income on equity invest','equity income','equity in earnings'],
         lambda ec2,ac2,i: f"=-IS!{ec2}28",indent=True)
    for i in range(m): ws.cell(row=11,column=3+n+i).font=fn(GRF)
    # Negate actuals for equity income reversal
    eq_rev_act=gv(find_val(cfs_rows,'income on equity invest','equity income','equity in earnings'),n,0)
    for i,v in enumerate(eq_rev_act): w(ws,11,3+i,v,WHITE,BLF,fmt=FMT_NN)

    # Row 12: Other Operating
    crow(12,"  Other Operating Activities",['other operating activities','other operating cash'],
         lambda ec2,ac2,i: "=0",indent=True)

    # Rows 13-15: Working capital changes (BS-derived for estimates)
    crow(13,"  Change in Accounts Receivable",['change in accounts receivable','increase (decrease) in receivables'],
         lambda ec2,ac2,i: f"=-(BS!{ec2}10-BS!{(lac if i==0 else ec(i-1))}10)",indent=True)
    crow(14,"  Change in Inventories",['change in inventor','decrease (increase) in inventor'],
         lambda ec2,ac2,i: f"=-(BS!{ec2}13-BS!{(lac if i==0 else ec(i-1))}13)",indent=True)
    crow(15,"  Change in Accounts Payable",['change in accounts payable','increase (decrease) in payable'],
         lambda ec2,ac2,i: f"=BS!{ec2}30-BS!{(lac if i==0 else ec(i-1))}30",indent=True)

    # Row 16: CFO
    w(ws,16,2,"Net Cash from Operating Activities",WHITE,BKF,True,align=LFT)
    cfo_act=gv(find_val(cfs_rows,'net cash from operating','net cash provided by operating','operating activities'),n,0)
    for i,v in enumerate(cfo_act): w(ws,16,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(m):
        w(ws,16,3+n+i,f"=SUM({ec(i)}6:{ec(i)}15)",WHITE,BKF,True,fmt=FMT_NN,bc=BROWN)

    spacer(ws,17)
    w(ws,18,2,"INVESTING ACTIVITIES",SECT,SECT_TXT,True,align=LFT,mc=nc)

    # Row 19: Capex
    capex_act=gv(find_val(cfs_rows,'capital expenditure','capital expenditures','purchase of property',
                          'purchases of property','purchase of pp&e'),n,0)
    capex_act=[-abs(v) for v in capex_act]
    w(ws,19,2,"Capital Expenditure",WHITE,BKF,align=LFT)
    for i,v in enumerate(capex_act): w(ws,19,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(m):
        w(ws,19,3+n+i,f"=-IS!{ec(i)}7*Assumptions!{asmpC(i)}49",WHITE,BKF,fmt=FMT_NN)

    inv_rows=[
        (20,"Sale of Property, Plant & Equipment",['sale of property','proceeds from sale of property'],'pp_sale',50),
        (21,"Cash Acquisitions",['cash acquisition','acquisition','business acquisition'],'acquisitions',51),
        (22,"Divestitures",['divestiture','proceeds from divestiture','disposal of business'],'divestitures',52),
        (23,"Investment in Marketable & Equity Securities",['investment in securities','other investing'],'oth_inv',53),
        (24,"Net (Increase) Decrease in Loans",['net loans','net increase decrease in loans'],'',54),
    ]
    for row,lbl,keys,akey,arow in inv_rows:
        vals=gv(find_val(cfs_rows,*keys) if keys else None,n,0)
        w(ws,row,2,lbl,WHITE,BKF,align=LFT)
        for i,v in enumerate(vals): w(ws,row,3+i,v,WHITE,BLF,fmt=FMT_NN)
        for i in range(m):
            w(ws,row,3+n+i,f"=Assumptions!{asmpC(i)}{arow}",WHITE,BKF,fmt=FMT_NN)

    # Row 25: CFI
    w(ws,25,2,"Net Cash from Investing Activities",WHITE,BKF,True,align=LFT)
    cfi_act=gv(find_val(cfs_rows,'net cash from investing','net cash used in investing','investing activities'),n,0)
    for i,v in enumerate(cfi_act): w(ws,25,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(m):
        w(ws,25,3+n+i,f"=SUM({ec(i)}19:{ec(i)}24)",WHITE,BKF,True,fmt=FMT_NN,bc=BROWN)

    spacer(ws,26)
    w(ws,27,2,"FINANCING ACTIVITIES",SECT,SECT_TXT,True,align=LFT,mc=nc)

    fin_rows=[
        (28,"Total Debt Issued",['total debt issued','proceeds from debt','issuance of long-term debt',
                                 'proceeds from issuance of long-term debt'],'',55),
        (29,"Total Debt Repaid",['total debt repaid','repayment of debt','repayment of long-term debt'],'',56),
        (30,"Issuance of Common Stock",['issuance of common stock','proceeds from issuance of common'],'',57),
        (31,"Repurchase of Common Stock",['repurchase of common stock','purchase of common stock','buyback'],'',58),
    ]
    for row,lbl,keys,_,arow in fin_rows:
        vals=gv(find_val(cfs_rows,*keys) if keys else None,n,0)
        w(ws,row,2,lbl,WHITE,BKF,align=LFT)
        for i,v in enumerate(vals): w(ws,row,3+i,v,WHITE,BLF,fmt=FMT_NN)
        for i in range(m):
            w(ws,row,3+n+i,f"=Assumptions!{asmpC(i)}{arow}",WHITE,BKF,fmt=FMT_NN)

    # Row 32: Dividends (derived from IS shares × DPS)
    div_act=gv(find_val(cfs_rows,'common dividends','dividends paid'),n,0)
    div_act=[-abs(v) for v in div_act]
    w(ws,32,2,"Common Dividends Paid",WHITE,BKF,align=LFT)
    for i,v in enumerate(div_act): w(ws,32,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(m):
        w(ws,32,3+n+i,f"=-IS!{ec(i)}44*Assumptions!{asmpC(i)}19",WHITE,BKF,fmt=FMT_NN)

    # Row 33: Other Financing
    oth_fin_act=gv(find_val(cfs_rows,'other financing activities','other financing'),n,0)
    w(ws,33,2,"Other Financing Activities",WHITE,BKF,align=LFT)
    for i,v in enumerate(oth_fin_act): w(ws,33,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(m):
        w(ws,33,3+n+i,f"=Assumptions!{asmpC(i)}59",WHITE,BKF,fmt=FMT_NN)

    # Row 34: CFF
    w(ws,34,2,"Net Cash from Financing Activities",WHITE,BKF,True,align=LFT)
    cff_act=gv(find_val(cfs_rows,'net cash from financing','net cash used in financing','financing activities'),n,0)
    for i,v in enumerate(cff_act): w(ws,34,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(m):
        w(ws,34,3+n+i,f"=SUM({ec(i)}28:{ec(i)}33)",WHITE,BKF,True,fmt=FMT_NN,bc=BROWN)

    spacer(ws,35)
    # Row 36: FX
    fx_act=gv(find_val(cfs_rows,'effect of','fx adjustment','foreign exchange effect'),n,0)
    w(ws,36,2,"Effect of FX on Cash",WHITE,BKF,align=LFT)
    for i,v in enumerate(fx_act): w(ws,36,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(m):
        w(ws,36,3+n+i,f"=Assumptions!{asmpC(i)}60",WHITE,BKF,fmt=FMT_NN)

    # Row 37: Misc
    misc_act=gv(find_val(cfs_rows,'miscellaneous','other cash flow adjustment'),n,0)
    w(ws,37,2,"Miscellaneous Cash Flow Adjustments",WHITE,BKF,align=LFT)
    for i,v in enumerate(misc_act): w(ws,37,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(m):
        w(ws,37,3+n+i,f"=Assumptions!{asmpC(i)}61",WHITE,BKF,fmt=FMT_NN)

    # Row 38: Net Change in Cash
    w(ws,38,2,"Net Change in Cash",WHITE,BKF,True,align=LFT)
    net_chg_act=gv(find_val(cfs_rows,'net change in cash','net increase decrease in cash'),n,0)
    for i,v in enumerate(net_chg_act): w(ws,38,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(m):
        w(ws,38,3+n+i,f"={ec(i)}16+{ec(i)}25+{ec(i)}34+{ec(i)}36+{ec(i)}37",WHITE,BKF,True,fmt=FMT_NN,bc=BROWN)

    # Row 39: Beginning Cash
    w(ws,39,2,"Beginning Cash",WHITE,BKF,align=LFT)
    beg_act=gv(find_val(cfs_rows,'beginning cash','cash at beginning','beginning of period'),n,0)
    for i,v in enumerate(beg_act): w(ws,39,3+i,v,WHITE,BLF,fmt=FMT_NN)
    # FY+1E beginning = last actual BS cash
    w(ws,39,3+n,f"=BS!{lac}7",WHITE,GRF,fmt=FMT_NN)
    for i in range(1,m):
        w(ws,39,3+n+i,f"={ec(i-1)}41",WHITE,BKF,fmt=FMT_NN)

    spacer(ws,40,h=3)

    # Row 41: Ending Cash — BS!{col}7 links here for estimates
    w(ws,41,2,"Ending Cash",WHITE,BKF,True,align=LFT)
    end_act=gv(find_val(cfs_rows,'ending cash','cash at end','end of period cash'),n,0)
    for i,v in enumerate(end_act): w(ws,41,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(m):
        w(ws,41,3+n+i,f"={ec(i)}39+{ec(i)}38",WHITE,BKF,True,fmt=FMT_NN,bc=BROWN)

    spacer(ws,42)
    w(ws,43,2,"SUPPLEMENTAL",SECT,SECT_TXT,True,align=LFT,mc=nc)

    # Row 44: FCF = CFO + Capex
    w(ws,44,2,"  Free Cash Flow (CFO + Capex)",WHITE,GYF,italic=True,align=IND)
    fcf_act=gv(find_val(cfs_rows,'free cash flow'),n,0)
    for i,v in enumerate(fcf_act if any(fcf_act) else [a+b for a,b in zip(cfo_act,capex_act)]):
        w(ws,44,3+i,v,WHITE,BLF,italic=True,fmt=FMT_NN)
    for i in range(m):
        w(ws,44,3+n+i,f"={ec(i)}16+{ec(i)}19",WHITE,GYF,italic=True,fmt=FMT_NN)

    # Row 45: FCF Margin
    w(ws,45,2,"  % FCF Margin",WHITE,GYF,italic=True,align=IND)
    for c in range(3,3+n+m):
        cl=get_column_letter(c)
        w(ws,45,c,f"={cl}44/IS!{cl}7",WHITE,GYF,italic=True,fmt=FMT_P)

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# DCF TAB
# ═══════════════════════════════════════════════════════════════════════════════
def build_dcf(wb, md):
    n_act=md['n_act']; n_est=md['n_est']
    est_yrs=md['est_yrs']
    ticker=md['ticker']; company=md['company']; curr=md['currency']; u=md['units']
    dcf_asmp=md['dcf']; n=n_act; m=n_est

    ws=wb.create_sheet("DCF")
    ws.sheet_properties.tabColor=RED_BG
    ws.column_dimensions['A'].width=1
    ws.column_dimensions['B'].width=38
    ws.column_dimensions['C'].width=18
    ws.column_dimensions['D'].width=14
    for i in range(m+1): ws.column_dimensions[get_column_letter(5+i)].width=13
    ws.column_dimensions[get_column_letter(6+m)].width=14
    ws.row_dimensions[1].height=3.75
    w(ws,2,2,f"{company} ({ticker}) — DCF Valuation ({curr}{u})",NAVY,WHF,True,sz=13,align=LFT,mc=7+m)
    ws.row_dimensions[2].height=22

    # IS/CFS estimate column letters (same as IS/BS/CFS tabs)
    ec_cols=[get_column_letter(3+n+i) for i in range(m)]
    lac=get_column_letter(3+n-1)                     # last actual col on IS/CFS
    asmpC_all=[get_column_letter(3+i) for i in range(m)]  # assumptions cols
    term_col=get_column_letter(5+m)                  # terminal column on DCF

    # ── SECTION 1: DCF ASSUMPTIONS ──────────────────────────────────────────
    spacer(ws,3)
    w(ws,4,2,"DCF ASSUMPTIONS",SECT,SECT_TXT,True,align=LFT,mc=7+m)
    w(ws,5,2,"Assumption",BLUE_H,WHF,True,align=LFT)
    w(ws,5,4,"Value",BLUE_H,WHF,True,align=CTR)
    w(ws,5,5,"Notes",BLUE_H,WHF,True,align=LFT,mc=5+m)

    dcf_rows=[
        (6, "WACC",                          dcf_asmp['wacc'],   FMT_P,  "Blended cost of capital — adjust to company/sector"),
        (7, "Terminal Growth Rate (TGR)",     dcf_asmp['tgr'],    FMT_P,  "Long-run nominal GDP / sector growth"),
        (8, "Projection Period (years)",      m,                  FMT_N,  f"Based on {m} estimate years in model"),
        (9, f"Cash & Equivalents ({curr}{u})",dcf_asmp['cash'],  FMT_NN, "Last actual ending cash"),
        (10,f"Short-Term Investments ({curr}{u})",dcf_asmp['st_inv'],FMT_NN,"Last actual ST investments"),
        (11,f"Total Debt ({curr}{u})",        dcf_asmp['total_debt'],FMT_NN,"Last actual total debt"),
        (12,f"Net Cash ({curr}{u})",          "=(D9+D10-D11)",    FMT_NN, "Computed: Cash + STI − Total Debt"),
        (13,f"Diluted Shares Outstanding ({u})",dcf_asmp['dil_shares'],FMT_N,"Last actual diluted shares"),
        (14,f"Current Share Price ({curr})",  dcf_asmp['share_price'],FMT_EP,"Enter current market price"),
        (15,"NTM EV/EBITDA (comps ref)",      dcf_asmp['ntm_ev_ebitda'],FMT_MX,"Peer sector median — adjust to comps"),
    ]
    for r,lbl,val,fmt,note in dcf_rows:
        w(ws,r,2,lbl,WHITE,BKF,align=LFT)
        fc=BKF if isinstance(val,str) and str(val).startswith('=') else BLF
        w(ws,r,4,val,WHITE,fc,fmt=fmt,align=CTR)
        w(ws,r,5,note,WHITE,GYF,italic=True,align=LFT,mc=5+m)

    # ── SECTION 2: FCF PROJECTIONS ──────────────────────────────────────────
    spacer(ws,15)
    # Headers row 16
    yr_hdrs=["","Base (Actual)"]+[f"Yr {i+1}" for i in range(m)]+["Terminal"]
    for i,lbl in enumerate(yr_hdrs):
        is_e=i>1
        w(ws,16,3+i,lbl,GREEN_H if is_e else BLUE_H,WHF,True,align=CTR)

    fcf_rows=[
        (17,"Revenue",            f"IS!{{c}}7",   FMT_NN,False),
        (18,"  % Revenue Growth", None,           FMT_P, True),
        (19,"EBIT",               f"IS!{{c}}22",  FMT_NN,False),
        (20,"  % EBIT Margin",    None,           FMT_P, True),
        (21,"Tax on EBIT",        None,           FMT_NN,False),
        (22,"NOPAT",              None,           FMT_NN,False),
        (23,"(+) D&A",            f"IS!{{c}}50",  FMT_NN,False),
        (24,"(+) SBC",            "=0",           FMT_NN,False),
        (25,"(−) Capex",          f"CFS!{{c}}19", FMT_NN,False),
        (26,"(+/−) Change in WC", None,           FMT_NN,False),
        (27,"Unlevered FCF",      None,           FMT_NN,True),
        (28,"  % uFCF Margin",    None,           FMT_P, True),
    ]

    tax_base_str=f"IF(IS!{lac}22>0,-IS!{lac}22*D6,0)"
    for r,lbl,ref_tmpl,fmt,bold in fcf_rows:
        is_grey=lbl.startswith("  ") and "%" in lbl
        fc=GYF if is_grey else BKF
        w(ws,r,2,lbl,WHITE,fc,bold,italic=is_grey,align=IND if is_grey else LFT)
        # Base year (col D = 4)
        if ref_tmpl and "{c}" in ref_tmpl:
            w(ws,r,4,f"={ref_tmpl.format(c=lac)}",WHITE,GRF,bold,fmt=fmt,align=CTR)
        elif r==18: w(ws,r,4,"",WHITE,GYF,italic=True,fmt=fmt,align=CTR)
        elif r==20: w(ws,r,4,f"=D19/D17",WHITE,GYF,italic=True,fmt=fmt,align=CTR)
        elif r==21: w(ws,r,4,f"=IF(IS!{lac}22>0,-IS!{lac}22*D6,0)",WHITE,BKF,fmt=fmt,align=CTR)
        elif r==22: w(ws,r,4,"=D19+D21",WHITE,BKF,fmt=fmt,align=CTR)
        elif r==24: w(ws,r,4,"=0",WHITE,BKF,fmt=fmt,align=CTR)
        elif r==26: w(ws,r,4,"=0",WHITE,BKF,fmt=fmt,align=CTR)
        elif r==27: w(ws,r,4,"=D22+D23+D24+D25+D26",WHITE,BKF,True,fmt=fmt,align=CTR,bc=RED_BG)
        elif r==28: w(ws,r,4,"=D27/D17",WHITE,GYF,italic=True,fmt=fmt,align=CTR)
        # Estimate years (cols E..E+m-1 = col 5..5+m-1)
        for i,ec2 in enumerate(ec_cols):
            col_idx=5+i; cl=get_column_letter(col_idx)
            prev_cl=get_column_letter(col_idx-1)
            if ref_tmpl and "{c}" in ref_tmpl:
                w(ws,r,col_idx,f"={ref_tmpl.format(c=ec2)}",WHITE,GRF,bold,fmt=fmt,align=CTR,bc=RED_BG if r==27 else None)
            elif r==18:
                prev_rev="D17" if i==0 else f"{get_column_letter(col_idx-1)}17"
                w(ws,r,col_idx,f"={cl}17/{prev_rev}-1",WHITE,GYF,italic=True,fmt=fmt,align=CTR)
            elif r==20:
                w(ws,r,col_idx,f"={cl}19/{cl}17",WHITE,GYF,italic=True,fmt=fmt,align=CTR)
            elif r==21:
                w(ws,r,col_idx,f"=IF(IS!{ec2}22>0,-IS!{ec2}22*Assumptions!{asmpC_all[i]}15,0)",WHITE,BKF,fmt=fmt,align=CTR)
            elif r==22:
                w(ws,r,col_idx,f"={cl}19+{cl}21",WHITE,BKF,fmt=fmt,align=CTR)
            elif r==24:
                w(ws,r,col_idx,"=0",WHITE,BKF,fmt=fmt,align=CTR)
            elif r==26:
                w(ws,r,col_idx,f"=-(CFS!{ec2}13+CFS!{ec2}14+CFS!{ec2}15)",WHITE,BKF,fmt=fmt,align=CTR)
            elif r==27:
                w(ws,r,col_idx,f"=SUM({cl}22:{cl}26)",WHITE,BKF,True,fmt=fmt,align=CTR,bc=RED_BG)
            elif r==28:
                w(ws,r,col_idx,f"={cl}27/{cl}17",WHITE,GYF,italic=True,fmt=fmt,align=CTR)
        # Terminal column
        tc_col=5+m; tc=get_column_letter(tc_col); last_est=get_column_letter(tc_col-1)
        if r==17:   w(ws,r,tc_col,f"={last_est}17*(1+D7)",WHITE,BKF,fmt=fmt,align=CTR)
        elif r==18: w(ws,r,tc_col,f"=D7",WHITE,GYF,italic=True,fmt=FMT_P,align=CTR)
        elif r==19: w(ws,r,tc_col,f"={last_est}19*(1+D7)",WHITE,BKF,fmt=fmt,align=CTR)
        elif r==20: w(ws,r,tc_col,f"={tc}19/{tc}17",WHITE,GYF,italic=True,fmt=FMT_P,align=CTR)
        elif r==21: w(ws,r,tc_col,f"=IF({tc}19>0,-{tc}19*D6,0)",WHITE,BKF,fmt=fmt,align=CTR)
        elif r==22: w(ws,r,tc_col,f"={tc}19+{tc}21",WHITE,BKF,fmt=fmt,align=CTR)
        elif r==23: w(ws,r,tc_col,f"={last_est}23*(1+D7)",WHITE,GRF,fmt=fmt,align=CTR)
        elif r==24: w(ws,r,tc_col,"=0",WHITE,BKF,fmt=fmt,align=CTR)
        elif r==25: w(ws,r,tc_col,f"={last_est}25*(1+D7)",WHITE,GRF,fmt=fmt,align=CTR)
        elif r==26: w(ws,r,tc_col,"=0",WHITE,BKF,fmt=fmt,align=CTR)
        elif r==27: w(ws,r,tc_col,f"={tc}22+{tc}23+{tc}24+{tc}25+{tc}26",WHITE,BKF,True,fmt=fmt,align=CTR,bc=RED_BG)
        elif r==28: w(ws,r,tc_col,f"={tc}27/{tc}17",WHITE,GYF,italic=True,fmt=FMT_P,align=CTR)

    # ── SECTION 3: DCF CALCULATION ──────────────────────────────────────────
    spacer(ws,29)
    w(ws,30,2,"DCF CALCULATION",SECT,SECT_TXT,True,align=LFT,mc=7+m)

    # Column sub-headers on their own row
    w(ws,31,4,"Base",BLUE_H,WHF,True,align=CTR)
    for i in range(m): w(ws,31,5+i,f"Yr {i+1}",BLUE_H,WHF,True,align=CTR)
    w(ws,31,5+m,"Terminal",BLUE_H,WHF,True,align=CTR)

    w(ws,32,2,"Period",WHITE,GYF,italic=True,align=LFT)
    for i in range(m): w(ws,32,5+i,i+1,WHITE,GYF,italic=True,fmt=FMT_N,align=CTR)
    w(ws,32,5+m,"∞",WHITE,GYF,italic=True,align=CTR)

    w(ws,33,2,"Discount Factor",WHITE,GYF,italic=True,align=LFT)
    for i in range(m): w(ws,33,5+i,f"=1/(1+DCF!D6)^{i+1}",WHITE,GYF,italic=True,fmt="0.0000",align=CTR)
    w(ws,33,5+m,f"=1/(1+DCF!D6)^{m}",WHITE,GYF,italic=True,fmt="0.0000",align=CTR)

    w(ws,34,2,"Unlevered FCF",WHITE,BKF,align=LFT)
    for i in range(m): w(ws,34,5+i,f"={get_column_letter(5+i)}27",WHITE,GRF,fmt=FMT_NN,align=CTR)

    w(ws,35,2,"Terminal Value (Gordon Growth)",WHITE,BKF,align=LFT)
    w(ws,35,5+m,f"={tc}27*(1+D7)/(D6-D7)",WHITE,BKF,fmt=FMT_NN,align=CTR)

    w(ws,36,2,"PV of FCF",WHITE,BKF,align=LFT)
    for i in range(m):
        w(ws,36,5+i,f"={get_column_letter(5+i)}34*{get_column_letter(5+i)}33",WHITE,BKF,fmt=FMT_NN,align=CTR,bc=RED_BG)
    w(ws,36,5+m,f"={tc}35*{tc}33",WHITE,BKF,fmt=FMT_NN,align=CTR,bc=RED_BG)

    # ── SECTION 4: VALUATION BRIDGE ─────────────────────────────────────────
    spacer(ws,37)
    w(ws,38,2,"VALUATION BRIDGE",SECT,SECT_TXT,True,align=LFT,mc=7+m)
    sum_pv="+".join([f"{get_column_letter(5+i)}36" for i in range(m)])
    bridge=[
        (39,f"Sum of PV of FCFs ({curr}{u})",f"={sum_pv}",FMT_NN,BKF,WHITE),
        (40,f"(+) PV of Terminal Value ({curr}{u})",f"={tc}36",FMT_NN,GRF,WHITE),
        (41,f"Enterprise Value ({curr}{u})",f"=D39+D40",FMT_NN,BKF,GREEN_TOT),
        (42,f"Enterprise Value ({curr}B)",f"=D41/1000","#,##0.0",BKF,WHITE),
        (43,"  % from Terminal Value",f"=D40/D41",FMT_P,GYF,WHITE),
        (44,f"(+) Net Cash ({curr}{u})",f"=D12",FMT_NN,GRF,WHITE),
        (45,f"Equity Value ({curr}{u})",f"=D41+D44",FMT_NN,BKF,GREEN_TOT),
        (46,f"(÷) Diluted Shares ({u})",f"=D13",FMT_N,GRF,WHITE),
        (47,f"Intrinsic Value Per Share ({curr})",f"=D45/D46",FMT_EP,WHF,RED_BG),
        (48,f"Current Share Price ({curr})",f"=D14",FMT_EP,GRF,WHITE),
        (49,"  Upside / (Downside)",f"=D47/D48-1",FMT_P,GYF,WHITE),
    ]
    for r,lbl,formula,fmt,fc,bg in bridge:
        bold=(r in [41,45,47]); italic=(fc==GYF)
        w(ws,r,2,lbl,bg,fc if r!=47 else WHF,bold,italic,align=LFT)
        w(ws,r,4,formula,bg,fc if r!=47 else WHF,bold,italic,fmt=fmt,align=CTR)

    spacer(ws,50)
    ntm_ebitda_col=ec_cols[0] if ec_cols else lac  # NTM = first estimate year
    cchk=[
        (51,"  TV as % of EV",f"=D40/D41",FMT_P),
        (52,f"  Implied EV / NTM EBITDA",f"=D41/IS!{ntm_ebitda_col}51",FMT_MX),
        (53,f"  Implied EV / NTM Revenue",f"=D41/IS!{ntm_ebitda_col}7",FMT_MX),
        (54,f"  NTM EBITDA ({curr}{u})",f"=IS!{ntm_ebitda_col}51",FMT_NN),
    ]
    for r,lbl,formula,fmt in cchk:
        w(ws,r,2,lbl,WHITE,GYF,italic=True,align=IND)
        w(ws,r,4,formula,WHITE,GYF,italic=True,fmt=fmt,align=CTR)

    # ── SECTION 5: SENSITIVITY ANALYSIS ─────────────────────────────────────
    spacer(ws,54)
    w(ws,55,2,"SENSITIVITY ANALYSIS — IMPLIED SHARE PRICE",SECT,SECT_TXT,True,align=LFT,mc=7+m)

    # Table 1: WACC × TGR
    tgr_vals=[0.015,0.020,0.025,0.030,0.035,0.040]
    wacc_vals=[0.075,0.080,0.090,0.100,0.105]
    base_wacc=dcf_asmp['wacc']; base_tgr=dcf_asmp['tgr']

    w(ws,56,3,f"WACC \\ TGR",NAVY,WHF,True,align=CTR)
    for j,tgr in enumerate(tgr_vals):
        cell=ws.cell(row=56,column=4+j)
        cell.value=tgr; cell.number_format=FMT_P
        cell.font=fn(WHF,True); cell.fill=fl(NAVY); cell.alignment=CTR

    fcf_pv_tmpl="+".join([f"{get_column_letter(5+i)}27/((1+{{wacc}})^{i+1})" for i in range(m)])
    tv_tmpl=f"{tc}27*(1+{{tgr}})/({{wacc}}-{{tgr}})/(1+{{wacc}})^{m}"

    for i,wacc in enumerate(wacc_vals):
        row=57+i
        hdr=ws.cell(row=row,column=3)
        hdr.value=wacc; hdr.number_format=FMT_P
        is_base_row=abs(wacc-base_wacc)<0.001
        hdr.font=fn(WHF,True); hdr.fill=fl(RED_BG if is_base_row else NAVY); hdr.alignment=CTR
        for j,tgr in enumerate(tgr_vals):
            col=4+j
            is_base=abs(wacc-base_wacc)<0.001 and abs(tgr-base_tgr)<0.001
            fcf_pv=fcf_pv_tmpl.format(wacc=wacc)
            tv=tv_tmpl.format(wacc=wacc,tgr=tgr)
            eq_val=f"=({fcf_pv}+{tv}+D12)/D13"
            cell=ws.cell(row=row,column=col)
            cell.value=eq_val; cell.number_format=FMT_EP
            cell.font=fn(BKF,is_base); cell.fill=fl(YELLOW if is_base else WHITE); cell.alignment=CTR
    w(ws,62,3,f"Base case: WACC={base_wacc:.1%}, TGR={base_tgr:.1%}",WHITE,GYF,italic=True,align=LFT,mc=9)
    ws.row_dimensions[62].height=12

    # Table 2: WACC × EV/EBITDA exit multiple
    mult_vals=[8.0,10.0,12.0,14.0,16.0,18.0]
    base_mult=dcf_asmp['ntm_ev_ebitda']
    last_est_ebitda_col=ec_cols[-1] if ec_cols else lac

    spacer(ws,63)
    w(ws,64,3,f"WACC \\ EV/EBITDA",NAVY,WHF,True,align=CTR)
    for j,mult in enumerate(mult_vals):
        cell=ws.cell(row=64,column=4+j)
        cell.value=mult; cell.number_format=FMT_MX
        cell.font=fn(WHF,True); cell.fill=fl(NAVY); cell.alignment=CTR

    for i,wacc in enumerate(wacc_vals):
        row=65+i
        hdr=ws.cell(row=row,column=3)
        hdr.value=wacc; hdr.number_format=FMT_P
        is_base_row=abs(wacc-base_wacc)<0.001
        hdr.font=fn(WHF,True); hdr.fill=fl(RED_BG if is_base_row else NAVY); hdr.alignment=CTR
        for j,mult in enumerate(mult_vals):
            col=4+j
            is_base=abs(wacc-base_wacc)<0.001 and abs(mult-base_mult)<0.01
            fcf_pv=fcf_pv_tmpl.format(wacc=wacc)
            tv_exit=f"IS!{last_est_ebitda_col}51*{mult}/(1+{wacc})^{m}"
            eq_val=f"=({fcf_pv}+{tv_exit}+D12)/D13"
            cell=ws.cell(row=row,column=col)
            cell.value=eq_val; cell.number_format=FMT_EP
            cell.font=fn(BKF,is_base); cell.fill=fl(YELLOW if is_base else WHITE); cell.alignment=CTR
    w(ws,70,3,f"Base case: WACC={base_wacc:.1%}, Exit EV/EBITDA={base_mult:.1f}x",WHITE,GYF,italic=True,align=LFT,mc=9)
    ws.row_dimensions[70].height=12

    # ── SECTION 6: EV/NTM REVENUE MULTIPLES ─────────────────────────────────
    spacer(ws,71)
    w(ws,72,2,"EV / NTM REVENUE MULTIPLES",SECT,SECT_TXT,True,align=LFT,mc=7+m)
    w(ws,73,2,f"NTM Revenue ({curr}{u})",WHITE,BKF,align=LFT)
    w(ws,73,4,f"=IS!{ntm_ebitda_col}7",WHITE,GRF,fmt=FMT_NN,align=CTR)
    w(ws,74,2,f"Net Cash ({curr}{u})",WHITE,BKF,align=LFT)
    w(ws,74,4,"=D12",WHITE,GRF,fmt=FMT_NN,align=CTR)
    w(ws,75,2,f"Diluted Shares ({u})",WHITE,BKF,align=LFT)
    w(ws,75,4,"=D13",WHITE,GRF,fmt=FMT_NN,align=CTR)
    w(ws,76,2,"Peer context: adjust multiple range to sector. EV/NTM Revenue = EV ÷ forward revenue.",WHITE,GYF,italic=True,align=LFT,mc=7+m)
    ws.row_dimensions[76].height=14

    mult_rev=[1.0,1.5,2.0,2.5,3.0,3.5,4.0]
    base_rev=2.5
    w(ws,77,3,"Multiple →",NAVY,WHF,True,align=CTR)
    for j,mult in enumerate(mult_rev):
        cell=ws.cell(row=77,column=4+j)
        cell.value=mult; cell.number_format=FMT_MX
        cell.font=fn(WHF,True); cell.fill=fl(NAVY); cell.alignment=CTR

    for j,mult in enumerate(mult_rev):
        is_base=abs(mult-base_rev)<0.01
        w(ws,78,4+j,f"=D73*{mult}",LT_BLUE,BKF,fmt=FMT_NN,align=CTR)
        ws.cell(row=78,column=4+j).fill=fl(YELLOW if is_base else LT_BLUE)
        w(ws,79,4+j,f"={get_column_letter(4+j)}78+D74",WHITE,BKF,fmt=FMT_NN,align=CTR)
        ws.cell(row=79,column=4+j).fill=fl(YELLOW if is_base else WHITE)
        w(ws,80,4+j,f"={get_column_letter(4+j)}79/D75",GREEN_H,WHF,True,fmt=FMT_EP,align=CTR)
        ws.cell(row=80,column=4+j).fill=fl(YELLOW if is_base else GREEN_H)

    w(ws,78,2,f"Implied EV ({curr}{u})",WHITE,BKF,align=LFT)
    w(ws,79,2,f"Equity Value ({curr}{u})",WHITE,BKF,align=LFT)
    w(ws,80,2,f"Implied Share Price ({curr})",WHITE,WHF,True,align=LFT)
    w(ws,81,3,f"Base case: {base_rev:.1f}x NTM Revenue.",WHITE,GYF,italic=True,align=LFT,mc=10)
    ws.row_dimensions[81].height=12

    return ws


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    parser=argparse.ArgumentParser(description="Generic financial model builder")
    parser.add_argument("ticker",type=str.upper,help="Ticker symbol (e.g. GS, MC, DGE)")
    parser.add_argument("--company","-c",default=None,help="Company name for title")
    parser.add_argument("--currency","--ccy",default="",help="Currency symbol (e.g. $ £ €)")
    parser.add_argument("--units","-u",default="M",help="Units label (default: M)")
    parser.add_argument("--output-dir","-o",type=Path,default=None)
    args=parser.parse_args()

    ticker=args.ticker
    company=args.company or ticker
    research_dir=args.output_dir or DEFAULT_RESEARCH_DIR
    ticker_dir=research_dir/ticker

    fin_path=ticker_dir/f"{ticker}_tikr_financials.csv"
    est_path=ticker_dir/f"{ticker}_tikr_forward_estimates.csv"
    out_path=ticker_dir/f"{ticker}_Financial_Model.xlsx"

    # Estimates are always required
    if not est_path.exists():
        print(f"Error: File not found: {est_path}",file=sys.stderr)
        sys.exit(1)

    # Financials: prefer TIKR CSV, fall back to SEC combined CSVs
    sec_combined = ticker_dir / "combined" / f"{ticker}_10K_income_statement_combined.csv"
    if not fin_path.exists() and not sec_combined.exists():
        print(f"Error: No financials found. Expected one of:",file=sys.stderr)
        print(f"  {fin_path}",file=sys.stderr)
        print(f"  {sec_combined}",file=sys.stderr)
        sys.exit(1)

    fin_source = "TIKR" if fin_path.exists() else "SEC EDGAR"
    print(f"Building model for {ticker} ({company}) ...")
    print(f"  Financials: {fin_path if fin_path.exists() else ticker_dir/'combined'} [{fin_source}]")
    print(f"  Estimates:  {est_path}")

    try:
        md=build_model_data(ticker,company,args.currency,args.units,fin_path,est_path)
    except Exception as e:
        print(f"Error parsing CSV data: {e}",file=sys.stderr)
        raise

    print(f"  Actuals:    {md['n_act']} years ({', '.join(md['act_yrs'])})")
    print(f"  Estimates:  {md['n_est']} years ({', '.join(md['est_yrs'])})")

    wb=Workbook()
    wb.remove(wb.active)
    build_assumptions(wb,md)
    build_is(wb,md)
    build_bs(wb,md)
    build_cfs(wb,md)
    build_dcf(wb,md)

    sheet_order=["IS","BS","CFS","Assumptions","DCF"]
    for target_idx,name in enumerate(sheet_order):
        if name in wb.sheetnames:
            current_idx=wb.sheetnames.index(name)
            if current_idx!=target_idx:
                wb.move_sheet(name,offset=target_idx-current_idx)

    out_path.parent.mkdir(parents=True,exist_ok=True)
    wb.save(out_path)
    print(f"Saved → {out_path}")


if __name__=="__main__":
    main()
