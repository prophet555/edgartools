#!/usr/bin/env python3
"""
Fix HEIO_model.xlsx:
- Fill in all actuals (IS, BS, CFS) from TIKR CSV data
- Fix IS estimate formulas (Assumptions row refs were off by +1)
- Fix CFS estimate formulas (labels/formulas were completely scrambled)
- Complete BS estimate formulas (most were missing)
- Build out DCF tab properly
"""

from pathlib import Path
import openpyxl

PATH = Path.home() / "Library/CloudStorage/OneDrive-AG/01_equities_research/HEIO/HEIO_model.xlsx"

wb = openpyxl.load_workbook(PATH)
IS  = wb["IS"]
BS  = wb["BS"]
CFS = wb["CFS"]
DCF = wb["DCF"]
ASS = wb["Assumptions"]

# ── Column layout ─────────────────────────────────────────────────────────────
# B(2)=labels, C(3)=2023 actual, D(4)=2024 actual, E(5)=2025 actual
# F(6)=2026E, G(7)=2027E, H(8)=2028E, I(9)=2029E
# (model has no 2022 column — the "12/31/22" in row 3 col B is a label-column header mistake)

ACT_COLS = [3, 4, 5]   # C, D, E
EST_COLS = [6, 7, 8, 9] # F, G, H, I
EST_YEARS = ["2026E", "2027E", "2028E", "2029E"]

def col(n):
    from openpyxl.utils import get_column_letter
    return get_column_letter(n)

def set_val(ws, r, c, v):
    ws.cell(row=r, column=c).value = v

# ─────────────────────────────────────────────────────────────────────────────
# Fix header row 3 col B on all tabs ("12/31/22" → blank — it's the label col)
# ─────────────────────────────────────────────────────────────────────────────
for ws in [IS, BS, CFS]:
    ws.cell(row=3, column=2).value = None

# ══════════════════════════════════════════════════════════════════════════════
# IS TAB — fill actuals + fix estimate formulas
# ══════════════════════════════════════════════════════════════════════════════

# IS actuals: [2023, 2024, 2025]
IS_ACTUALS = {
    4:  [30362,  29821,  28753],   # Total Revenues
    5:  [-19659, -19025, -18128],  # Cost of Goods Sold
    6:  [10703,  10796,  10625],   # Gross Profit
    7:  [-4392,  -4527,  -4494],   # SG&A
    8:  [-2291,  -2297,  -2273],   # D&A
    9:  [-156,   -136,   -56],     # Other OpEx
    10: [3864,   3836,   3802],    # Operating Income / EBIT
    11: [-640,   -680,   -620],    # Interest Expense
    12: [97,     128,    116],     # Interest Income
    13: [218,    -705,   255],     # Equity Income
    14: [-323,   -217,   -47],     # FX Gains (Loss)
    15: [-117,   -92,    -135],    # Other Non-Op
    16: [3099,   2270,   3371],    # EBT Excl. Unusual
    17: [2401,   1161,   2138],    # Net Income to Company
    18: [-1227,  -663,   -1186],   # Minority Interest
    19: [1174,   498,    952],     # Net Income
    20: [283.97, 282.87, 281.11],  # Diluted Shares
    21: [4.12,   1.76,   3.39],    # EPS
}

for row, vals in IS_ACTUALS.items():
    for ci, v in zip(ACT_COLS, vals):
        set_val(IS, row, ci, v)

# Fix IS estimate formulas: all had Assumptions row off by +1.
# Correct mapping: Revenue Growth=row5, Gross Margin=6, SGA%=7, DA=8,
# OtherOpEx%=9, IntExp=10, IntInc=11, EqInc=12, FX=13, OthNI=14,
# TaxRate=15, MI%=16, Shares=17, SBC=18, DPS=19

# Generate corrected formulas for each estimate column
EST_COL_LETTERS = [col(c) for c in EST_COLS]  # F, G, H, I
PREV_COL = {
    6: "E",   # 2026E: prior year = 2025 = col E
    7: "F",
    8: "G",
    9: "H",
}

for ci, cl_est in zip(EST_COLS, EST_COL_LETTERS):
    prev = PREV_COL[ci]
    ac = EST_COL_LETTERS.index(cl_est)  # 0-based index into EST_COL_LETTERS
    ass_col = col(3 + ac)  # Assumptions columns: C,D,E,F for 2026E-2029E

    # Row 4: Revenue = prev_rev * (1 + rev_growth)
    IS.cell(row=4, column=ci).value = f"={prev}4*(1+Assumptions!{ass_col}5)"

    # Row 5: COGS = -Revenue*(1-gross_margin)
    IS.cell(row=5, column=ci).value = f"=-{cl_est}4*(1-Assumptions!{ass_col}6)"

    # Row 6: Gross Profit = Revenue + COGS
    IS.cell(row=6, column=ci).value = f"={cl_est}4+{cl_est}5"

    # Row 7: SG&A = -Revenue * sga_pct
    IS.cell(row=7, column=ci).value = f"=-{cl_est}4*Assumptions!{ass_col}7"

    # Row 8: D&A = DA assumption (fixed $)
    IS.cell(row=8, column=ci).value = f"=-Assumptions!{ass_col}8"

    # Row 9: Other OpEx = -Revenue * otherOpEx_pct
    IS.cell(row=9, column=ci).value = f"=-{cl_est}4*Assumptions!{ass_col}9"

    # Row 10: EBIT = GP + SGA + DA + OtherOpEx (all negative already)
    IS.cell(row=10, column=ci).value = f"={cl_est}6+{cl_est}7+{cl_est}8+{cl_est}9"

    # Row 11: Interest Expense = -assumption
    IS.cell(row=11, column=ci).value = f"=-Assumptions!{ass_col}10"

    # Row 12: Interest Income
    IS.cell(row=12, column=ci).value = f"=Assumptions!{ass_col}11"

    # Row 13: Equity Income
    IS.cell(row=13, column=ci).value = f"=Assumptions!{ass_col}12"

    # Row 14: FX
    IS.cell(row=14, column=ci).value = f"=Assumptions!{ass_col}13"

    # Row 15: Other Non-Op
    IS.cell(row=15, column=ci).value = f"=Assumptions!{ass_col}14"

    # Row 16: EBT Excl Unusual = sum of EBIT + non-op items
    IS.cell(row=16, column=ci).value = (
        f"={cl_est}10+{cl_est}11+{cl_est}12+{cl_est}13+{cl_est}14+{cl_est}15"
    )

    # Row 17: Net Income to Company = EBT * (1 - tax_rate)
    IS.cell(row=17, column=ci).value = f"={cl_est}16*(1-Assumptions!{ass_col}15)"

    # Row 18: Minority Interest = -NIC * mi_pct
    IS.cell(row=18, column=ci).value = f"=-{cl_est}17*Assumptions!{ass_col}16"

    # Row 19: Net Income
    IS.cell(row=19, column=ci).value = f"={cl_est}17+{cl_est}18"

    # Row 20: Shares
    IS.cell(row=20, column=ci).value = f"=Assumptions!{ass_col}17"

    # Row 21: EPS
    IS.cell(row=21, column=ci).value = f"={cl_est}19/{cl_est}20"

# ══════════════════════════════════════════════════════════════════════════════
# BS TAB — fill actuals + complete all estimate formulas
# ══════════════════════════════════════════════════════════════════════════════

# BS actuals: [2023, 2024, 2025]
BS_ACTUALS = {
    4:  [2377,   2350,   4816],    # Cash
    5:  [3376,   3125,   3070],    # Accounts Receivable
    6:  [1307,   1066,   1124],    # Other Receivables
    7:  [3721,   3572,   3263],    # Inventory
    8:  [532,    562,    520],     # Prepaid
    9:  [86,     224,    143],     # Other Current Assets
    10: [11399,  10899,  12936],   # Total Current Assets
    11: [29505,  29822,  31137],   # Gross PP&E
    12: [-14733, -15145, -16600],  # Accumulated Depreciation
    13: [14772,  14677,  14537],   # Net PP&E
    14: [4311,   3681,   3407],    # LT Investments
    15: [12238,  12301,  11588],   # Goodwill
    16: [9543,   9400,   8423],    # Other Intangibles
    17: [1292,   1264,   1213],    # Deferred Tax Assets
    18: [1018,   1089,   1174],    # Other LT Assets
    19: [54758,  53475,  53474],   # TOTAL ASSETS
    20: [5735,   5986,   5830],    # Accounts Payable
    21: [1944,   2042,   2133],    # Accrued Expenses
    22: [1443,   1090,   961],     # Short-term Borrowings
    23: [2443,   1862,   1781],    # Current LT Debt
    24: [306,    314,    346],     # Capital Lease Current
    25: [1752,   1734,   1586],    # Income Tax Payable
    26: [1202,   1210,   1295],    # Other Current Liab
    27: [14825,  14238,  13932],   # Total Current Liabilities
    28: [13085,  12753,  15021],   # LT Debt
    29: [961,    1030,   1170],    # Capital Leases LT
    30: [586,    519,    542],     # Pension
    31: [2213,   2155,   1820],    # Deferred Tax Liab
    32: [694,    676,    654],     # Other Non-Current Liab
    33: [32364,  31371,  33139],   # TOTAL LIABILITIES
    34: [461,    461,    461],     # Common Stock
    35: [1257,   1257,   1257],    # APIC
    36: [10247,  10308,  10599],   # Retained Earnings
    37: [-390,   -390,   -696],    # Treasury Stock
    38: [-1842,  -2090,  -2991],   # AOCI
    39: [9733,   9546,   8630],    # Total Common Equity
    40: [12661,  12558,  11705],   # Minority Interest
    41: [22394,  22104,  20335],   # Total Equity
    42: [54758,  53475,  53474],   # TOTAL L&E
}

for row, vals in BS_ACTUALS.items():
    for ci, v in zip(ACT_COLS, vals):
        set_val(BS, row, ci, v)

# BS estimate formulas
# Assumptions layout (relevant rows):
#   22=DSO, 23=DPO, 24=Inv Days, 25=OtherRec%Rev, 26=Prepaid%Rev,
#   27=OtherCA$, 28=LTInv$, 29=Goodwill$, 30=OtherInt$, 31=OtherLTA$,
#   32=DTA$, 33=AccruedExp%Rev, 34=STB$, 35=CurrLTD$, 36=CapLeaseCurr$,
#   37=ItaxPayable%Rev, 38=OtherCurrLiab$, 39=LTDebt$, 40=CapLeasesLT$,
#   41=Pension$, 42=DTLiab$, 43=OtherNCL$, 44=CS$, 45=APIC$,
#   46=TreasuryStock$, 47=AOCI$, 48=MIEquity$
# IS: row4=Revenue, row5=COGS, row8=DA, row19=NI
# CFS: row22=EndingCash

PREV_BS = {6: "E", 7: "F", 8: "G", 9: "H"}

for ci, cl_est in zip(EST_COLS, EST_COL_LETTERS):
    prev = PREV_BS[ci]
    ac   = EST_COL_LETTERS.index(cl_est)
    ass  = col(3 + ac)  # Assumptions col: C,D,E,F

    # Row 4: Cash = CFS ending cash
    BS.cell(row=4, column=ci).value = f"=CFS!{cl_est}22"

    # Row 5: AR = Revenue * DSO / 365
    BS.cell(row=5, column=ci).value = f"=IS!{cl_est}4*Assumptions!{ass}22/365"

    # Row 6: Other Receivables = Revenue * OtherRec%
    BS.cell(row=6, column=ci).value = f"=IS!{cl_est}4*Assumptions!{ass}25"

    # Row 7: Inventory = abs(COGS) * Inv_Days / 365
    BS.cell(row=7, column=ci).value = f"=-IS!{cl_est}5*Assumptions!{ass}24/365"

    # Row 8: Prepaid = Revenue * Prepaid%
    BS.cell(row=8, column=ci).value = f"=IS!{cl_est}4*Assumptions!{ass}26"

    # Row 9: Other Current Assets = fixed
    BS.cell(row=9, column=ci).value = f"=Assumptions!{ass}27"

    # Row 10: Total Current Assets = sum
    BS.cell(row=10, column=ci).value = (
        f"={cl_est}4+{cl_est}5+{cl_est}6+{cl_est}7+{cl_est}8+{cl_est}9"
    )

    # Row 11: Gross PP&E = prior + CapEx
    BS.cell(row=11, column=ci).value = f"={prev}11+Assumptions!{ass}51"

    # Row 12: Accumulated Depreciation = prior - IS D&A (D&A deepens the negative)
    BS.cell(row=12, column=ci).value = f"={prev}12+IS!{cl_est}8"

    # Row 13: Net PP&E = Gross + Accum Dep
    BS.cell(row=13, column=ci).value = f"={cl_est}11+{cl_est}12"

    # Row 14: LT Investments = fixed assumption
    BS.cell(row=14, column=ci).value = f"=Assumptions!{ass}28"

    # Row 15: Goodwill = fixed assumption
    BS.cell(row=15, column=ci).value = f"=Assumptions!{ass}29"

    # Row 16: Other Intangibles = amortises per assumption
    BS.cell(row=16, column=ci).value = f"=Assumptions!{ass}30"

    # Row 17: Deferred Tax Assets = fixed assumption
    BS.cell(row=17, column=ci).value = f"=Assumptions!{ass}32"

    # Row 18: Other LT Assets = fixed assumption
    BS.cell(row=18, column=ci).value = f"=Assumptions!{ass}31"

    # Row 19: TOTAL ASSETS = TCA + non-current items
    BS.cell(row=19, column=ci).value = (
        f"={cl_est}10+{cl_est}13+{cl_est}14+{cl_est}15+{cl_est}16+{cl_est}17+{cl_est}18"
    )

    # Row 20: AP = abs(COGS) * DPO / 365
    BS.cell(row=20, column=ci).value = f"=-IS!{cl_est}5*Assumptions!{ass}23/365"

    # Row 21: Accrued Expenses = Revenue * Accrued%
    BS.cell(row=21, column=ci).value = f"=IS!{cl_est}4*Assumptions!{ass}33"

    # Row 22: STB = fixed
    BS.cell(row=22, column=ci).value = f"=Assumptions!{ass}34"

    # Row 23: Current LT Debt = fixed
    BS.cell(row=23, column=ci).value = f"=Assumptions!{ass}35"

    # Row 24: Cap Lease Current = fixed
    BS.cell(row=24, column=ci).value = f"=Assumptions!{ass}36"

    # Row 25: Income Tax Payable = Revenue * ITP%
    BS.cell(row=25, column=ci).value = f"=IS!{cl_est}4*Assumptions!{ass}37"

    # Row 26: Other Current Liab = fixed
    BS.cell(row=26, column=ci).value = f"=Assumptions!{ass}38"

    # Row 27: Total Current Liabilities
    BS.cell(row=27, column=ci).value = (
        f"={cl_est}20+{cl_est}21+{cl_est}22+{cl_est}23+{cl_est}24+{cl_est}25+{cl_est}26"
    )

    # Row 28: LT Debt = fixed
    BS.cell(row=28, column=ci).value = f"=Assumptions!{ass}39"

    # Row 29: Capital Leases LT = fixed
    BS.cell(row=29, column=ci).value = f"=Assumptions!{ass}40"

    # Row 30: Pension = fixed
    BS.cell(row=30, column=ci).value = f"=Assumptions!{ass}41"

    # Row 31: Deferred Tax Liab = fixed
    BS.cell(row=31, column=ci).value = f"=Assumptions!{ass}42"

    # Row 32: Other NCL = fixed
    BS.cell(row=32, column=ci).value = f"=Assumptions!{ass}43"

    # Row 33: TOTAL LIABILITIES
    BS.cell(row=33, column=ci).value = (
        f"={cl_est}27+{cl_est}28+{cl_est}29+{cl_est}30+{cl_est}31+{cl_est}32"
    )

    # Row 34: Common Stock = fixed
    BS.cell(row=34, column=ci).value = f"=Assumptions!{ass}44"

    # Row 35: APIC = fixed
    BS.cell(row=35, column=ci).value = f"=Assumptions!{ass}45"

    # Row 36: Retained Earnings = prior + NI - Dividends (DPS * shares)
    BS.cell(row=36, column=ci).value = (
        f"={prev}36+IS!{cl_est}19-Assumptions!{ass}19*Assumptions!{ass}17"
    )

    # Row 37: Treasury Stock = per assumption (buybacks increase treasury)
    BS.cell(row=37, column=ci).value = f"=Assumptions!{ass}46"

    # Row 38: AOCI = fixed
    BS.cell(row=38, column=ci).value = f"=Assumptions!{ass}47"

    # Row 39: Total Common Equity
    BS.cell(row=39, column=ci).value = (
        f"={cl_est}34+{cl_est}35+{cl_est}36+{cl_est}37+{cl_est}38"
    )

    # Row 40: Minority Interest Equity = fixed
    BS.cell(row=40, column=ci).value = f"=Assumptions!{ass}48"

    # Row 41: Total Equity
    BS.cell(row=41, column=ci).value = f"={cl_est}39+{cl_est}40"

    # Row 42: TOTAL L&E
    BS.cell(row=42, column=ci).value = f"={cl_est}33+{cl_est}41"

# ══════════════════════════════════════════════════════════════════════════════
# CFS TAB — fill actuals + fix estimate formulas
# ══════════════════════════════════════════════════════════════════════════════

# CFS actuals: [2023, 2024, 2025]
# Using BS cash for Beginning/Ending (internally consistent)
CFS_ACTUALS = {
    4:  [1174,   498,    952],     # Net Income
    5:  [1975,   2159,   2147],    # D&A (CFS add-back)
    6:  [0,      0,      0],       # SBC (not broken out, assumed ~0 in actuals)
    7:  [-42,    347,    54],      # Change in AR
    8:  [-4,     -39,    2],       # Change in Inventory
    9:  [-100,   543,    271],     # Change in AP
    10: [-32,    -6,     98],      # Change in Other WC
    11: [4430,   5503,   5012],    # Net Cash from Operations
    12: [-2434,  -2184,  -2133],   # CapEx
    13: [-3576,  -2435,  -2455],   # Net Cash from Investing
    14: [6751,   3076,   6582],    # Debt Issued
    15: [-5004,  -4446,  -4430],   # Debt Repaid
    16: [-942,   -5,     -658],    # Buybacks
    17: [-1335,  -1199,  -1276],   # Dividends Paid
    18: [-816,   -2574,  196],     # Net Cash from Financing
    19: [-231,   -166,   -151],    # FX Effect
    20: [-388,   -27,    2466],    # Net Change in Cash (= BS ending - BS beginning)
    21: [2765,   2377,   2350],    # Beginning Cash (= prior year BS cash)
    22: [2377,   2350,   4816],    # Ending Cash (= BS cash)
    23: [1996,   3319,   2879],    # Free Cash Flow
}

for row, vals in CFS_ACTUALS.items():
    for ci, v in zip(ACT_COLS, vals):
        set_val(CFS, row, ci, v)

# Fix CFS estimate formulas (all were scrambled)
for ci, cl_est in zip(EST_COLS, EST_COL_LETTERS):
    prev = PREV_BS[ci]
    ac   = EST_COL_LETTERS.index(cl_est)
    ass  = col(3 + ac)

    # Row 4: Net Income from IS
    CFS.cell(row=4, column=ci).value = f"=IS!{cl_est}19"

    # Row 5: D&A = from IS D&A line (D&A assumption drives IS row 8, which is negative; negate for CFS add-back)
    CFS.cell(row=5, column=ci).value = f"=-IS!{cl_est}8"

    # Row 6: SBC from Assumptions
    CFS.cell(row=6, column=ci).value = f"=Assumptions!{ass}18"

    # Row 7: Change in AR = -(AR_t - AR_{t-1})
    CFS.cell(row=7, column=ci).value = f"=-(BS!{cl_est}5-BS!{prev}5)"

    # Row 8: Change in Inventory = -(Inv_t - Inv_{t-1})
    CFS.cell(row=8, column=ci).value = f"=-(BS!{cl_est}7-BS!{prev}7)"

    # Row 9: Change in AP = AP_t - AP_{t-1}
    CFS.cell(row=9, column=ci).value = f"=BS!{cl_est}20-BS!{prev}20"

    # Row 10: Change in Other WC = 0 (simplification)
    CFS.cell(row=10, column=ci).value = 0

    # Row 11: Net Cash from Operations
    CFS.cell(row=11, column=ci).value = (
        f"={cl_est}4+{cl_est}5+{cl_est}6+{cl_est}7+{cl_est}8+{cl_est}9+{cl_est}10"
    )

    # Row 12: CapEx = negative (cash outflow)
    CFS.cell(row=12, column=ci).value = f"=-Assumptions!{ass}51"

    # Row 13: Net Cash from Investing = CapEx only (simplified)
    CFS.cell(row=13, column=ci).value = f"={cl_est}12"

    # Row 14: Debt Issued
    CFS.cell(row=14, column=ci).value = f"=Assumptions!{ass}53"

    # Row 15: Debt Repaid (negative)
    CFS.cell(row=15, column=ci).value = f"=-Assumptions!{ass}54"

    # Row 16: Buybacks (negative)
    CFS.cell(row=16, column=ci).value = f"=-Assumptions!{ass}52"

    # Row 17: Dividends Paid = -(DPS * Diluted Shares)
    CFS.cell(row=17, column=ci).value = f"=-Assumptions!{ass}19*Assumptions!{ass}17"

    # Row 18: Net Cash from Financing
    CFS.cell(row=18, column=ci).value = (
        f"={cl_est}14+{cl_est}15+{cl_est}16+{cl_est}17"
    )

    # Row 19: FX Effect
    CFS.cell(row=19, column=ci).value = f"=Assumptions!{ass}55"

    # Row 20: Net Change in Cash = CFO + CFI + CFF + FX
    CFS.cell(row=20, column=ci).value = (
        f"={cl_est}11+{cl_est}13+{cl_est}18+{cl_est}19"
    )

    # Row 21: Beginning Cash = prior year BS cash
    if ci == 6:  # 2026E: prior year = 2025 actual = BS!E4
        CFS.cell(row=21, column=ci).value = f"=BS!E4"
    else:         # later years: prior year ending cash
        CFS.cell(row=21, column=ci).value = f"={prev}22"

    # Row 22: Ending Cash = Beginning + Net Change
    CFS.cell(row=22, column=ci).value = f"={cl_est}21+{cl_est}20"

    # Row 23: Free Cash Flow = CFO + CapEx
    CFS.cell(row=23, column=ci).value = f"={cl_est}11+{cl_est}12"

# ══════════════════════════════════════════════════════════════════════════════
# DCF TAB — build out properly
# ══════════════════════════════════════════════════════════════════════════════
# Existing: rows 4-11 have basic assumptions (WACC, TGR, Cash, Debt, NetDebt, Shares, Price)
# Add: FCF build, PV calcs, valuation bridge, sensitivity table

# Clear everything beyond row 11 to rebuild
for row in range(12, 60):
    for c in range(1, 12):
        DCF.cell(row=row, column=c).value = None

# Also fix Net Debt formula (was hardcoded, make it a formula)
DCF.cell(row=9, column=3).value = "=C7-C8"   # Cash - Total Debt

# Estimate year columns in DCF: D(4)=2026E, E(5)=2027E, F(6)=2028E, G(7)=2029E
DCF_EST = {6: "D", 7: "E", 8: "F", 9: "G"}  # DCF_col_idx: letter mapping
# Note: DCF uses its OWN column layout starting at D for estimates
DCF_EST_COLS = [4, 5, 6, 7]   # D, E, F, G in DCF
DCF_EST_LETS = ["D", "E", "F", "G"]

IS_EST_LETS = ["F", "G", "H", "I"]  # IS tab estimate columns

r = 13

# Section header
DCF.cell(row=r, column=2).value = "FCF BUILD"
r += 1

# Column headers for estimate years
DCF.cell(row=r, column=3).value = "Metric"
for i, yr in enumerate(EST_YEARS):
    DCF.cell(row=r, column=4+i).value = yr
DCF.cell(row=r, column=8).value = "Terminal"
r += 1

# Track FCF build rows
DCFR = {}

DCFR["rev"] = r
DCF.cell(row=r, column=2).value = "Revenue"
for i, (il, dl) in enumerate(zip(IS_EST_LETS, DCF_EST_LETS)):
    DCF.cell(row=r, column=4+i).value = f"=IS!{il}4"
r += 1

DCFR["ebit"] = r
DCF.cell(row=r, column=2).value = "EBIT"
for i, il in enumerate(IS_EST_LETS):
    DCF.cell(row=r, column=4+i).value = f"=IS!{il}10"
r += 1

DCFR["tax_ebit"] = r
DCF.cell(row=r, column=2).value = "Tax on EBIT"
for i, (il, dl) in enumerate(zip(IS_EST_LETS, DCF_EST_LETS)):
    DCF.cell(row=r, column=4+i).value = f"=-MAX(0,IS!{il}10)*Assumptions!{col(3+i)}15"
r += 1

DCFR["nopat"] = r
DCF.cell(row=r, column=2).value = "NOPAT"
for i, dl in enumerate(DCF_EST_LETS):
    DCF.cell(row=r, column=4+i).value = f"={dl}{DCFR['ebit']}+{dl}{DCFR['tax_ebit']}"
r += 1

DCFR["da"] = r
DCF.cell(row=r, column=2).value = "(+) D&A"
for i, il in enumerate(IS_EST_LETS):
    DCF.cell(row=r, column=4+i).value = f"=-IS!{il}8"
r += 1

DCFR["capex"] = r
DCF.cell(row=r, column=2).value = "(-) CapEx"
for i, il in enumerate(IS_EST_LETS):
    DCF.cell(row=r, column=4+i).value = f"=CFS!{il}12"
r += 1

DCFR["dwc"] = r
DCF.cell(row=r, column=2).value = "(+/-) Change in WC"
for i, il in enumerate(IS_EST_LETS):
    DCF.cell(row=r, column=4+i).value = f"=CFS!{il}7+CFS!{il}8+CFS!{il}9"
r += 1

DCFR["ufcf"] = r
DCF.cell(row=r, column=2).value = "Unlevered Free Cash Flow"
for i, dl in enumerate(DCF_EST_LETS):
    DCF.cell(row=r, column=4+i).value = (
        f"={dl}{DCFR['nopat']}+{dl}{DCFR['da']}+{dl}{DCFR['capex']}+{dl}{DCFR['dwc']}"
    )
r += 1

# Terminal Value (Gordon Growth using last FCF)
DCF.cell(row=r, column=2).value = "Terminal Value (Gordon Growth)"
last_fcf_col = DCF_EST_LETS[-1]  # G = last estimate year
DCF.cell(row=r, column=8).value = (
    f"={last_fcf_col}{DCFR['ufcf']}*(1+C6)/(C5-C6)"
)
TV_ROW = r
r += 2

# Section: PV Calculations
DCF.cell(row=r, column=2).value = "PV CALCULATIONS"
r += 1

DCF.cell(row=r, column=2).value = "Period"
for i in range(4):
    DCF.cell(row=r, column=4+i).value = i + 1
DCF.cell(row=r, column=8).value = 4  # TV discounted at period 4
PERIOD_ROW = r; r += 1

DCF.cell(row=r, column=2).value = "Discount Factor"
for i, dl in enumerate(DCF_EST_LETS):
    DCF.cell(row=r, column=4+i).value = f"=1/(1+C5)^{dl}{PERIOD_ROW}"
DCF.cell(row=r, column=8).value = f"=1/(1+C5)^{last_fcf_col}{PERIOD_ROW}"
DISC_ROW = r; r += 1

DCF.cell(row=r, column=2).value = "PV of FCF"
for i, dl in enumerate(DCF_EST_LETS):
    DCF.cell(row=r, column=4+i).value = f"={dl}{DCFR['ufcf']}*{dl}{DISC_ROW}"
DCF.cell(row=r, column=8).value = f"=H{TV_ROW}*H{DISC_ROW}"
PV_FCF_ROW = r; r += 2

# Section: Valuation Bridge
DCF.cell(row=r, column=2).value = "VALUATION"
r += 1

sum_pvfcf = "+".join([f"{dl}{PV_FCF_ROW}" for dl in DCF_EST_LETS])
SUM_ROW = r
DCF.cell(row=r, column=2).value = "Sum of PV FCFs"
DCF.cell(row=r, column=3).value = f"={sum_pvfcf}"
r += 1

PVTV_ROW = r
DCF.cell(row=r, column=2).value = "PV of Terminal Value"
DCF.cell(row=r, column=3).value = f"=H{PV_FCF_ROW}"
r += 1

EV_ROW = r
DCF.cell(row=r, column=2).value = "Enterprise Value"
DCF.cell(row=r, column=3).value = f"=C{SUM_ROW}+C{PVTV_ROW}"
r += 1

DCF.cell(row=r, column=2).value = "(-) Net Debt"
DCF.cell(row=r, column=3).value = f"=C9"   # Net Debt from row 9
ND_ROW = r; r += 1

EQV_ROW = r
DCF.cell(row=r, column=2).value = "Equity Value"
DCF.cell(row=r, column=3).value = f"=C{EV_ROW}-C{ND_ROW}"
r += 1

IVS_ROW = r
DCF.cell(row=r, column=2).value = "Intrinsic Value Per Share (€)"
DCF.cell(row=r, column=3).value = f"=C{EQV_ROW}/C10"
r += 1

DCF.cell(row=r, column=2).value = "Current Price (€)"
DCF.cell(row=r, column=3).value = f"=C11"
DCF.cell(row=r+1, column=2).value = "Upside / (Downside)"
DCF.cell(row=r+1, column=3).value = f"=C{IVS_ROW}/C11-1"
r += 3

# Section: Sensitivity — WACC × TGR → Implied Share Price
DCF.cell(row=r, column=2).value = "SENSITIVITY: Implied Share Price (€)"
r += 1

wacc_vals = [0.065, 0.075, 0.085, 0.095, 0.105]
tgr_vals  = [0.010, 0.015, 0.020, 0.025, 0.030, 0.035]

# Corner label
DCF.cell(row=r, column=3).value = "WACC \\ TGR →"
for j, tg in enumerate(tgr_vals):
    DCF.cell(row=r, column=4+j).value = tg
r += 1

for wc in wacc_vals:
    DCF.cell(row=r, column=3).value = wc
    for j, tg in enumerate(tgr_vals):
        # Compute implied share price for this WACC/TGR
        # uFCF refs from DCF tab
        pv_terms = "+".join([
            f"({dl}{DCFR['ufcf']}/(1+{wc})^{i+1})"
            for i, dl in enumerate(DCF_EST_LETS)
        ])
        tv_term = f"({last_fcf_col}{DCFR['ufcf']}*(1+{tg})/({wc}-{tg}))/(1+{wc})^4"
        ev_formula = f"=({pv_terms}+{tv_term}-C9)/C10"
        DCF.cell(row=r, column=4+j).value = ev_formula
    r += 1

wb.save(PATH)
print(f"Saved fixed model → {PATH}")

# Quick sanity check
wb2 = openpyxl.load_workbook(PATH)
IS2 = wb2["IS"]
print(f"\nSanity checks:")
print(f"  IS!E4 (2025 Revenue): {IS2.cell(4,5).value}")
print(f"  IS!F4 formula (2026E Rev): {IS2.cell(4,6).value}")
print(f"  IS!E19 (2025 NI): {IS2.cell(19,5).value}")
CFS2 = wb2["CFS"]
print(f"  CFS!E22 (2025 Ending Cash): {CFS2.cell(22,5).value}")
print(f"  CFS!F22 formula (2026E End Cash): {CFS2.cell(22,6).value}")
BS2  = wb2["BS"]
print(f"  BS!E4 (2025 Cash): {BS2.cell(4,5).value}")
print(f"  BS!F4 formula (2026E Cash): {BS2.cell(4,6).value}")
DCF2 = wb2["DCF"]
print(f"  DCF EV formula: {DCF2.cell(EV_ROW,3).value}")
print(f"  DCF IVS formula: {DCF2.cell(IVS_ROW,3).value}")
