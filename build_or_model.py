#!/usr/bin/env python3
"""Build L'Oréal SA (OR) 5-tab Financial Model Excel workbook."""

from pathlib import Path
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = (
    Path.home() / "Library/CloudStorage/OneDrive-AG/01_equities_research"
    / "OR" / "OR_financial_model.xlsx"
)

# ── Colours ───────────────────────────────────────────────────────────────────
NAVY       = "1F3864"
BLUE_HDR   = "2F5496"
GREEN_HDR  = "375623"
BROWN_TAB  = "843C0C"
PURPLE_TAB = "7030A0"
RED_TAB    = "C00000"
SECT_BG    = "D6E4F7"
GREEN_HL   = "E2EFDA"
BLUE_IN    = "0000FF"
GREEN_LK   = "008000"
GREY       = "888888"
YELLOW     = "FFFF00"
WHITE      = "FFFFFF"
BLACK      = "000000"
LIGHT_BLUE = "D9E1F2"

# ── Helpers ───────────────────────────────────────────────────────────────────
def fl(c): return PatternFill("solid", fgColor=c)
def fnt(bold=False, italic=False, color=BLACK, sz=11):
    return Font(name="Calibri", bold=bold, italic=italic, color=color, size=sz)
def top_border(color=BLUE_HDR):
    s = Side(style="medium", color=color)
    return Border(top=s)
def cl(n): return get_column_letter(n)

# ── Layout constants ──────────────────────────────────────────────────────────
ACT_YRS = ["12/31/22", "12/31/23", "12/31/24", "12/31/25"]
EST_YRS = ["12/31/26E", "12/31/27E", "12/31/28E", "12/31/29E", "12/31/30E"]
NA, NE   = len(ACT_YRS), len(EST_YRS)
C0       = 3   # first data column index
CLAST    = C0 + NA + NE - 1   # = 11  (col K)

FMT_NUM  = '#,##0;(#,##0);-'
FMT_PCT  = '0.0%'
FMT_EPS  = '€#,##0.00'
FMT_MUL  = '0.0x'

# ── Actuals ───────────────────────────────────────────────────────────────────
# IS (€M)
REV   = [38260.60, 41182.50, 43486.80, 44052.00]
COGS  = [-10577.40,-10767.00,-11227.00,-11313.40]
GP    = [27683.20, 30415.50, 32259.80, 32738.60]
SGA   = [-19094.30,-20985.50,-22228.50,-22475.10]
RD    = [-1138.60, -1288.90, -1354.70, -1380.60]
OPEX  = [-20232.90,-22274.40,-23583.20,-23855.70]
EBIT  = [7450.30, 8141.10, 8676.60, 8882.90]
INTX  = [-70.40,  -226.70,  -373.40,  -365.80]
INTI  = [538.00,   585.70,   596.30,   519.60]
EQINC = [1.40,      0.20,    -1.30,    -5.50]
OTHNI = [-65.80,  -49.30,   -28.90,   -29.00]
EBTEX = [7853.50, 8451.00,  8869.30,  9002.20]
UNUSU = [-241.50,  -449.90,  -437.70,  -505.40]
EBT   = [7612.00, 8001.10,  8431.60,  8496.80]
TAX   = [-1899.40,-1810.60, -2015.10, -2363.10]
EACO  = [5712.60, 6190.50,  6416.50,  6133.70]
MI    = [-6.00,    -6.50,    -7.80,    -6.50]
NI    = [5706.60, 6184.00,  6408.70,  6127.20]
DSHR  = [537.66,   537.02,   536.08,   535.37]
BSHR  = [535.90,   535.43,   534.48,   533.72]
EPS_D = [10.61,    11.52,    11.95,    11.44]
EPS_B = [10.65,    11.55,    11.99,    11.48]
DPS   = [6.00,     6.60,     7.00,     7.20]
SBC   = [169.00,   168.50,   239.10,   248.30]
DA    = [1307.70, 1254.90,  1439.90,  1494.10]
EBITDA= [8330.00, 8970.00,  9642.00,  9894.50]
TAXR  = [0.250,    0.226,    0.239,    0.278]

# BS (€M)
CASH   = [2617.70, 4288.10, 4052.30, 9865.00]
STI    = [23.10, 0, 0, 0]
TCAST  = [2640.80, 4288.10, 4052.30, 9865.00]
AR     = [4755.50, 5092.70, 5601.80, 5500.30]
OREC   = [1005.50, 1056.00,  902.30,  897.90]
TREC   = [5761.00, 6148.70, 6504.10, 6398.20]
INV    = [4079.40, 4482.40, 4630.10, 4543.30]
PREP   = [617.30,  655.80,  522.50,  599.40]
OTHCA  = [951.10,  750.40,  764.50,  792.60]
TCA    = [14049.60,16325.40,16473.50,22198.50]
GPPE   = [11798.80,12698.20,13386.90,13136.10]
ADEP   = [-6834.40,-7138.10,-7421.70,-7261.00]
NPPE   = [4964.40, 5560.10, 5965.20, 5875.10]
LTINV  = [11132.20,11158.20,14337.20,12518.40]
GW     = [11717.70,13102.60,13382.00,14469.70]
OINTH  = [3640.10, 4287.10, 4594.80, 5072.00]
LNREC  = [139.00,  171.00,  156.00,  184.10]
DTAL   = [801.10,  921.20,  973.30,  962.70]
OLTASS = [400.10,  329.50,  471.40,  540.70]
TASS   = [46844.20,51855.10,56353.40,61821.20]
AP     = [6345.60, 6347.00, 6468.50, 6727.70]
ACCR   = [1853.30, 2176.50, 2161.00, 2177.20]
STB    = [876.90,   67.00,   36.30,   54.10]
CLTD   = [135.90, 2024.50, 1345.00, 1994.80]
CLCO   = [407.00,  459.80,  468.60,  434.50]
ITP    = [264.20,  208.10,  275.10,  254.80]
OTHCL  = [3836.70, 3616.80, 3881.50, 3723.40]
TCL    = [13719.60,14899.70,14636.00,15366.50]
LTD    = [3017.60, 4746.70, 5187.10, 8069.70]
CLEASE = [1213.50, 1394.20, 1458.00, 1362.90]
PENS   = [457.90,  562.00,  668.90,  684.00]
DTL    = [905.60,  846.60,  964.50, 1013.20]
OTHNCL = [343.50,  324.30,  301.10,  321.10]
TLIAB  = [19657.70,22773.50,23215.60,26817.40]
CS     = [107.00,  106.90,  106.90,  106.80]
APIC   = [3368.70, 3370.20, 3444.30, 3509.80]
RE     = [17382.20,6184.00, 6408.70, 6127.20]
AOCI   = [6320.60,19413.20,23173.40,25209.40]
TCE    = [27178.50,29074.30,33133.30,34953.20]
MIE    = [8.00,     7.30,    4.50,   50.60]
TEQ    = [27186.50,29081.60,33137.80,35003.80]
TLE    = [46844.20,51855.10,56353.40,61821.20]

# CFS (€M)
CFNI   = [5706.60, 6184.00, 6408.70, 6127.20]
CFDA   = [1307.70, 1254.90, 1439.90, 1494.10]
CFADC  = [166.50,  155.00,  145.20,  158.30]
CFMI   = [6.00,    6.50,    7.80,    6.50]
CFGLA  = [7.60,    6.90,   15.20,   37.80]
CFAWR  = [0,      19.80,    1.60,    0]
CFEQI  = [-0.50,  -0.20,    2.90,    6.70]
CFSBC  = [169.00, 168.50,  239.10,  248.30]
CFOTH  = [-73.30, 204.10,  252.20,  250.60]
CFAR   = [-717.60,-427.30, -506.70, -152.60]
CFINV  = [-865.40,-438.30, -121.20, -141.90]
CFAP   = [247.90, 138.80,   78.80,  538.30]
CFOWC  = [323.80, 331.90,  322.50,   83.60]
CFO    = [6278.30, 7604.60, 8286.00, 8656.90]
CFCAP  = [-1343.20,-1488.70,-1641.70,-1495.30]
CFSPPE = [9.20,   12.80,   13.60,    5.70]
CFACQ  = [-746.90,-2497.20,-148.90,-2426.60]
CFIINV = [-142.80,-170.70,-1927.00, 2509.00]
CFOIV  = [-0.10,   0.10,    0.10,    0.10]
CFI    = [-2223.80,-4143.70,-3703.90,-1407.10]
CFDI   = [3019.90, 3567.10, 1529.40, 4057.30]
CFDR   = [-4010.70,-1254.30,-2258.10,-1085.90]
CFSI   = [103.20,   1.50,   69.90,   65.60]
CFBB   = [-502.30,-503.30, -497.50, -501.50]
CFDIV  = [-2689.90,-3425.60,-3614.90,-3917.00]
CFOFIN = [-0.10,  0,       -13.90,  -16.30]
CFF    = [-4079.90,-1614.60,-4785.10,-1397.80]
CFFX   = [-70.70,-175.90,  -32.80,  -39.20]
CFNC   = [-96.10,1670.40, -235.80, 5812.70]
CFBEG  = [2713.80,2617.70, 4288.10, 4052.30]
CFEND  = [2617.70,4288.10, 4052.30, 9865.00]
CFFCF  = [4935.10,6115.90, 6644.30, 7161.60]

# ── Assumption values (forward estimates) ─────────────────────────────────────
# Based on L'Oréal historical trajectory (conservative/consensus)
A_REV_G = [0.065,  0.060,  0.055,  0.050,  0.050]   # Revenue growth %
A_GM    = [0.745,  0.748,  0.750,  0.752,  0.755]   # Gross margin %
A_SGA   = [0.510,  0.510,  0.508,  0.506,  0.505]   # SG&A % of rev
A_RD    = [0.0312, 0.0311, 0.0310, 0.0305, 0.0300]  # R&D % of rev
A_INTX  = [-380,   -390,   -395,   -400,   -405]    # Interest exp €M
A_INTI  = [490,    460,    440,    420,    400]      # Interest inc €M
A_OTHNI = [-29,    -29,    -28,    -28,    -27]      # Other non-op €M
A_TAXR  = [0.255,  0.255,  0.250,  0.250,  0.250]   # Tax rate %
A_MI    = [-7.0,   -7.5,   -7.5,   -8.0,   -8.0]   # Minority int €M
A_DSHR  = [534.5,  533.5,  532.5,  531.5,  530.5]  # Diluted shares M
A_DPS_G = [0.060,  0.055,  0.055,  0.050,  0.050]  # DPS growth %
A_SBC   = [255,    262,    270,    278,    285]      # SBC €M
A_DA    = [1550,   1610,   1670,   1735,   1800]    # D&A €M
A_DSO   = [46.0,   45.5,   45.0,   44.5,   44.0]   # Days sales out
A_DI    = [145,    143,    141,    139,    137]      # Days inventory
A_DPO   = [215,    215,    216,    216,    217]      # Days payable out
A_LTINV = [13000,  13500,  14000,  14500,  15000]   # LT investments €M
A_GW_ADD= [500,    300,    300,    300,    300]      # Goodwill additions
A_OI_ADD= [400,    400,    400,    400,    400]      # Other intangibles add
A_CAPX  = [0.034,  0.035,  0.035,  0.036,  0.036]  # Capex % of rev
A_STKI  = [70,     70,     70,     70,     70]      # Stock issuances €M
A_BBK   = [-510,  -520,   -525,   -530,   -535]     # Buybacks €M
A_ACQ   = [-400,  -400,   -400,   -400,   -400]     # Acquisitions €M
A_FX    = [-40,   -40,    -40,    -40,    -40]      # FX on cash €M

# ── Row-position tracking dictionaries (filled by build_* functions) ──────────
ASS_R = {}   # assumption_name -> row number in Assumptions tab
IS_R  = {}   # metric_name -> row number in IS tab
BS_R  = {}   # metric_name -> row number in BS tab
CFS_R = {}   # metric_name -> row number in CFS tab

def ass_ref(name, est_idx):
    """Return formula string 'Assumptions!Xr' for the est_idx-th estimate column."""
    r = ASS_R[name]
    c = cl(C0 + NA + est_idx)
    return f"=Assumptions!{c}{r}"

def is_ref(name, col_idx):
    """Return IS!Xr reference for col_idx (0-indexed from C0)."""
    r = IS_R[name]
    c = cl(C0 + col_idx)
    return f"IS!{c}{r}"

def bs_ref(name, col_idx):
    r = BS_R[name]
    c = cl(C0 + col_idx)
    return f"BS!{c}{r}"

def cfs_ref(name, col_idx):
    r = CFS_R[name]
    c = cl(C0 + col_idx)
    return f"CFS!{c}{r}"

# ── Generic tab setup ─────────────────────────────────────────────────────────
def setup_ws(ws, stmt_name, tab_hex):
    ws.sheet_properties.tabColor = tab_hex
    ws.column_dimensions["A"].width = 1
    ws.column_dimensions["B"].width = 35
    for i in range(NA + NE):
        ws.column_dimensions[cl(C0 + i)].width = 14
    ws.row_dimensions[1].height = 4
    # Title row 2
    title = f"L'Oréal SA (OR) — {stmt_name} (€M)"
    c = ws.cell(row=2, column=2, value=title)
    c.fill = fl(NAVY); c.font = fnt(bold=True, color=WHITE, sz=13)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=CLAST)
    ws.row_dimensions[2].height = 22
    # Header row 3
    ws.row_dimensions[3].height = 16
    ws.cell(row=3, column=2).fill = fl(NAVY)
    for i, yr in enumerate(ACT_YRS):
        c = ws.cell(row=3, column=C0 + i, value=yr)
        c.fill = fl(BLUE_HDR); c.font = fnt(bold=True, color=WHITE)
        c.alignment = Alignment(horizontal="center")
    for i, yr in enumerate(EST_YRS):
        c = ws.cell(row=3, column=C0 + NA + i, value=yr)
        c.fill = fl(GREEN_HDR); c.font = fnt(bold=True, color=WHITE)
        c.alignment = Alignment(horizontal="center")
    return 4  # next row

def section(ws, row, label):
    c = ws.cell(row=row, column=2, value=label)
    c.fill = fl(SECT_BG); c.font = fnt(bold=True, color=NAVY)
    c.alignment = Alignment(horizontal="left")
    for col in range(C0, CLAST + 1):
        ws.cell(row=row, column=col).fill = fl(SECT_BG)
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=CLAST)
    ws.row_dimensions[row].height = 14
    return row + 1

def row_write(ws, row, label, vals, fmt=FMT_NUM, bold=False, italic=False,
              indent=0, is_total=False, bg=None, helper=False, num_act=NA):
    """Write a data row; vals is list of NA+NE items (numbers or formula strings)."""
    lbl = "  " * indent + label
    c = ws.cell(row=row, column=2, value=lbl)
    c.font = fnt(bold=bold and not helper, italic=helper, color=BLACK)
    c.alignment = Alignment(horizontal="left")
    if bg: c.fill = fl(bg)
    bt = top_border(BLUE_HDR) if is_total else None
    for i, v in enumerate(vals):
        col = C0 + i
        cell = ws.cell(row=row, column=col)
        if v is not None and v != "":
            cell.value = v
        cell.number_format = fmt
        is_act = i < num_act
        if helper:
            cell.font = fnt(italic=True, color=GREY)
        elif is_act:
            cell.font = fnt(bold=bold, color=BLUE_IN)
        else:
            v_str = str(v) if v is not None else ""
            if v_str.startswith("=") and any(s + "!" in v_str for s in ["IS","BS","CFS"]) and "Assumptions" not in v_str:
                cell.font = fnt(bold=bold, color=GREEN_LK)
            else:
                cell.font = fnt(bold=bold, color=BLACK)
        cell.alignment = Alignment(horizontal="right")
        if bg: cell.fill = fl(bg)
        if bt: cell.border = bt
    return row + 1

# ── BUILD ASSUMPTIONS TAB ─────────────────────────────────────────────────────
def build_assumptions(wb):
    ws = wb.create_sheet("Assumptions")
    ws.sheet_properties.tabColor = PURPLE_TAB
    ws.column_dimensions["A"].width = 1
    ws.column_dimensions["B"].width = 38
    # Estimate cols only (C-G)
    for i in range(NE):
        ws.column_dimensions[cl(C0 + i)].width = 14
    # Notes column
    ws.column_dimensions[cl(C0 + NE)].width = 40
    ws.row_dimensions[1].height = 4
    # Title
    title = "L'Oréal SA (OR) — Assumptions (€M)"
    c = ws.cell(row=2, column=2, value=title)
    c.fill = fl(NAVY); c.font = fnt(bold=True, color=WHITE, sz=13)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=C0 + NE)
    ws.row_dimensions[2].height = 22
    # Headers row 3 (estimate years only)
    ws.row_dimensions[3].height = 16
    ws.cell(row=3, column=2).fill = fl(NAVY)
    for i, yr in enumerate(EST_YRS):
        c = ws.cell(row=3, column=C0 + i, value=yr)
        c.fill = fl(GREEN_HDR); c.font = fnt(bold=True, color=WHITE)
        c.alignment = Alignment(horizontal="center")
    nc = ws.cell(row=3, column=C0 + NE, value="Notes / Source")
    nc.fill = fl(NAVY); nc.font = fnt(bold=True, color=WHITE)

    def ass_row(row, label, vals, fmt=FMT_NUM, note=""):
        c = ws.cell(row=row, column=2, value=label)
        c.font = fnt(color=BLACK); c.alignment = Alignment(horizontal="left")
        for i, v in enumerate(vals):
            cell = ws.cell(row=row, column=C0 + i, value=v)
            cell.number_format = fmt
            cell.font = fnt(color=BLUE_IN)
            cell.alignment = Alignment(horizontal="right")
        nc = ws.cell(row=row, column=C0 + NE, value=note)
        nc.font = fnt(italic=True, color=GREY, sz=9)

    r = 4
    # IS Drivers
    c = ws.cell(row=r, column=2, value="INCOME STATEMENT DRIVERS")
    c.fill = fl(SECT_BG); c.font = fnt(bold=True, color=NAVY)
    for col in range(C0, C0 + NE + 1): ws.cell(row=r, column=col).fill = fl(SECT_BG)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=C0 + NE)
    r += 1

    ASS_R["rev_growth"] = r
    ass_row(r, "Revenue Growth %", A_REV_G, FMT_PCT,
            "Based on L'Oréal 5-yr avg ~7%; moderating to ~5% by 2030E"); r += 1
    ASS_R["gross_margin"] = r
    ass_row(r, "Gross Margin %", A_GM, FMT_PCT,
            "Trending up from 74.3% actual; premium mix + pricing"); r += 1
    ASS_R["sga_pct"] = r
    ass_row(r, "SG&A % of Revenue", A_SGA, FMT_PCT,
            "Stable; marketing leverage offset by investments"); r += 1
    ASS_R["rd_pct"] = r
    ass_row(r, "R&D % of Revenue", A_RD, FMT_PCT,
            "~3.1% of revenue; slight efficiency gain"); r += 1
    ASS_R["int_exp"] = r
    ass_row(r, "Interest Expense (€M)", A_INTX, FMT_NUM,
            "Based on 2025 actual €366M; slight increase with higher debt"); r += 1
    ASS_R["int_inc"] = r
    ass_row(r, "Interest Income (€M)", A_INTI, FMT_NUM,
            "Declining from elevated 2025 cash position"); r += 1
    ASS_R["oth_nonop"] = r
    ass_row(r, "Other Non-Op Income (€M)", A_OTHNI, FMT_NUM,
            "Stable at ~€29M expense based on history"); r += 1
    ASS_R["tax_rate"] = r
    ass_row(r, "Effective Tax Rate %", A_TAXR, FMT_PCT,
            "Normalizing toward 25% from elevated 2025 (27.8%)"); r += 1
    ASS_R["mi"] = r
    ass_row(r, "Minority Interest (€M)", A_MI, FMT_NUM,
            "Stable, modest growth reflecting NCI earnings"); r += 1
    ASS_R["dil_shares"] = r
    ass_row(r, "Diluted Shares (M)", A_DSHR, FMT_NUM,
            "Gradual decline ~1M/yr from buyback program"); r += 1
    ASS_R["dps_growth"] = r
    ass_row(r, "Dividend Per Share Growth %", A_DPS_G, FMT_PCT,
            "L'Oréal has raised DPS every year; moderating growth"); r += 1
    ASS_R["sbc"] = r
    ass_row(r, "Stock-Based Compensation (€M)", A_SBC, FMT_NUM,
            "Growing with headcount and options; ~0.6% of rev"); r += 1
    ASS_R["da"] = r
    ass_row(r, "Depreciation & Amortization (€M)", A_DA, FMT_NUM,
            "Growing with capex investments; ~3.5% of revenue"); r += 1
    r += 1  # blank

    # BS Drivers
    c = ws.cell(row=r, column=2, value="BALANCE SHEET DRIVERS")
    c.fill = fl(SECT_BG); c.font = fnt(bold=True, color=NAVY)
    for col in range(C0, C0 + NE + 1): ws.cell(row=r, column=col).fill = fl(SECT_BG)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=C0 + NE)
    r += 1

    ASS_R["dso"] = r
    ass_row(r, "Days Sales Outstanding (days)", A_DSO, FMT_NUM,
            "2025 actual: ~45.6 days; modest improvement expected"); r += 1
    ASS_R["di"] = r
    ass_row(r, "Days Inventory Outstanding (days)", A_DI, FMT_NUM,
            "2025 actual: ~146 days; slight improvement from efficiency"); r += 1
    ASS_R["dpo"] = r
    ass_row(r, "Days Payable Outstanding (days)", A_DPO, FMT_NUM,
            "2025 actual: ~217 days; stable / slight increase"); r += 1
    ASS_R["lt_inv"] = r
    ass_row(r, "Long-term Investments (€M, fixed)", A_LTINV, FMT_NUM,
            "Strategic equity investments; gradual growth"); r += 1
    ASS_R["gw_add"] = r
    ass_row(r, "Goodwill Net Additions (€M)", A_GW_ADD, FMT_NUM,
            "Annual M&A contribution; lower than 2022-2025 avg"); r += 1
    ASS_R["oi_add"] = r
    ass_row(r, "Other Intangibles Net Additions (€M)", A_OI_ADD, FMT_NUM,
            "Acquisitions + amortization net"); r += 1
    r += 1

    # CFS Drivers
    c = ws.cell(row=r, column=2, value="CASH FLOW DRIVERS")
    c.fill = fl(SECT_BG); c.font = fnt(bold=True, color=NAVY)
    for col in range(C0, C0 + NE + 1): ws.cell(row=r, column=col).fill = fl(SECT_BG)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=C0 + NE)
    r += 1

    ASS_R["capex_pct"] = r
    ass_row(r, "Capital Expenditure % of Revenue", A_CAPX, FMT_PCT,
            "2025 actual ~3.4%; stable investment profile"); r += 1
    ASS_R["stk_iss"] = r
    ass_row(r, "Stock Issuances (€M)", A_STKI, FMT_NUM,
            "Options exercise proceeds; ~€70M/yr"); r += 1
    ASS_R["buyback"] = r
    ass_row(r, "Share Buybacks (€M)", A_BBK, FMT_NUM,
            "~€500M/yr buyback programme; modest increase"); r += 1
    ASS_R["acquis"] = r
    ass_row(r, "Acquisitions (€M)", A_ACQ, FMT_NUM,
            "Conservative M&A; bolt-on acquisitions assumed"); r += 1
    ASS_R["fx_effect"] = r
    ass_row(r, "FX Effect on Cash (€M)", A_FX, FMT_NUM,
            "Based on ~€40M annual FX drag historically"); r += 1
    r += 2

    # Model Notes
    c = ws.cell(row=r, column=2, value="MODEL NOTES")
    c.fill = fl(SECT_BG); c.font = fnt(bold=True, color=NAVY)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=C0 + NE)
    for col in range(C0, C0 + NE + 1): ws.cell(row=r, column=col).fill = fl(SECT_BG)
    r += 1
    notes = [
        "Blue (#0000FF) — hardcoded inputs (actuals & all Assumptions cells)",
        "Black (#000000) — calculated formulas on IS, BS, CFS, DCF tabs",
        "Green (#008000) — cross-sheet links pulling from another tab",
        "Grey italic — helper / ratio rows (margins, growth %)",
        "All values in €M (millions of euros) unless otherwise noted",
        "Actuals sourced from TIKR Terminal (L'Oréal SA, cid=874249)",
        "Forward assumptions are analyst-consensus-calibrated; edit Assumptions tab to adjust",
    ]
    for note in notes:
        nc = ws.cell(row=r, column=2, value=note)
        nc.font = fnt(italic=True, color=GREY, sz=9)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=C0 + NE)
        r += 1

    return ws

# ── BUILD IS TAB ──────────────────────────────────────────────────────────────
def build_is(wb):
    ws = wb.create_sheet("IS")
    ws.sheet_properties.tabColor = BLUE_HDR
    r = setup_ws(ws, "Income Statement", BLUE_HDR)

    def est_vals(act_vals, formulas):
        return act_vals + formulas

    # Estimate formulas for revenue
    rev_est = []
    for ei in range(NE):
        prev_col = cl(C0 + NA + ei - 1) if ei > 0 else cl(C0 + NA - 1)
        prev_row = "5"   # to be filled after we know the row
        rev_est.append(None)  # placeholder
    # We'll write directly below after defining row numbers

    r = section(ws, r, "REVENUE")
    IS_R["rev"] = r
    # Revenue estimate: prior col * (1 + growth)
    rev_f = []
    for ei in range(NE):
        prev_c = cl(C0 + NA + ei - 1) if ei > 0 else cl(C0 + NA - 1)
        grw = f"Assumptions!{cl(C0 + ei)}{ASS_R['rev_growth']}"
        rev_f.append(f"={prev_c}{r}*(1+{grw})")
    r = row_write(ws, r, "Total Revenues", REV + rev_f, FMT_NUM, bold=True)

    IS_R["rev_growth_h"] = r
    rev_g_f = []
    for ei in range(NE):
        prev_c = cl(C0 + NA + ei - 1) if ei > 0 else cl(C0 + NA - 1)
        curr_c = cl(C0 + NA + ei)
        rev_g_f.append(f"={curr_c}{r-1}/{prev_c}{r-1}-1")
    rev_g_act = [None, 0.076, 0.056, 0.013]
    r = row_write(ws, r, "  YoY Growth %", rev_g_act + rev_g_f, FMT_PCT, helper=True)

    r = section(ws, r, "COST OF REVENUE")
    IS_R["cogs"] = r
    cogs_f = [f"=-{cl(C0+NA+ei)}{IS_R['rev']}*(1-Assumptions!{cl(C0+ei)}{ASS_R['gross_margin']})" for ei in range(NE)]
    r = row_write(ws, r, "  Cost of Goods Sold", COGS + cogs_f, FMT_NUM)

    IS_R["gp"] = r
    gp_f = [f"={cl(C0+NA+ei)}{IS_R['rev']}+{cl(C0+NA+ei)}{IS_R['cogs']}" for ei in range(NE)]
    r = row_write(ws, r, "Gross Profit", GP + gp_f, FMT_NUM, bold=True, is_total=True)

    IS_R["gm_h"] = r
    gm_f = [f"={cl(C0+NA+ei)}{IS_R['gp']}/{cl(C0+NA+ei)}{IS_R['rev']}" for ei in range(NE)]
    gm_act = [0.724, 0.739, 0.742, 0.743]
    r = row_write(ws, r, "  Gross Margin %", gm_act + gm_f, FMT_PCT, helper=True)

    r = section(ws, r, "OPERATING EXPENSES")
    IS_R["sga"] = r
    sga_f = [f"=-{cl(C0+NA+ei)}{IS_R['rev']}*Assumptions!{cl(C0+ei)}{ASS_R['sga_pct']}" for ei in range(NE)]
    r = row_write(ws, r, "  Selling General & Admin Expenses", SGA + sga_f, FMT_NUM, indent=1)

    IS_R["rd"] = r
    rd_f = [f"=-{cl(C0+NA+ei)}{IS_R['rev']}*Assumptions!{cl(C0+ei)}{ASS_R['rd_pct']}" for ei in range(NE)]
    r = row_write(ws, r, "  R&D Expenses", RD + rd_f, FMT_NUM, indent=1)

    IS_R["opex"] = r
    opex_f = [f"={cl(C0+NA+ei)}{IS_R['sga']}+{cl(C0+NA+ei)}{IS_R['rd']}" for ei in range(NE)]
    r = row_write(ws, r, "Total Operating Expenses", OPEX + opex_f, FMT_NUM, bold=True, is_total=True)

    r = section(ws, r, "OPERATING INCOME")
    IS_R["ebit"] = r
    ebit_f = [f"={cl(C0+NA+ei)}{IS_R['gp']}+{cl(C0+NA+ei)}{IS_R['opex']}" for ei in range(NE)]
    r = row_write(ws, r, "Operating Income (EBIT)", EBIT + ebit_f, FMT_NUM, bold=True, is_total=True)

    IS_R["ebit_m_h"] = r
    ebitm_f = [f"={cl(C0+NA+ei)}{IS_R['ebit']}/{cl(C0+NA+ei)}{IS_R['rev']}" for ei in range(NE)]
    ebitm_act = [0.195, 0.198, 0.200, 0.202]
    r = row_write(ws, r, "  EBIT Margin %", ebitm_act + ebitm_f, FMT_PCT, helper=True)

    r = section(ws, r, "OTHER INCOME / EXPENSE")
    IS_R["int_exp"] = r
    intx_f = [f"=Assumptions!{cl(C0+ei)}{ASS_R['int_exp']}" for ei in range(NE)]
    r = row_write(ws, r, "  Interest Expense", INTX + intx_f, FMT_NUM, indent=1)

    IS_R["int_inc"] = r
    inti_f = [f"=Assumptions!{cl(C0+ei)}{ASS_R['int_inc']}" for ei in range(NE)]
    r = row_write(ws, r, "  Interest Income", INTI + inti_f, FMT_NUM, indent=1)

    IS_R["eq_inc"] = r
    r = row_write(ws, r, "  Income (Loss) on Equity Investments",
                  EQINC + [-5]*NE, FMT_NUM, indent=1)

    IS_R["oth_ni"] = r
    othni_f = [f"=Assumptions!{cl(C0+ei)}{ASS_R['oth_nonop']}" for ei in range(NE)]
    r = row_write(ws, r, "  Other Non-Operating Income (Expenses)", OTHNI + othni_f, FMT_NUM, indent=1)

    IS_R["ebt_ex"] = r
    ebtex_f = [
        f"={cl(C0+NA+ei)}{IS_R['ebit']}+{cl(C0+NA+ei)}{IS_R['int_exp']}+"
        f"{cl(C0+NA+ei)}{IS_R['int_inc']}+{cl(C0+NA+ei)}{IS_R['eq_inc']}+"
        f"{cl(C0+NA+ei)}{IS_R['oth_ni']}" for ei in range(NE)]
    r = row_write(ws, r, "EBT Excl. Unusual Items", EBTEX + ebtex_f, FMT_NUM, bold=True, is_total=True)

    IS_R["unusual"] = r
    r = row_write(ws, r, "  Unusual / Non-Recurring Items", UNUSU + [0]*NE, FMT_NUM, indent=1)

    IS_R["ebt"] = r
    ebt_f = [f"={cl(C0+NA+ei)}{IS_R['ebt_ex']}+{cl(C0+NA+ei)}{IS_R['unusual']}" for ei in range(NE)]
    r = row_write(ws, r, "EBT Incl. Unusual Items", EBT + ebt_f, FMT_NUM, bold=True, is_total=True)

    IS_R["tax"] = r
    tax_f = [f"=-{cl(C0+NA+ei)}{IS_R['ebt']}*Assumptions!{cl(C0+ei)}{ASS_R['tax_rate']}" for ei in range(NE)]
    r = row_write(ws, r, "  Income Tax Expense", TAX + tax_f, FMT_NUM, indent=1)

    IS_R["eac"] = r
    eac_f = [f"={cl(C0+NA+ei)}{IS_R['ebt']}+{cl(C0+NA+ei)}{IS_R['tax']}" for ei in range(NE)]
    r = row_write(ws, r, "Earnings from Continuing Operations", EACO + eac_f, FMT_NUM, bold=True, is_total=True)

    IS_R["mi"] = r
    mi_f = [f"=Assumptions!{cl(C0+ei)}{ASS_R['mi']}" for ei in range(NE)]
    r = row_write(ws, r, "  Minority Interest", MI + mi_f, FMT_NUM, indent=1)

    IS_R["ni"] = r
    ni_f = [f"={cl(C0+NA+ei)}{IS_R['eac']}+{cl(C0+NA+ei)}{IS_R['mi']}" for ei in range(NE)]
    r = row_write(ws, r, "Net Income", NI + ni_f, FMT_NUM, bold=True, is_total=True)

    IS_R["ni_m_h"] = r
    nim_f = [f"={cl(C0+NA+ei)}{IS_R['ni']}/{cl(C0+NA+ei)}{IS_R['rev']}" for ei in range(NE)]
    nim_act = [0.149, 0.150, 0.147, 0.139]
    r = row_write(ws, r, "  Net Income Margin %", nim_act + nim_f, FMT_PCT, helper=True)

    r = section(ws, r, "PER SHARE DATA")
    IS_R["dil_sh"] = r
    dshr_f = [f"=Assumptions!{cl(C0+ei)}{ASS_R['dil_shares']}" for ei in range(NE)]
    r = row_write(ws, r, "Weighted Avg. Diluted Shares (M)", DSHR + dshr_f, FMT_NUM)

    IS_R["bas_sh"] = r
    bshr_f = [f"=Assumptions!{cl(C0+ei)}{ASS_R['dil_shares']}-1" for ei in range(NE)]
    r = row_write(ws, r, "Weighted Avg. Basic Shares (M)", BSHR + bshr_f, FMT_NUM)

    IS_R["eps_dil"] = r
    epsd_f = [f"={cl(C0+NA+ei)}{IS_R['ni']}/{cl(C0+NA+ei)}{IS_R['dil_sh']}" for ei in range(NE)]
    r = row_write(ws, r, "EPS (Diluted)", EPS_D + epsd_f, FMT_EPS, bold=True)

    IS_R["eps_bas"] = r
    epsb_f = [f"={cl(C0+NA+ei)}{IS_R['ni']}/{cl(C0+NA+ei)}{IS_R['bas_sh']}" for ei in range(NE)]
    r = row_write(ws, r, "EPS (Basic)", EPS_B + epsb_f, FMT_EPS)

    IS_R["dps"] = r
    # DPS est: prior × (1 + growth)
    dps_f = []
    for ei in range(NE):
        prev_c = cl(C0 + NA + ei - 1) if ei > 0 else cl(C0 + NA - 1)
        grw = f"Assumptions!{cl(C0+ei)}{ASS_R['dps_growth']}"
        dps_f.append(f"={prev_c}{r}*(1+{grw})")
    r = row_write(ws, r, "Dividends Per Share (€)", DPS + dps_f, FMT_EPS)

    r = section(ws, r, "SUPPLEMENTAL METRICS")
    IS_R["sbc"] = r
    sbc_f = [f"=Assumptions!{cl(C0+ei)}{ASS_R['sbc']}" for ei in range(NE)]
    r = row_write(ws, r, "  Stock-Based Compensation", SBC + sbc_f, FMT_NUM, indent=1)

    IS_R["da"] = r
    da_f = [f"=Assumptions!{cl(C0+ei)}{ASS_R['da']}" for ei in range(NE)]
    r = row_write(ws, r, "  Depreciation & Amortization", DA + da_f, FMT_NUM, indent=1)

    IS_R["ebitda"] = r
    ebitda_f = [f"={cl(C0+NA+ei)}{IS_R['ebit']}+{cl(C0+NA+ei)}{IS_R['da']}" for ei in range(NE)]
    r = row_write(ws, r, "EBITDA", EBITDA + ebitda_f, FMT_NUM, bold=True, is_total=True)

    IS_R["ebitda_m_h"] = r
    ebitdm_f = [f"={cl(C0+NA+ei)}{IS_R['ebitda']}/{cl(C0+NA+ei)}{IS_R['rev']}" for ei in range(NE)]
    ebitdm_act = [0.218, 0.218, 0.222, 0.225]
    r = row_write(ws, r, "  EBITDA Margin %", ebitdm_act + ebitdm_f, FMT_PCT, helper=True)

    return ws

# ── BUILD BS TAB ──────────────────────────────────────────────────────────────
def build_bs(wb):
    ws = wb.create_sheet("BS")
    ws.sheet_properties.tabColor = GREEN_HDR
    r = setup_ws(ws, "Balance Sheet", GREEN_HDR)

    rev_r  = IS_R["rev"]
    cogs_r = IS_R["cogs"]
    da_r   = IS_R["da"]
    ni_r   = IS_R["ni"]
    dps_r  = IS_R["dps"]
    dshr_r = IS_R["dil_sh"]
    sbc_r  = IS_R["sbc"]

    r = section(ws, r, "ASSETS — CURRENT ASSETS")
    # Cash row: actuals blue; estimates = green link to CFS ending cash (set later)
    BS_R["cash"] = r
    cash_est = [None]*NE   # will be replaced by CFS link after CFS is built
    r = row_write(ws, r, "  Cash and Equivalents", CASH + cash_est, FMT_NUM, indent=1)

    BS_R["sti"] = r
    r = row_write(ws, r, "  Short-Term Investments", STI + [0]*NE, FMT_NUM, indent=1)

    BS_R["tot_cash"] = r
    tc_f = [f"={cl(C0+NA+ei)}{BS_R['cash']}+{cl(C0+NA+ei)}{BS_R['sti']}" for ei in range(NE)]
    r = row_write(ws, r, "  Total Cash & Short-Term Investments", TCAST + tc_f, FMT_NUM, indent=1)

    BS_R["ar"] = r
    # AR = Revenue * DSO / 365
    ar_f = [f"=IS!{cl(C0+NA+ei)}{rev_r}*Assumptions!{cl(C0+ei)}{ASS_R['dso']}/365" for ei in range(NE)]
    r = row_write(ws, r, "  Accounts Receivable", AR + ar_f, FMT_NUM, indent=1)

    BS_R["oth_rec"] = r
    r = row_write(ws, r, "  Other Receivables", OREC + [890]*NE, FMT_NUM, indent=1)

    BS_R["tot_rec"] = r
    tr_f = [f"={cl(C0+NA+ei)}{BS_R['ar']}+{cl(C0+NA+ei)}{BS_R['oth_rec']}" for ei in range(NE)]
    r = row_write(ws, r, "  Total Receivables", TREC + tr_f, FMT_NUM, indent=1)

    BS_R["inv"] = r
    # Inventory = COGS * DI / 365  (use absolute value of COGS)
    inv_f = [f"=-IS!{cl(C0+NA+ei)}{cogs_r}*Assumptions!{cl(C0+ei)}{ASS_R['di']}/365" for ei in range(NE)]
    r = row_write(ws, r, "  Inventory", INV + inv_f, FMT_NUM, indent=1)

    BS_R["prep"] = r
    prep_f = [f"=IS!{cl(C0+NA+ei)}{rev_r}*0.013" for ei in range(NE)]
    r = row_write(ws, r, "  Prepaid Expenses", PREP + prep_f, FMT_NUM, indent=1)

    BS_R["oth_ca"] = r
    r = row_write(ws, r, "  Other Current Assets", OTHCA + [800]*NE, FMT_NUM, indent=1)

    BS_R["tca"] = r
    tca_f = [f"={cl(C0+NA+ei)}{BS_R['tot_cash']}+{cl(C0+NA+ei)}{BS_R['tot_rec']}+"
             f"{cl(C0+NA+ei)}{BS_R['inv']}+{cl(C0+NA+ei)}{BS_R['prep']}+"
             f"{cl(C0+NA+ei)}{BS_R['oth_ca']}" for ei in range(NE)]
    r = row_write(ws, r, "Total Current Assets", TCA + tca_f, FMT_NUM, bold=True, is_total=True)

    r = section(ws, r, "ASSETS — NON-CURRENT ASSETS")
    BS_R["nppe"] = r
    # Net PP&E = prior + capex - D&A
    nppe_f = []
    for ei in range(NE):
        prev_c = cl(C0+NA+ei-1) if ei > 0 else cl(C0+NA-1)
        cap_c = f"IS!{cl(C0+NA+ei)}{rev_r}*Assumptions!{cl(C0+ei)}{ASS_R['capex_pct']}"
        da_c  = f"IS!{cl(C0+NA+ei)}{da_r}"
        nppe_f.append(f"={prev_c}{BS_R['nppe']}+{cap_c}-{da_c}")
    r = row_write(ws, r, "  Net Property, Plant & Equipment", NPPE + nppe_f, FMT_NUM, indent=1)

    BS_R["lt_inv"] = r
    ltinv_f = [f"=Assumptions!{cl(C0+ei)}{ASS_R['lt_inv']}" for ei in range(NE)]
    r = row_write(ws, r, "  Long-Term Investments", LTINV + ltinv_f, FMT_NUM, indent=1)

    BS_R["gw"] = r
    gw_f = []
    for ei in range(NE):
        prev_c = cl(C0+NA+ei-1) if ei > 0 else cl(C0+NA-1)
        gw_f.append(f"={prev_c}{BS_R['gw']}+Assumptions!{cl(C0+ei)}{ASS_R['gw_add']}")
    r = row_write(ws, r, "  Goodwill", GW + gw_f, FMT_NUM, indent=1)

    BS_R["oth_int"] = r
    oith_f = []
    for ei in range(NE):
        prev_c = cl(C0+NA+ei-1) if ei > 0 else cl(C0+NA-1)
        oith_f.append(f"={prev_c}{BS_R['oth_int']}+Assumptions!{cl(C0+ei)}{ASS_R['oi_add']}")
    r = row_write(ws, r, "  Other Intangibles", OINTH + oith_f, FMT_NUM, indent=1)

    BS_R["dta_lt"] = r
    r = row_write(ws, r, "  Deferred Tax Assets (LT)", DTAL + [1000]*NE, FMT_NUM, indent=1)

    BS_R["oth_lta"] = r
    r = row_write(ws, r, "  Other Long-Term Assets", OLTASS + [550]*NE, FMT_NUM, indent=1)

    BS_R["tnca"] = r
    tnca_f = [f"={cl(C0+NA+ei)}{BS_R['nppe']}+{cl(C0+NA+ei)}{BS_R['lt_inv']}+"
              f"{cl(C0+NA+ei)}{BS_R['gw']}+{cl(C0+NA+ei)}{BS_R['oth_int']}+"
              f"{cl(C0+NA+ei)}{BS_R['dta_lt']}+{cl(C0+NA+ei)}{BS_R['oth_lta']}" for ei in range(NE)]
    r = row_write(ws, r, "Total Non-Current Assets", [x-y for x,y in zip(TASS,TCA)] + tnca_f, FMT_NUM, bold=True, is_total=True)

    BS_R["tass"] = r
    tass_f = [f"={cl(C0+NA+ei)}{BS_R['tca']}+{cl(C0+NA+ei)}{BS_R['tnca']}" for ei in range(NE)]
    r = row_write(ws, r, "TOTAL ASSETS", TASS + tass_f, FMT_NUM, bold=True, is_total=True, bg=GREEN_HL)

    r = section(ws, r, "LIABILITIES — CURRENT LIABILITIES")
    BS_R["ap"] = r
    # AP = COGS * DPO / 365 (abs)
    ap_f = [f"=-IS!{cl(C0+NA+ei)}{cogs_r}*Assumptions!{cl(C0+ei)}{ASS_R['dpo']}/365" for ei in range(NE)]
    r = row_write(ws, r, "  Accounts Payable", AP + ap_f, FMT_NUM, indent=1)

    BS_R["accr"] = r
    r = row_write(ws, r, "  Accrued Expenses", ACCR + [2200]*NE, FMT_NUM, indent=1)

    BS_R["stb"] = r
    r = row_write(ws, r, "  Short-Term Borrowings", STB + [55]*NE, FMT_NUM, indent=1)

    BS_R["cltd"] = r
    r = row_write(ws, r, "  Current Portion of Long-Term Debt", CLTD + [1500]*NE, FMT_NUM, indent=1)

    BS_R["clco"] = r
    r = row_write(ws, r, "  Current Portion of Capital Lease Oblig.", CLCO + [440]*NE, FMT_NUM, indent=1)

    BS_R["itp"] = r
    r = row_write(ws, r, "  Income Taxes Payable", ITP + [260]*NE, FMT_NUM, indent=1)

    BS_R["oth_cl"] = r
    r = row_write(ws, r, "  Other Current Liabilities", OTHCL + [3800]*NE, FMT_NUM, indent=1)

    BS_R["tcl"] = r
    tcl_f = [f"={cl(C0+NA+ei)}{BS_R['ap']}+{cl(C0+NA+ei)}{BS_R['accr']}+"
             f"{cl(C0+NA+ei)}{BS_R['stb']}+{cl(C0+NA+ei)}{BS_R['cltd']}+"
             f"{cl(C0+NA+ei)}{BS_R['clco']}+{cl(C0+NA+ei)}{BS_R['itp']}+"
             f"{cl(C0+NA+ei)}{BS_R['oth_cl']}" for ei in range(NE)]
    r = row_write(ws, r, "Total Current Liabilities", TCL + tcl_f, FMT_NUM, bold=True, is_total=True)

    r = section(ws, r, "LIABILITIES — NON-CURRENT LIABILITIES")
    BS_R["ltd"] = r
    r = row_write(ws, r, "  Long-Term Debt", LTD + [8000]*NE, FMT_NUM, indent=1)

    BS_R["clease_lt"] = r
    r = row_write(ws, r, "  Capital Leases (LT)", CLEASE + [1350]*NE, FMT_NUM, indent=1)

    BS_R["pension"] = r
    r = row_write(ws, r, "  Pension & Post-Retirement Benefits", PENS + [700]*NE, FMT_NUM, indent=1)

    BS_R["dtl"] = r
    r = row_write(ws, r, "  Deferred Tax Liability (LT)", DTL + [1050]*NE, FMT_NUM, indent=1)

    BS_R["oth_ncl"] = r
    r = row_write(ws, r, "  Other Non-Current Liabilities", OTHNCL + [330]*NE, FMT_NUM, indent=1)

    BS_R["tliab"] = r
    tliab_f = [f"={cl(C0+NA+ei)}{BS_R['tcl']}+{cl(C0+NA+ei)}{BS_R['ltd']}+"
               f"{cl(C0+NA+ei)}{BS_R['clease_lt']}+{cl(C0+NA+ei)}{BS_R['pension']}+"
               f"{cl(C0+NA+ei)}{BS_R['dtl']}+{cl(C0+NA+ei)}{BS_R['oth_ncl']}" for ei in range(NE)]
    r = row_write(ws, r, "Total Liabilities", TLIAB + tliab_f, FMT_NUM, bold=True, is_total=True)

    r = section(ws, r, "STOCKHOLDERS' EQUITY")
    BS_R["cs"] = r
    r = row_write(ws, r, "  Common Stock", CS + [107]*NE, FMT_NUM, indent=1)

    BS_R["apic"] = r
    # APIC = prior + SBC + stock issuances
    apic_f = []
    for ei in range(NE):
        prev_c = cl(C0+NA+ei-1) if ei > 0 else cl(C0+NA-1)
        apic_f.append(f"={prev_c}{BS_R['apic']}+IS!{cl(C0+NA+ei)}{sbc_r}+"
                      f"Assumptions!{cl(C0+ei)}{ASS_R['stk_iss']}")
    r = row_write(ws, r, "  Additional Paid-In Capital", APIC + apic_f, FMT_NUM, indent=1)

    BS_R["re"] = r
    # RE = prior + Net Income - dividends paid (DPS * shares * -1 for cash out)
    re_f = []
    for ei in range(NE):
        prev_c = cl(C0+NA+ei-1) if ei > 0 else cl(C0+NA-1)
        ni_c = f"IS!{cl(C0+NA+ei)}{ni_r}"
        div_c = f"IS!{cl(C0+NA+ei)}{dps_r}*IS!{cl(C0+NA+ei)}{dshr_r}"
        re_f.append(f"={prev_c}{BS_R['re']}+{ni_c}-{div_c}")
    r = row_write(ws, r, "  Retained Earnings / Accumulated Deficit", RE + re_f, FMT_NUM, indent=1)

    BS_R["aoci"] = r
    r = row_write(ws, r, "  AOCI / Other Comprehensive Income", AOCI + [25500]*NE, FMT_NUM, indent=1)

    BS_R["tce"] = r
    tce_f = [f"={cl(C0+NA+ei)}{BS_R['cs']}+{cl(C0+NA+ei)}{BS_R['apic']}+"
             f"{cl(C0+NA+ei)}{BS_R['re']}+{cl(C0+NA+ei)}{BS_R['aoci']}" for ei in range(NE)]
    r = row_write(ws, r, "Total Common Equity", TCE + tce_f, FMT_NUM, bold=True, is_total=True)

    BS_R["mi_eq"] = r
    r = row_write(ws, r, "  Minority Interest (NCI)", MIE + [55]*NE, FMT_NUM, indent=1)

    BS_R["teq"] = r
    teq_f = [f"={cl(C0+NA+ei)}{BS_R['tce']}+{cl(C0+NA+ei)}{BS_R['mi_eq']}" for ei in range(NE)]
    r = row_write(ws, r, "Total Equity", TEQ + teq_f, FMT_NUM, bold=True, is_total=True)

    BS_R["tle"] = r
    tle_f = [f"={cl(C0+NA+ei)}{BS_R['tliab']}+{cl(C0+NA+ei)}{BS_R['teq']}" for ei in range(NE)]
    r = row_write(ws, r, "TOTAL LIABILITIES & EQUITY", TLE + tle_f, FMT_NUM, bold=True, is_total=True, bg=GREEN_HL)

    BS_R["chk"] = r
    chk_f = [f"={cl(C0+NA+ei)}{BS_R['tass']}-{cl(C0+NA+ei)}{BS_R['tle']}" for ei in range(NE)]
    chk_act = [0]*NA
    r = row_write(ws, r, "  Balance Check (should = 0)", chk_act + chk_f, FMT_NUM, helper=True)

    return ws

# ── BUILD CFS TAB ─────────────────────────────────────────────────────────────
def build_cfs(wb):
    ws = wb.create_sheet("CFS")
    ws.sheet_properties.tabColor = BROWN_TAB
    r = setup_ws(ws, "Cash Flow Statement", BROWN_TAB)

    ni_r    = IS_R["ni"];   da_r   = IS_R["da"]
    sbc_r   = IS_R["sbc"];  rev_r  = IS_R["rev"]
    ar_r    = BS_R["ar"];   inv_r  = BS_R["inv"]
    ap_r    = BS_R["ap"];   tca_r  = BS_R["tca"]
    tcl_r   = BS_R["tcl"];  dps_r  = IS_R["dps"]
    dshr_r  = IS_R["dil_sh"]

    r = section(ws, r, "OPERATING ACTIVITIES")
    CFS_R["ni"] = r
    ni_f = [f"=IS!{cl(C0+NA+ei)}{ni_r}" for ei in range(NE)]
    r = row_write(ws, r, "  Net Income / (Loss)", CFNI + ni_f, FMT_NUM, indent=1)

    CFS_R["da"] = r
    da_f = [f"=IS!{cl(C0+NA+ei)}{da_r}" for ei in range(NE)]
    r = row_write(ws, r, "  Depreciation & Amortization", CFDA + da_f, FMT_NUM, indent=1)

    CFS_R["adc"] = r
    r = row_write(ws, r, "  Amortization of Deferred Charges", CFADC + [150]*NE, FMT_NUM, indent=1)

    CFS_R["sbc"] = r
    sbc_f = [f"=IS!{cl(C0+NA+ei)}{sbc_r}" for ei in range(NE)]
    r = row_write(ws, r, "  Stock-Based Compensation", CFSBC + sbc_f, FMT_NUM, indent=1)

    CFS_R["mi_earn"] = r
    r = row_write(ws, r, "  Minority Interest in Earnings", CFMI + [7]*NE, FMT_NUM, indent=1)

    CFS_R["gl_asset"] = r
    r = row_write(ws, r, "  Gain/Loss on Asset Sales", CFGLA + [15]*NE, FMT_NUM, indent=1)

    CFS_R["oth_ops"] = r
    r = row_write(ws, r, "  Other Operating Activities", CFOTH + [200]*NE, FMT_NUM, indent=1)

    # Working capital changes
    CFS_R["chg_ar"] = r
    # Change in AR = -(AR_t - AR_{t-1})
    chg_ar_f = []
    for ei in range(NE):
        curr_c = cl(C0+NA+ei); prev_c = cl(C0+NA+ei-1) if ei > 0 else cl(C0+NA-1)
        chg_ar_f.append(f"=-(BS!{curr_c}{ar_r}-BS!{prev_c}{ar_r})")
    r = row_write(ws, r, "  Change in Accounts Receivable", CFAR + chg_ar_f, FMT_NUM, indent=1)

    CFS_R["chg_inv"] = r
    chg_inv_f = []
    for ei in range(NE):
        curr_c = cl(C0+NA+ei); prev_c = cl(C0+NA+ei-1) if ei > 0 else cl(C0+NA-1)
        chg_inv_f.append(f"=-(BS!{curr_c}{inv_r}-BS!{prev_c}{inv_r})")
    r = row_write(ws, r, "  Change in Inventories", CFINV + chg_inv_f, FMT_NUM, indent=1)

    CFS_R["chg_ap"] = r
    chg_ap_f = []
    for ei in range(NE):
        curr_c = cl(C0+NA+ei); prev_c = cl(C0+NA+ei-1) if ei > 0 else cl(C0+NA-1)
        chg_ap_f.append(f"=BS!{curr_c}{ap_r}-BS!{prev_c}{ap_r}")
    r = row_write(ws, r, "  Change in Accounts Payable", CFAP + chg_ap_f, FMT_NUM, indent=1)

    CFS_R["chg_owc"] = r
    r = row_write(ws, r, "  Change in Other Net Working Capital", CFOWC + [100]*NE, FMT_NUM, indent=1)

    CFS_R["cfo"] = r
    cfo_f = [
        f"={cl(C0+NA+ei)}{CFS_R['ni']}+{cl(C0+NA+ei)}{CFS_R['da']}+"
        f"{cl(C0+NA+ei)}{CFS_R['adc']}+{cl(C0+NA+ei)}{CFS_R['sbc']}+"
        f"{cl(C0+NA+ei)}{CFS_R['mi_earn']}+{cl(C0+NA+ei)}{CFS_R['gl_asset']}+"
        f"{cl(C0+NA+ei)}{CFS_R['oth_ops']}+{cl(C0+NA+ei)}{CFS_R['chg_ar']}+"
        f"{cl(C0+NA+ei)}{CFS_R['chg_inv']}+{cl(C0+NA+ei)}{CFS_R['chg_ap']}+"
        f"{cl(C0+NA+ei)}{CFS_R['chg_owc']}" for ei in range(NE)]
    r = row_write(ws, r, "Net Cash from Operating Activities", CFO + cfo_f, FMT_NUM, bold=True, is_total=True)

    r = section(ws, r, "INVESTING ACTIVITIES")
    CFS_R["capex"] = r
    capex_f = [f"=-IS!{cl(C0+NA+ei)}{rev_r}*Assumptions!{cl(C0+ei)}{ASS_R['capex_pct']}" for ei in range(NE)]
    r = row_write(ws, r, "  Capital Expenditure", CFCAP + capex_f, FMT_NUM, indent=1)

    CFS_R["sale_ppe"] = r
    r = row_write(ws, r, "  Sale of PP&E", CFSPPE + [8]*NE, FMT_NUM, indent=1)

    CFS_R["acquis"] = r
    acq_f = [f"=Assumptions!{cl(C0+ei)}{ASS_R['acquis']}" for ei in range(NE)]
    r = row_write(ws, r, "  Cash Acquisitions", CFACQ + acq_f, FMT_NUM, indent=1)

    CFS_R["inv_mkt"] = r
    r = row_write(ws, r, "  Investments in Securities", CFIINV + [0]*NE, FMT_NUM, indent=1)

    CFS_R["oth_inv"] = r
    r = row_write(ws, r, "  Other Investing Activities", CFOIV + [0]*NE, FMT_NUM, indent=1)

    CFS_R["cfi"] = r
    cfi_f = [f"={cl(C0+NA+ei)}{CFS_R['capex']}+{cl(C0+NA+ei)}{CFS_R['sale_ppe']}+"
             f"{cl(C0+NA+ei)}{CFS_R['acquis']}+{cl(C0+NA+ei)}{CFS_R['inv_mkt']}+"
             f"{cl(C0+NA+ei)}{CFS_R['oth_inv']}" for ei in range(NE)]
    r = row_write(ws, r, "Net Cash from Investing Activities", CFI + cfi_f, FMT_NUM, bold=True, is_total=True)

    r = section(ws, r, "FINANCING ACTIVITIES")
    CFS_R["debt_iss"] = r
    r = row_write(ws, r, "  Total Debt Issued", CFDI + [1000]*NE, FMT_NUM, indent=1)

    CFS_R["debt_rep"] = r
    r = row_write(ws, r, "  Total Debt Repaid", CFDR + [-1000]*NE, FMT_NUM, indent=1)

    CFS_R["stk_iss"] = r
    stki_f = [f"=Assumptions!{cl(C0+ei)}{ASS_R['stk_iss']}" for ei in range(NE)]
    r = row_write(ws, r, "  Issuance of Common Stock", CFSI + stki_f, FMT_NUM, indent=1)

    CFS_R["buyback"] = r
    bbk_f = [f"=Assumptions!{cl(C0+ei)}{ASS_R['buyback']}" for ei in range(NE)]
    r = row_write(ws, r, "  Repurchase of Common Stock", CFBB + bbk_f, FMT_NUM, indent=1)

    CFS_R["divid"] = r
    # Dividends = -(DPS * diluted shares)
    div_f = [f"=-IS!{cl(C0+NA+ei)}{dps_r}*IS!{cl(C0+NA+ei)}{dshr_r}" for ei in range(NE)]
    r = row_write(ws, r, "  Common Dividends Paid", CFDIV + div_f, FMT_NUM, indent=1)

    CFS_R["oth_fin"] = r
    r = row_write(ws, r, "  Other Financing Activities", CFOFIN + [-15]*NE, FMT_NUM, indent=1)

    CFS_R["cff"] = r
    cff_f = [f"={cl(C0+NA+ei)}{CFS_R['debt_iss']}+{cl(C0+NA+ei)}{CFS_R['debt_rep']}+"
             f"{cl(C0+NA+ei)}{CFS_R['stk_iss']}+{cl(C0+NA+ei)}{CFS_R['buyback']}+"
             f"{cl(C0+NA+ei)}{CFS_R['divid']}+{cl(C0+NA+ei)}{CFS_R['oth_fin']}" for ei in range(NE)]
    r = row_write(ws, r, "Net Cash from Financing Activities", CFF + cff_f, FMT_NUM, bold=True, is_total=True)

    CFS_R["fx"] = r
    fx_f = [f"=Assumptions!{cl(C0+ei)}{ASS_R['fx_effect']}" for ei in range(NE)]
    r = row_write(ws, r, "  Effect of FX on Cash", CFFX + fx_f, FMT_NUM, indent=1)

    CFS_R["net_chg"] = r
    nc_f = [f"={cl(C0+NA+ei)}{CFS_R['cfo']}+{cl(C0+NA+ei)}{CFS_R['cfi']}+"
            f"{cl(C0+NA+ei)}{CFS_R['cff']}+{cl(C0+NA+ei)}{CFS_R['fx']}" for ei in range(NE)]
    r = row_write(ws, r, "Net Change in Cash", CFNC + nc_f, FMT_NUM, bold=True, is_total=True)

    CFS_R["beg_cash"] = r
    beg_f = []
    for ei in range(NE):
        prev_c = cl(C0+NA+ei-1) if ei > 0 else cl(C0+NA-1)
        beg_f.append(f"=BS!{prev_c}{BS_R['cash']}")
    r = row_write(ws, r, "  Beginning Cash", CFBEG + beg_f, FMT_NUM, indent=1)

    CFS_R["end_cash"] = r
    end_f = [f"={cl(C0+NA+ei)}{CFS_R['beg_cash']}+{cl(C0+NA+ei)}{CFS_R['net_chg']}" for ei in range(NE)]
    r = row_write(ws, r, "Ending Cash", CFEND + end_f, FMT_NUM, bold=True, is_total=True)

    r = section(ws, r, "SUPPLEMENTAL")
    CFS_R["fcf"] = r
    fcf_f = [f"={cl(C0+NA+ei)}{CFS_R['cfo']}+{cl(C0+NA+ei)}{CFS_R['capex']}" for ei in range(NE)]
    r = row_write(ws, r, "  Free Cash Flow", CFFCF + fcf_f, FMT_NUM, indent=1)

    CFS_R["fcf_m"] = r
    fcfm_f = [f"={cl(C0+NA+ei)}{CFS_R['fcf']}/IS!{cl(C0+NA+ei)}{rev_r}" for ei in range(NE)]
    fcfm_act = [0.129, 0.149, 0.153, 0.163]
    r = row_write(ws, r, "  FCF Margin %", fcfm_act + fcfm_f, FMT_PCT, helper=True)

    return ws

def backfill_bs_cash(bs_ws):
    """After CFS is built, fill estimate-year cash cells in BS with CFS links."""
    for ei in range(NE):
        col = C0 + NA + ei
        row = BS_R["cash"]
        cell = bs_ws.cell(row=row, column=col)
        cell.value = f"=CFS!{cl(col)}{CFS_R['end_cash']}"
        cell.font = fnt(color=GREEN_LK)
        cell.number_format = FMT_NUM
        cell.alignment = Alignment(horizontal="right")

# ── BUILD DCF TAB ─────────────────────────────────────────────────────────────
def build_dcf(wb):
    ws = wb.create_sheet("DCF")
    ws.sheet_properties.tabColor = RED_TAB
    ws.column_dimensions["A"].width = 1
    ws.column_dimensions["B"].width = 38
    ws.column_dimensions["C"].width = 20
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 35
    for i in range(2, 9):
        ws.column_dimensions[cl(5+i)].width = 13
    ws.row_dimensions[1].height = 4

    title = "L'Oréal SA (OR) — DCF Valuation (€M)"
    c = ws.cell(row=2, column=2, value=title)
    c.fill = fl(NAVY); c.font = fnt(bold=True, color=WHITE, sz=13)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=2, start_column=2, end_row=2, end_column=12)
    ws.row_dimensions[2].height = 22

    def dcf_hdr(row, label):
        c = ws.cell(row=row, column=2, value=label)
        c.fill = fl(BLUE_HDR); c.font = fnt(bold=True, color=WHITE)
        ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=12)
        ws.row_dimensions[row].height = 16

    def dcf_inp(row, label, val, fmt=FMT_NUM, note=""):
        c = ws.cell(row=row, column=2, value=label)
        c.font = fnt(color=BLACK); c.alignment = Alignment(horizontal="left")
        v = ws.cell(row=row, column=3, value=val)
        v.number_format = fmt; v.font = fnt(color=BLUE_IN)
        v.alignment = Alignment(horizontal="right")
        if note:
            n = ws.cell(row=row, column=5, value=note)
            n.font = fnt(italic=True, color=GREY, sz=9)
        return row + 1

    r = 3
    dcf_hdr(r, "SECTION 1 — DCF ASSUMPTIONS"); r += 1

    WACC_ROW = r
    r = dcf_inp(r, "WACC", 0.085, FMT_PCT, "~8.5% for L'Oréal: risk-free ~3.5%, ERP 5%, β~0.9, low leverage")
    TGR_ROW = r
    r = dcf_inp(r, "Terminal Growth Rate", 0.025, FMT_PCT, "2.5% — in line with long-run EU nominal GDP")
    PROJ_YRS_ROW = r
    r = dcf_inp(r, "Projection Period (years)", 5, FMT_NUM, "5-year explicit forecast (2026E–2030E)")
    CASH_ROW = r
    r = dcf_inp(r, "Cash & Equivalents (€M)", 9865, FMT_NUM, f"=BS!{cl(C0+NA-1)}{BS_R['cash']} — 2025 actual")
    STI_ROW = r
    r = dcf_inp(r, "Short-Term Investments (€M)", 0, FMT_NUM, "Nil as of 2025")
    DEBT_ROW = r
    r = dcf_inp(r, "Total Debt (€M)", 11916, FMT_NUM, "2025 actual: LTD + ST debt + leases")
    NETCASH_ROW = r
    ws.cell(row=r, column=2, value="Net Cash (€M)").font = fnt(color=BLACK)
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="left")
    nc_cell = ws.cell(row=r, column=3, value=f"=C{CASH_ROW}+C{STI_ROW}-C{DEBT_ROW}")
    nc_cell.number_format = FMT_NUM; nc_cell.font = fnt(color=BLACK)
    nc_cell.alignment = Alignment(horizontal="right")
    r += 1
    DILSH_ROW = r
    r = dcf_inp(r, "Diluted Shares Outstanding (M)", 535.4, FMT_NUM, "2025 actual weighted avg diluted shares")
    CURPX_ROW = r
    r = dcf_inp(r, "Current Share Price (€)", 366.60, FMT_EPS, "As of 12/31/25 from TIKR data")
    NTM_EV_REV_ROW = r
    r = dcf_inp(r, "NTM EV/Revenue Reference (peers)", 3.5, FMT_MUL, "L'Oréal historical ~3-5x; premium consumer staples")
    r += 1

    # Section 2 — FCF Build
    dcf_hdr(r, "SECTION 2 — UNLEVERED FCF PROJECTIONS"); r += 1
    # Column headers for DCF section
    ws.cell(row=r, column=2, value="").fill = fl(NAVY)
    for ei, yr in enumerate(EST_YRS):
        c = ws.cell(row=r, column=4+ei, value=yr)
        c.fill = fl(GREEN_HDR); c.font = fnt(bold=True, color=WHITE)
        c.alignment = Alignment(horizontal="center")
    tv_c = ws.cell(row=r, column=4+NE, value="Terminal")
    tv_c.fill = fl(NAVY); tv_c.font = fnt(bold=True, color=WHITE)
    tv_c.alignment = Alignment(horizontal="center")
    HDR_ROW_DCF = r; r += 1

    def dcf_frow(row, label, vals, fmt=FMT_NUM, bold=False, helper=False, bg=None):
        c = ws.cell(row=row, column=2, value=label)
        c.font = fnt(bold=bold and not helper, italic=helper, color=BLACK)
        c.alignment = Alignment(horizontal="left")
        for i, v in enumerate(vals):
            col = 4 + i
            cell = ws.cell(row=row, column=col)
            if v is not None and v != "":
                cell.value = v
            cell.number_format = fmt
            v_str = str(v) if v else ""
            if helper:
                cell.font = fnt(italic=True, color=GREY)
            elif v_str.startswith("=") and any(s+"!" in v_str for s in ["IS","BS","CFS"]):
                cell.font = fnt(bold=bold, color=GREEN_LK)
            else:
                cell.font = fnt(bold=bold, color=BLACK)
            cell.alignment = Alignment(horizontal="right")
            if bg: cell.fill = fl(bg)
        return row + 1

    DCF_REV_ROW = r
    rev_vals = [f"=IS!{cl(C0+NA+ei)}{IS_R['rev']}" for ei in range(NE)] + [None]
    r = dcf_frow(r, "Revenue (€M)", rev_vals)

    rev_g_vals = [f"=IS!{cl(C0+NA+ei)}{IS_R['rev_growth_h']}" for ei in range(NE)] + [None]
    r = dcf_frow(r, "  Revenue Growth %", rev_g_vals, FMT_PCT, helper=True)

    DCF_EBIT_ROW = r
    ebit_vals = [f"=IS!{cl(C0+NA+ei)}{IS_R['ebit']}" for ei in range(NE)] + [None]
    r = dcf_frow(r, "EBIT (€M)", ebit_vals)

    ebit_m_vals = [f"=IS!{cl(C0+NA+ei)}{IS_R['ebit_m_h']}" for ei in range(NE)] + [None]
    r = dcf_frow(r, "  EBIT Margin %", ebit_m_vals, FMT_PCT, helper=True)

    DCF_TAX_EBIT_ROW = r
    # Tax on EBIT = max(0, EBIT * tax_rate)
    tax_ebit = [f"=-MAX(0,IS!{cl(C0+NA+ei)}{IS_R['ebit']})*Assumptions!{cl(C0+ei)}{ASS_R['tax_rate']}" for ei in range(NE)] + [None]
    r = dcf_frow(r, "  Tax on EBIT", tax_ebit)

    DCF_NOPAT_ROW = r
    nopat = [f"={cl(4+ei)}{DCF_EBIT_ROW}+{cl(4+ei)}{DCF_TAX_EBIT_ROW}" for ei in range(NE)] + [None]
    r = dcf_frow(r, "NOPAT", nopat, bold=True)

    DCF_DA_ROW = r
    da_vals = [f"=IS!{cl(C0+NA+ei)}{IS_R['da']}" for ei in range(NE)] + [None]
    r = dcf_frow(r, "  (+) D&A", da_vals)

    DCF_SBC_ROW = r
    sbc_vals = [f"=IS!{cl(C0+NA+ei)}{IS_R['sbc']}" for ei in range(NE)] + [None]
    r = dcf_frow(r, "  (+) SBC", sbc_vals)

    DCF_CAPEX_ROW = r
    capex_vals = [f"=CFS!{cl(C0+NA+ei)}{CFS_R['capex']}" for ei in range(NE)] + [None]
    r = dcf_frow(r, "  (-) Capex", capex_vals)

    DCF_WC_ROW = r
    # WC change = -(change in TCA excl cash - change in TCL excl STD)
    wc_vals = []
    for ei in range(NE):
        curr_c = cl(C0+NA+ei); prev_c = cl(C0+NA+ei-1) if ei > 0 else cl(C0+NA-1)
        wc_vals.append(f"=-(BS!{curr_c}{BS_R['tca']}-BS!{curr_c}{BS_R['cash']}"
                       f"-(BS!{prev_c}{BS_R['tca']}-BS!{prev_c}{BS_R['cash']}))"
                       f"+(BS!{curr_c}{BS_R['tcl']}-BS!{curr_c}{BS_R['stb']}-BS!{curr_c}{BS_R['cltd']}"
                       f"-(BS!{prev_c}{BS_R['tcl']}-BS!{prev_c}{BS_R['stb']}-BS!{prev_c}{BS_R['cltd']}))")
    wc_vals.append(None)
    r = dcf_frow(r, "  (+/-) Change in Working Capital", wc_vals)

    DCF_UFCF_ROW = r
    ufcf = [f"={cl(4+ei)}{DCF_NOPAT_ROW}+{cl(4+ei)}{DCF_DA_ROW}+{cl(4+ei)}{DCF_SBC_ROW}+"
            f"{cl(4+ei)}{DCF_CAPEX_ROW}+{cl(4+ei)}{DCF_WC_ROW}" for ei in range(NE)]
    # Terminal year column (col 4+NE)
    ufcf.append(None)
    r = dcf_frow(r, "Unlevered Free Cash Flow (uFCF)", ufcf, bold=True, bg=None)
    bt = top_border(BLUE_HDR)
    for ei in range(NE):
        ws.cell(row=r-1, column=4+ei).border = bt

    ufcf_m = [f"={cl(4+ei)}{DCF_UFCF_ROW}/IS!{cl(C0+NA+ei)}{IS_R['rev']}" for ei in range(NE)] + [None]
    r = dcf_frow(r, "  uFCF Margin %", ufcf_m, FMT_PCT, helper=True)
    r += 1

    # Section 3 — DCF Calculation
    dcf_hdr(r, "SECTION 3 — DCF CALCULATION"); r += 1

    # Period, discount factors, FCF, TV, PV
    ws.cell(row=r, column=2, value="Period").font = fnt(italic=True, color=GREY)
    for ei in range(NE):
        c = ws.cell(row=r, column=4+ei, value=ei+1)
        c.font = fnt(italic=True, color=GREY); c.alignment = Alignment(horizontal="center")
    PERIOD_ROW = r; r += 1

    ws.cell(row=r, column=2, value="Discount Factor").font = fnt(color=BLACK)
    for ei in range(NE):
        c = ws.cell(row=r, column=4+ei, value=f"=1/(1+C{WACC_ROW})^{cl(4+ei)}{PERIOD_ROW}")
        c.font = fnt(color=BLACK); c.number_format = "0.0000"; c.alignment = Alignment(horizontal="right")
    DISC_ROW = r; r += 1

    ws.cell(row=r, column=2, value="uFCF (€M)").font = fnt(color=GREEN_LK)
    for ei in range(NE):
        c = ws.cell(row=r, column=4+ei, value=f"={cl(4+ei)}{DCF_UFCF_ROW}")
        c.font = fnt(color=GREEN_LK); c.number_format = FMT_NUM; c.alignment = Alignment(horizontal="right")
    FCF_ROW_DCF = r; r += 1

    # Terminal Value at end of Year 5 using Gordon Growth
    TV_COL = 4 + NE
    ws.cell(row=r, column=2, value="Terminal Value (€M)").font = fnt(bold=True, color=BLACK)
    ws.cell(row=r, column=TV_COL, value=f"={cl(4+NE-1)}{FCF_ROW_DCF}*(1+C{TGR_ROW})/(C{WACC_ROW}-C{TGR_ROW})")
    ws.cell(row=r, column=TV_COL).number_format = FMT_NUM
    ws.cell(row=r, column=TV_COL).font = fnt(bold=True, color=BLACK)
    TV_ROW = r; r += 1

    ws.cell(row=r, column=2, value="PV of FCF (€M)").font = fnt(color=BLACK)
    for ei in range(NE):
        c = ws.cell(row=r, column=4+ei, value=f"={cl(4+ei)}{FCF_ROW_DCF}*{cl(4+ei)}{DISC_ROW}")
        c.font = fnt(color=BLACK); c.number_format = FMT_NUM; c.alignment = Alignment(horizontal="right")
    PV_FCF_ROW = r; r += 1

    ws.cell(row=r, column=2, value="PV of Terminal Value (€M)").font = fnt(bold=True, color=BLACK)
    ws.cell(row=r, column=TV_COL, value=f"={cl(TV_COL)}{TV_ROW}*{cl(4+NE-1)}{DISC_ROW}")
    ws.cell(row=r, column=TV_COL).number_format = FMT_NUM
    ws.cell(row=r, column=TV_COL).font = fnt(bold=True, color=BLACK)
    PV_TV_ROW = r; r += 1
    r += 1

    # Section 4 — Valuation Bridge
    dcf_hdr(r, "SECTION 4 — VALUATION BRIDGE"); r += 1

    def vrow(label, formula, fmt=FMT_NUM, bold=False, bg=None, color=BLACK):
        ws.cell(row=r, column=2, value=label).font = fnt(bold=bold, color=BLACK)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="left")
        c = ws.cell(row=r, column=3, value=formula)
        c.number_format = fmt; c.font = fnt(bold=bold, color=color)
        c.alignment = Alignment(horizontal="right")
        if bg: c.fill = fl(bg)
        return r + 1

    SUM_PV_FCF_ROW = r
    f = "=SUM(" + ",".join([f"{cl(4+ei)}{PV_FCF_ROW}" for ei in range(NE)]) + ")"
    r = vrow("Sum of PV of FCFs (€M)", f)
    PV_TV_BRIDGE_ROW = r
    r = vrow("(+) PV of Terminal Value (€M)", f"={cl(TV_COL)}{PV_TV_ROW}")
    EV_ROW = r
    r = vrow("Enterprise Value (€M)", f"=C{SUM_PV_FCF_ROW}+C{PV_TV_BRIDGE_ROW}", bold=True, bg=GREEN_HL)
    ws.cell(row=EV_ROW-1, column=3).border = top_border(BLUE_HDR)
    TV_PCT_ROW = r
    r = vrow("  % from Terminal Value", f"=C{PV_TV_BRIDGE_ROW}/C{EV_ROW}", FMT_PCT, color=GREY)
    ws.cell(row=TV_PCT_ROW-1, column=2).font = fnt(italic=True, color=GREY)
    NET_CASH_BRIDGE_ROW = r
    r = vrow("(+) Net Cash (€M)", f"=C{NETCASH_ROW}")
    EQV_ROW = r
    r = vrow("Equity Value (€M)", f"=C{EV_ROW}+C{NET_CASH_BRIDGE_ROW}", bold=True, bg=GREEN_HL)
    ws.cell(row=EQV_ROW-1, column=3).border = top_border(BLUE_HDR)
    SHARE_ROW = r
    r = vrow("(/) Diluted Shares (M)", f"=C{DILSH_ROW}")
    IVS_ROW = r
    ws.cell(row=r, column=2, value="Intrinsic Value Per Share (€)").font = fnt(bold=True, color=WHITE)
    ws.cell(row=r, column=2).fill = fl(RED_TAB)
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="left")
    ivs = ws.cell(row=r, column=3, value=f"=C{EQV_ROW}/C{SHARE_ROW}")
    ivs.number_format = FMT_EPS; ivs.font = fnt(bold=True, color=WHITE)
    ivs.fill = fl(RED_TAB); ivs.alignment = Alignment(horizontal="right")
    r += 1

    # Cross-checks
    r += 1
    cc_items = [
        ("  TV as % of EV", f"=C{PV_TV_BRIDGE_ROW}/C{EV_ROW}", FMT_PCT),
        ("  Current Share Price (€)", f"=C{CURPX_ROW}", FMT_EPS),
        ("  Upside / (Downside) to Current Price",
         f"=C{IVS_ROW}/C{CURPX_ROW}-1", FMT_PCT),
        (f"  Implied EV/NTM Revenue", f"=C{EV_ROW}/IS!{cl(C0+NA)}{IS_R['rev']}", FMT_MUL),
        (f"  Implied EV/NTM EBITDA", f"=C{EV_ROW}/IS!{cl(C0+NA)}{IS_R['ebitda']}", FMT_MUL),
    ]
    for lbl, fm, fmt in cc_items:
        ws.cell(row=r, column=2, value=lbl).font = fnt(italic=True, color=GREY)
        ws.cell(row=r, column=2).alignment = Alignment(horizontal="left")
        c = ws.cell(row=r, column=3, value=fm)
        c.number_format = fmt; c.font = fnt(italic=True, color=GREY)
        c.alignment = Alignment(horizontal="right")
        r += 1

    r += 1

    # Section 5 — Sensitivity
    dcf_hdr(r, "SECTION 5 — SENSITIVITY: IMPLIED SHARE PRICE"); r += 1

    # WACC rows (5): base ±100bp & ±200bp
    # TGR cols (6): 1.5%, 2.0%, 2.5%(base), 3.0%, 3.5%, 4.0%
    wacc_vals = [0.065, 0.075, 0.085, 0.095, 0.105]  # WACC
    tgr_vals  = [0.015, 0.020, 0.025, 0.030, 0.035, 0.040]  # TGR

    SENS_START_ROW = r
    # Table header
    ws.cell(row=r, column=3, value="WACC \\ TGR →").font = fnt(bold=True, color=WHITE)
    ws.cell(row=r, column=3).fill = fl(BLUE_HDR)
    ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
    for j, tg in enumerate(tgr_vals):
        c = ws.cell(row=r, column=4+j, value=tg)
        c.number_format = FMT_PCT; c.fill = fl(BLUE_HDR)
        c.font = fnt(bold=True, color=WHITE); c.alignment = Alignment(horizontal="center")
    r += 1

    base_wacc, base_tgr = 0.085, 0.025
    for wc in wacc_vals:
        # Row label
        lbl_c = ws.cell(row=r, column=3, value=wc)
        lbl_c.number_format = FMT_PCT
        is_base_wacc = abs(wc - base_wacc) < 0.0001
        lbl_c.fill = fl(RED_TAB) if is_base_wacc else fl(BLUE_HDR)
        lbl_c.font = fnt(bold=True, color=WHITE)
        lbl_c.alignment = Alignment(horizontal="center")
        for j, tg in enumerate(tgr_vals):
            is_base_cell = is_base_wacc and abs(tg - base_tgr) < 0.0001
            # Compute implied share price for this WACC/TGR combo
            # Use the uFCF values from the DCF calculation
            ufcf_refs = "+".join([
                f"({cl(4+ei)}{FCF_ROW_DCF}/(1+{wc})^{ei+1})"
                for ei in range(NE)
            ])
            last_ufcf = f"{cl(4+NE-1)}{FCF_ROW_DCF}"
            tv_formula = f"({last_ufcf}*(1+{tg})/({wc}-{tg}))/(1+{wc})^5"
            ev_formula = f"{ufcf_refs}+{tv_formula}+C{NETCASH_ROW}"
            cell_formula = f"=({ev_formula})/C{DILSH_ROW}"
            c = ws.cell(row=r, column=4+j, value=cell_formula)
            c.number_format = FMT_EPS
            c.font = fnt(bold=is_base_cell, color=BLACK)
            c.fill = fl(YELLOW) if is_base_cell else fl(WHITE)
            c.alignment = Alignment(horizontal="center")
        r += 1

    note_c = ws.cell(row=r, column=3,
                     value="Base case: WACC=8.5%, TGR=2.5% (yellow cell)")
    note_c.font = fnt(italic=True, color=GREY, sz=9)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4+len(tgr_vals)-1)
    r += 2

    # Section 6 — EV/NTM Revenue
    c = ws.cell(row=r, column=2, value="SECTION 6 — EV / NTM REVENUE MULTIPLES")
    c.fill = fl(GREEN_HDR); c.font = fnt(bold=True, color=WHITE)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)
    ws.row_dimensions[r].height = 16
    r += 1

    NTM_REV_ROW = r
    ws.cell(row=r, column=2, value="NTM Revenue (€M)").font = fnt(color=BLACK)
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="left")
    c = ws.cell(row=r, column=3, value=f"=IS!{cl(C0+NA)}{IS_R['rev']}")
    c.number_format = FMT_NUM; c.font = fnt(color=GREEN_LK)
    c.alignment = Alignment(horizontal="right")
    r += 1

    NC_ROW6 = r
    ws.cell(row=r, column=2, value="Net Cash (€M)").font = fnt(color=BLACK)
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="left")
    c = ws.cell(row=r, column=3, value=f"=C{NETCASH_ROW}")
    c.number_format = FMT_NUM; c.font = fnt(color=BLACK); c.alignment = Alignment(horizontal="right")
    r += 1

    SH_ROW6 = r
    ws.cell(row=r, column=2, value="Diluted Shares (M)").font = fnt(color=BLACK)
    ws.cell(row=r, column=2).alignment = Alignment(horizontal="left")
    c = ws.cell(row=r, column=3, value=f"=C{DILSH_ROW}")
    c.number_format = FMT_NUM; c.font = fnt(color=BLACK); c.alignment = Alignment(horizontal="right")
    r += 1
    ws.cell(row=r, column=2,
            value="Peer context: L'Oréal peers (LVMH, Estée Lauder, Henkel) trade 2.5–5.0x EV/NTM Rev; L'Oréal historical avg ~3.5x"
           ).font = fnt(italic=True, color=GREY, sz=9)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=12)
    r += 1

    multiples = [2.0, 2.5, 3.0, 3.5, 4.0, 4.5, 5.0]
    base_mult = 3.5
    # Header
    ws.cell(row=r, column=3, value="Multiple →").font = fnt(bold=True, color=WHITE)
    ws.cell(row=r, column=3).fill = fl(GREEN_HDR); ws.cell(row=r, column=3).alignment = Alignment(horizontal="center")
    for j, m in enumerate(multiples):
        c = ws.cell(row=r, column=4+j, value=m)
        c.number_format = FMT_MUL; c.fill = fl(GREEN_HDR)
        c.font = fnt(bold=True, color=WHITE); c.alignment = Alignment(horizontal="center")
    r += 1

    ev_row_m = r
    ws.cell(row=r, column=3, value="Implied EV (€M)").font = fnt(color=BLACK)
    ws.cell(row=r, column=3).fill = fl(LIGHT_BLUE)
    ws.cell(row=r, column=3).alignment = Alignment(horizontal="left")
    for j, m in enumerate(multiples):
        c = ws.cell(row=r, column=4+j, value=f"=C3*{m}")
        # Actually reference NTM_REV_ROW row 3 is wrong, it is NTM_REV_ROW
        c.value = f"=C{NTM_REV_ROW}*{m}"
        c.number_format = FMT_NUM; c.fill = fl(LIGHT_BLUE)
        c.font = fnt(color=BLACK); c.alignment = Alignment(horizontal="right")
        if abs(m - base_mult) < 0.01: c.fill = fl(YELLOW)
    r += 1

    eqv_row_m = r
    ws.cell(row=r, column=3, value="Equity Value (€M)").font = fnt(color=BLACK)
    ws.cell(row=r, column=3).alignment = Alignment(horizontal="left")
    for j, m in enumerate(multiples):
        c = ws.cell(row=r, column=4+j, value=f"={cl(4+j)}{ev_row_m}+C{NC_ROW6}")
        c.number_format = FMT_NUM; c.font = fnt(color=BLACK); c.alignment = Alignment(horizontal="right")
    r += 1

    ws.cell(row=r, column=3, value="Implied Share Price (€)").font = fnt(bold=True, color=WHITE)
    ws.cell(row=r, column=3).fill = fl(GREEN_HDR)
    ws.cell(row=r, column=3).alignment = Alignment(horizontal="left")
    for j, m in enumerate(multiples):
        c = ws.cell(row=r, column=4+j, value=f"={cl(4+j)}{eqv_row_m}/C{SH_ROW6}")
        c.number_format = FMT_EPS
        if abs(m - base_mult) < 0.01:
            c.fill = fl(YELLOW); c.font = fnt(bold=True, color=BLACK)
        else:
            c.fill = fl(GREEN_HDR); c.font = fnt(bold=True, color=WHITE)
        c.alignment = Alignment(horizontal="right")
    r += 1
    note_r = ws.cell(row=r, column=3,
                     value=f"NTM period = 12/31/26E; net cash treatment = add net cash to EV. Base case = {base_mult}x (yellow)")
    note_r.font = fnt(italic=True, color=GREY, sz=9)
    ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=4+len(multiples)-1)

    return ws

# ── MAIN ──────────────────────────────────────────────────────────────────────
def main():
    OUTPUT_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    # Build in order: Assumptions first (so row numbers are known), then IS/BS/CFS/DCF
    ass_ws  = build_assumptions(wb)
    is_ws   = build_is(wb)
    bs_ws   = build_bs(wb)
    cfs_ws  = build_cfs(wb)
    backfill_bs_cash(bs_ws)  # now that CFS rows are known
    dcf_ws  = build_dcf(wb)

    # Reorder tabs: IS, BS, CFS, Assumptions, DCF
    wb.move_sheet("IS",          offset=-(wb.index(wb["IS"])))
    wb.move_sheet("BS",          offset=1 - wb.index(wb["BS"]))
    wb.move_sheet("CFS",         offset=2 - wb.index(wb["CFS"]))
    wb.move_sheet("Assumptions", offset=3 - wb.index(wb["Assumptions"]))
    wb.move_sheet("DCF",         offset=4 - wb.index(wb["DCF"]))

    wb.save(OUTPUT_PATH)
    print(f"Saved → {OUTPUT_PATH}")

if __name__ == "__main__":
    main()
