#!/usr/bin/env python3
"""Build DGE (Diageo) Financial Model — 5 tabs: IS, BS, CFS, Assumptions, DCF."""

from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUT = Path("/Users/mh/Library/CloudStorage/OneDrive-AG/01_equities_research/DGE/DGE_Financial_Model.xlsx")

# ── colours ───────────────────────────────────────────────────────────────
NAVY="1F3864"; BLUE_H="2F5496"; GREEN_H="375623"; SECT="D6E4F7"
GREEN_TOT="E2EFDA"; RED_BG="C00000"; YELLOW="FFFF00"; LT_BLUE="D9E1F2"
WHITE="FFFFFF"; BLF="0000FF"; BKF="000000"; GRF="008000"; GYF="888888"; WHF="FFFFFF"
SECT_TXT="1F3864"; BROWN="843C0C"; PURPLE="7030A0"

FMT_N='#,##0'; FMT_NN='#,##0;(#,##0);-'; FMT_P='0.0%'; FMT_EP='$#,##0.00'; FMT_MX='0.0x'

def fl(h): return PatternFill("solid", fgColor=h)
def fn(c=BKF, bold=False, italic=False, sz=10):
    return Font(color=c, bold=bold, italic=italic, size=sz, name="Calibri")
def bdr(c=BLUE_H):
    s=Side(style='medium',color=c); return Border(top=s)
CTR=Alignment(horizontal='center',vertical='center')
LFT=Alignment(horizontal='left',vertical='center')
IND=Alignment(horizontal='left',vertical='center',indent=2)

# ── core cell writer ──────────────────────────────────────────────────────
def w(ws,r,c,val=None,bg=WHITE,fc=BKF,bold=False,italic=False,fmt=None,
      align=CTR,mc=None,bc=None,sz=10):
    cell=ws.cell(row=r,column=c,value=val)
    cell.fill=fl(bg); cell.font=fn(fc,bold,italic,sz); cell.alignment=align
    if fmt: cell.number_format=fmt
    if mc: ws.merge_cells(start_row=r,start_column=c,end_row=r,end_column=mc)
    if bc: cell.border=bdr(bc)
    return cell

def col_widths(ws, n_val=9):
    ws.column_dimensions['A'].width=1
    ws.column_dimensions['B'].width=35
    for i in range(n_val):
        ws.column_dimensions[get_column_letter(3+i)].width=13

def title(ws,r,txt,nc=11):
    w(ws,r,2,txt,NAVY,WHF,True,False,None,LFT,nc,sz=13)
    ws.row_dimensions[r].height=22

def hdrs(ws,r,labels):
    for i,lbl in enumerate(labels):
        is_e=isinstance(lbl,str) and lbl.strip().endswith('E')
        w(ws,r,2+i,lbl,GREEN_H if is_e else BLUE_H,WHF,True,align=CTR)
    ws.row_dimensions[r].height=18

def sect(ws,r,lbl,nc=11):
    w(ws,r,2,lbl,SECT,SECT_TXT,True,align=LFT,mc=nc)

def spacer(ws,r,h=5):
    ws.row_dimensions[r].height=h
    for c in range(1,12): ws.cell(row=r,column=c).fill=fl(WHITE)

# ── data ──────────────────────────────────────────────────────────────────
ACT_YRS=["7/1/22","6/30/23","6/30/24","6/30/25"]
EST_YRS=["6/30/26 E","6/30/27 E","6/30/28 E","6/30/29 E","6/30/30 E"]
# col mapping: B=2,C=3(FY22),D=4(FY23),E=5(FY24),F=6(FY25),G=7(26E)..K=11(30E)
# Assumptions: B=2,C=3(26E),D=4(27E),E=5(28E),F=6(29E),G=7(30E),H=8(Notes)

# ═══════════════════════════════════════════════════════════════════════════
# ASSUMPTIONS TAB
# ═══════════════════════════════════════════════════════════════════════════
def build_assumptions(wb):
    ws=wb.create_sheet("Assumptions")
    ws.sheet_properties.tabColor=PURPLE
    ws.column_dimensions['A'].width=1
    ws.column_dimensions['B'].width=38
    for i in range(5): ws.column_dimensions[get_column_letter(3+i)].width=13
    ws.column_dimensions['H'].width=45

    ws.row_dimensions[1].height=3.75
    title(ws,2,"Diageo PLC (DGE) — Assumptions (£M)",8)
    hdrs(ws,3,["Assumption"]+EST_YRS+["Notes"])

    # ── IS DRIVERS ─────────────────────────────────────────────────────────
    spacer(ws,4)
    sect(ws,5,"INCOME STATEMENT DRIVERS",8)

    is_rows=[
        # (row, label, [c3..g7 values], fmt, notes)
        (6,  "Revenue Growth %",         [-0.0347,-0.0030,0.0276,0.0509,0.0373],  FMT_P,  "TIKR consensus estimates, Mar-2026"),
        (7,  "Gross Margin %",            [0.5995,0.5986,0.6000,0.6020,0.6020],    FMT_P,  "TIKR consensus estimates, Mar-2026"),
        (8,  "SG&A % of Revenue",         [0.182,0.182,0.182,0.182,0.182],         FMT_P,  "Blended FY22-25 average ~18.2%"),
        (9,  "Other OpEx % of Revenue",   [0.1307,0.1325,0.1315,0.1365,0.1391],   FMT_P,  "Implied from EBIT margin targets"),
        (10, "Interest Expense (£M)",     [-840.73,-738.14,-665.11,-635.80,-563.50],FMT_NN,"TIKR consensus, Mar-2026"),
        (11, "Interest & Investment Income (£M)",[200,200,200,200,200],            FMT_NN, "Mgmt guidance; declining rate env."),
        (12, "Equity Income (£M)",        [150,150,150,150,150],                   FMT_NN, "Moët Hennessy JV; conservative"),
        (13, "Other Non-Op Income (£M)",  [-30,-30,-30,-30,-30],                   FMT_NN, "Recurring ~(30) per actuals"),
        (14, "Unusual Items (£M)",        [0,0,0,0,0],                             FMT_NN, "Normalized to zero for projections"),
        (15, "Tax Rate %",                [0.246,0.247,0.247,0.252,0.258],         FMT_P,  "TIKR consensus, Mar-2026"),
        (16, "Minority Interest (£M)",    [-184,-184,-184,-184,-184],              FMT_NN, "Flat at FY25 level"),
        (17, "Diluted Shares (M)",        [2210,2195,2180,2165,2150],              FMT_N,  "Declining ~0.6%/yr; no buyback assumed"),
        (18, "Basic Shares (M)",          [2204,2189,2174,2159,2145],              FMT_N,  "Consistent with diluted"),
        (19, "Dividends Per Share (£)",   [0.56,0.63,0.65,0.71,0.76],             FMT_EP, "TIKR consensus, Mar-2026"),
        (20, "R&D Expense (£M)",          [74,74,74,74,74],                        FMT_NN, "Flat at FY25 level"),
        (21, "Depreciation & Amortisation (£M)",[803,805,827,843,906],            FMT_NN, "TIKR consensus D&A, Mar-2026"),
    ]
    for row,lbl,vals,fmt,note in is_rows:
        w(ws,row,2,lbl,WHITE,BKF,align=LFT)
        for i,v in enumerate(vals):
            w(ws,row,3+i,v,WHITE,BLF,fmt=fmt)
        w(ws,row,8,note,WHITE,GYF,italic=True,align=LFT)

    # ── BS DRIVERS ─────────────────────────────────────────────────────────
    spacer(ws,22)
    sect(ws,23,"BALANCE SHEET DRIVERS",8)

    bs_rows=[
        (24,"DSO — Accounts Receivable (days)",   [65,65,65,65,65],       FMT_N, "FY23-25 avg ~64d"),
        (25,"DIO — Inventory (days)",              [480,480,480,480,480],  FMT_N, "Spirits ageing; FY25 actual ~486d"),
        (26,"DPO — Accounts Payable (days)",       [145,145,145,145,145],  FMT_N, "FY23-25 avg ~144d"),
        (27,"Other Current Assets (£M)",           [334,334,334,334,334],  FMT_NN,"Flat at FY25 level"),
        (28,"Prepaid Expenses (£M)",               [133,133,133,133,133],  FMT_NN,"Flat at FY25 level"),
        (29,"Net PP&E (£M)",                       [9500,9700,9900,10100,10300],FMT_NN,"Gradual capex-led growth"),
        (30,"Long-Term Investments (£M)",          [5600,5850,6100,6350,6600],FMT_NN,"Equity method investments, ~5%/yr"),
        (31,"Goodwill (£M)",                       [3000,3050,3100,3150,3200],FMT_NN,"Modest bolt-on M&A"),
        (32,"Other Intangibles (£M)",              [11700,11600,11500,11400,11300],FMT_NN,"Gradual amortisation"),
        (33,"Deferred Tax Assets LT (£M)",         [150,150,150,150,150],  FMT_NN,"Flat at FY25 level"),
        (34,"Other Long-Term Assets (£M)",         [1900,1850,1800,1800,1800],FMT_NN,"Modest decline"),
        (35,"Accrued Expenses (£M)",               [2957,2900,2900,2900,2900],FMT_NN,"FY25 level"),
        (36,"Short-Term Borrowings (£M)",          [22,22,22,22,22],       FMT_NN,"FY25 level"),
        (37,"Current Portion of LT Debt (£M)",     [2500,2500,2500,2500,2500],FMT_NN,"Assume steady maturities"),
        (38,"Current Capital Lease Obligations (£M)",[112,112,112,112,112],FMT_NN,"FY25 level"),
        (39,"Current Income Taxes Payable (£M)",   [138,138,138,138,138],  FMT_NN,"FY25 level"),
        (40,"Unearned Revenue — Current (£M)",     [82,82,82,82,82],       FMT_NN,"FY25 level"),
        (41,"Other Current Liabilities (£M)",      [1364,1364,1364,1364,1364],FMT_NN,"FY25 level"),
        (42,"Long-Term Debt (£M)",                 [20250,18000,16500,14500,13000],FMT_NN,"De-leveraging per consensus Net Debt"),
        (43,"Capital Leases — Long-Term (£M)",     [541,541,541,541,541],  FMT_NN,"FY25 level"),
        (44,"Pension & Post-Retirement (£M)",      [409,409,409,409,409],  FMT_NN,"FY25 level"),
        (45,"Deferred Tax Liability NC (£M)",      [2944,2944,2944,2944,2944],FMT_NN,"FY25 level"),
        (46,"Other Non-Current Liabilities (£M)",  [516,516,516,516,516],  FMT_NN,"FY25 level"),
    ]
    for row,lbl,vals,fmt,note in bs_rows:
        w(ws,row,2,lbl,WHITE,BKF,align=LFT)
        for i,v in enumerate(vals):
            w(ws,row,3+i,v,WHITE,BLF,fmt=fmt)
        w(ws,row,8,note,WHITE,GYF,italic=True,align=LFT)

    # ── CFS DRIVERS ────────────────────────────────────────────────────────
    spacer(ws,47)
    sect(ws,48,"CASH FLOW DRIVERS",8)

    cfs_rows=[
        (49,"Capex % of Revenue",              [0.062,0.061,0.057,0.054,0.055],  FMT_P, "TIKR consensus capex, Mar-2026"),
        (50,"Sale of PP&E (£M)",               [63,63,63,63,63],                  FMT_NN,"FY25 level"),
        (51,"Cash Acquisitions (£M)",          [-35,-35,-35,-35,-35],             FMT_NN,"FY25 level; modest bolt-ons"),
        (52,"Divestitures (£M)",               [143,143,143,143,143],             FMT_NN,"FY25 level"),
        (53,"Investment in Securities (£M)",   [-84,-84,-84,-84,-84],             FMT_NN,"FY25 level"),
        (54,"Net Loans Originated (£M)",       [-195,-195,-195,-195,-195],        FMT_NN,"FY25 level"),
        (55,"Total Debt Issued (£M)",          [3000,2000,2000,1500,1500],        FMT_NN,"De-leveraging path"),
        (56,"Total Debt Repaid (£M)",          [-3500,-3500,-2500,-2500,-2000],   FMT_NN,"Net debt reduction per consensus"),
        (57,"Issuance of Common Stock (£M)",   [15,15,15,15,15],                  FMT_NN,"FY25 level"),
        (58,"Repurchase of Common Stock (£M)", [0,0,0,0,0],                       FMT_NN,"No buybacks assumed"),
        (59,"Other Financing Activities (£M)", [-109,-109,-109,-109,-109],        FMT_NN,"FY25 level"),
        (60,"FX Adjustments on Cash (£M)",     [-35,-35,-35,-35,-35],             FMT_NN,"FY25 level"),
        (61,"Miscellaneous Cash Adjustments (£M)",[21,21,21,21,21],              FMT_NN,"FY25 level"),
    ]
    for row,lbl,vals,fmt,note in cfs_rows:
        w(ws,row,2,lbl,WHITE,BKF,align=LFT)
        for i,v in enumerate(vals):
            w(ws,row,3+i,v,WHITE,BLF,fmt=fmt)
        w(ws,row,8,note,WHITE,GYF,italic=True,align=LFT)

    # ── MODEL NOTES ────────────────────────────────────────────────────────
    spacer(ws,62)
    sect(ws,63,"MODEL NOTES",8)
    notes=[
        (64,"COLOUR CODING",None,True),
        (65,"Blue (#0000FF) — hardcoded inputs (actuals & user-editable assumptions)",None,False),
        (66,"Black (#000000) — calculated formulas",None,False),
        (67,"Green (#008000) — cross-sheet links pulling from another tab",None,False),
        (68,"Grey italic (#888888) — helper / ratio rows (margins, growth %)",None,False),
        (69,"UNITS: All monetary values in £ millions (£M). Shares in millions.",None,False),
        (70,"FISCAL YEAR: Diageo FY ends 30 June each year.",None,False),
    ]
    for row,txt,_,bold in notes:
        w(ws,row,2,txt,WHITE,GYF,bold=bold,italic=not bold,align=LFT,mc=8)

    return ws

# ═══════════════════════════════════════════════════════════════════════════
# INCOME STATEMENT TAB
# ═══════════════════════════════════════════════════════════════════════════
# Row reference constants for IS (used in CFS/DCF cross-refs)
IS_R = {
    'rev':6,'tot_rev':7,'cogs':11,'gp':14,'sga':18,'ooe':19,'topex':20,
    'oi':22,'int_exp':26,'int_inc':27,'eq_inc':28,'oth_no':29,
    'ebt_ex':30,'unusual':32,'ebt':33,'tax':34,'econt':35,
    'mi':37,'ni':38,'dshares':42,'deps':43,'bshares':44,'beps':45,'dps':46,
    'rd':49,'da':50,'ebitda':51,'ebitdam':52
}

def build_is(wb):
    ws=wb.create_sheet("IS")
    ws.sheet_properties.tabColor=BLUE_H
    col_widths(ws)
    ws.row_dimensions[1].height=3.75
    title(ws,2,"Diageo PLC (DGE) — Income Statement (£M)")
    hdrs(ws,3,[""]+ ACT_YRS + EST_YRS)

    # actuals: col C=3(FY22),D=4(FY23),E=5(FY24),F=6(FY25)
    # estimates: col G=7(26E),H=8(27E),I=9(28E),J=10(29E),K=11(30E)

    spacer(ws,4)
    sect(ws,5,"REVENUE")

    # Row 6: Revenues (single line = Total)
    rev_act=[20516,20555,20269,20245]
    w(ws,6,2,"Revenues",WHITE,BKF,align=LFT)
    for i,v in enumerate(rev_act): w(ws,6,3+i,v,WHITE,BLF,fmt=FMT_NN)
    # estimates: prior_col * (1 + Assumptions!growth)
    asmp_cols=["C","D","E","F","G"]
    for i in range(5):
        ec=7+i  # G..K
        pc=get_column_letter(ec-1) if i>0 else "F"
        ac=asmp_cols[i]
        w(ws,6,ec,f"={pc}6*(1+Assumptions!{ac}6)",WHITE,BKF,fmt=FMT_NN)

    # Row 7: Total Revenues (bold, border)
    w(ws,7,2,"Total Revenues",WHITE,BKF,True,align=LFT)
    for c in range(3,12): w(ws,7,c,f"={get_column_letter(c)}6",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)

    # Row 8: YoY Growth %
    w(ws,8,2,"  % Change YoY",WHITE,GYF,italic=True,align=IND)
    w(ws,8,3,"",WHITE,GYF); # FY22 blank
    for c in range(4,12): w(ws,8,c,f"={get_column_letter(c)}7/{get_column_letter(c-1)}7-1",WHITE,GYF,italic=True,fmt=FMT_P)

    spacer(ws,9)
    sect(ws,10,"COST OF REVENUE")

    # Row 11: COGS
    cogs_act=[-7923,-8209,-7997,-8010]
    w(ws,11,2,"Cost of Goods Sold",WHITE,BKF,align=LFT)
    for i,v in enumerate(cogs_act): w(ws,11,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(5):
        ec=7+i; ac=asmp_cols[i]
        w(ws,11,ec,f"=-{get_column_letter(ec)}6*(1-Assumptions!{ac}7)",WHITE,BKF,fmt=FMT_NN)

    # Row 12: Total Cost
    w(ws,12,2,"Total Cost of Revenue",WHITE,BKF,True,align=LFT)
    for c in range(3,12): w(ws,12,c,f"={get_column_letter(c)}11",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)

    spacer(ws,13)
    # Row 14: Gross Profit
    w(ws,14,2,"Gross Profit",WHITE,BKF,True,align=LFT)
    gp_act=[12593,12346,12272,12235]
    for i,v in enumerate(gp_act): w(ws,14,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for c in range(7,12): w(ws,14,c,f"={get_column_letter(c)}7+{get_column_letter(c)}12",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)

    # Row 15: Gross Margin %
    w(ws,15,2,"  % Gross Margin",WHITE,GYF,italic=True,align=IND)
    for c in range(3,12): w(ws,15,c,f"={get_column_letter(c)}14/{get_column_letter(c)}7",WHITE,GYF,italic=True,fmt=FMT_P)

    spacer(ws,16)
    sect(ws,17,"OPERATING EXPENSES")

    # Row 18: SG&A
    sga_act=[-3603,-3610,-3732,-3693]
    w(ws,18,2,"Selling, General & Admin Expenses",WHITE,BKF,align=LFT)
    for i,v in enumerate(sga_act): w(ws,18,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(5):
        ec=7+i; ac=asmp_cols[i]
        w(ws,18,ec,f"=-{get_column_letter(ec)}7*Assumptions!{ac}8",WHITE,BKF,fmt=FMT_NN)

    # Row 19: Other OpEx
    ooe_act=[-2603,-2370,-2541,-2816]
    w(ws,19,2,"Other Operating Expenses",WHITE,BKF,align=LFT)
    for i,v in enumerate(ooe_act): w(ws,19,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(5):
        ec=7+i; ac=asmp_cols[i]
        w(ws,19,ec,f"=-{get_column_letter(ec)}7*Assumptions!{ac}9",WHITE,BKF,fmt=FMT_NN)

    # Row 20: Total OpEx
    w(ws,20,2,"Total Operating Expenses",WHITE,BKF,True,align=LFT)
    topex_act=[-6206,-5980,-6273,-6509]
    for i,v in enumerate(topex_act): w(ws,20,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for c in range(7,12): w(ws,20,c,f"={get_column_letter(c)}18+{get_column_letter(c)}19",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)

    spacer(ws,21)
    # Row 22: Operating Income / EBIT
    w(ws,22,2,"Operating Income (EBIT)",WHITE,BKF,True,align=LFT)
    oi_act=[6387,6366,5999,5726]
    for i,v in enumerate(oi_act): w(ws,22,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for c in range(7,12): w(ws,22,c,f"={get_column_letter(c)}14+{get_column_letter(c)}20",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)

    # Row 23: EBIT Margin
    w(ws,23,2,"  % EBIT Margin",WHITE,GYF,italic=True,align=IND)
    for c in range(3,12): w(ws,23,c,f"={get_column_letter(c)}22/{get_column_letter(c)}7",WHITE,GYF,italic=True,fmt=FMT_P)

    spacer(ws,24)
    sect(ws,25,"OTHER INCOME / EXPENSE")

    # Rows 26-29
    int_exp_act=[-1114,-1059,-1212,-1196]
    w(ws,26,2,"Interest Expense",WHITE,BKF,align=LFT)
    for i,v in enumerate(int_exp_act): w(ws,26,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(5): w(ws,26,7+i,f"=Assumptions!{asmp_cols[i]}10",WHITE,BKF,fmt=FMT_NN)

    int_inc_act=[627,325,294,380]
    w(ws,27,2,"Interest & Investment Income",WHITE,BKF,align=LFT)
    for i,v in enumerate(int_inc_act): w(ws,27,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(5): w(ws,27,7+i,f"=Assumptions!{asmp_cols[i]}11",WHITE,BKF,fmt=FMT_NN)

    eq_inc_act=[555,443,414,193]
    w(ws,28,2,"Income on Equity Investments",WHITE,BKF,align=LFT)
    for i,v in enumerate(eq_inc_act): w(ws,28,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(5): w(ws,28,7+i,f"=Assumptions!{asmp_cols[i]}12",WHITE,BKF,fmt=FMT_NN)

    oth_no_act=[-38,-20,-41,-29]
    w(ws,29,2,"Other Non-Operating Income (Expenses)",WHITE,BKF,align=LFT)
    for i,v in enumerate(oth_no_act): w(ws,29,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(5): w(ws,29,7+i,f"=Assumptions!{asmp_cols[i]}13",WHITE,BKF,fmt=FMT_NN)

    # Row 30: EBT Excl Unusual
    w(ws,30,2,"EBT Excl. Unusual Items",WHITE,BKF,True,align=LFT)
    ebt_ex_act=[6417,6055,5454,5074]
    for i,v in enumerate(ebt_ex_act): w(ws,30,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for c in range(7,12): w(ws,30,c,f"={get_column_letter(c)}22+{get_column_letter(c)}26+{get_column_letter(c)}27+{get_column_letter(c)}28+{get_column_letter(c)}29",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)

    spacer(ws,31)
    # Row 32: Unusual items
    unusual_act=[-609,-413,-74,-1537] # sum of unusual lines from CSV (merger+impairment+GainLoss+writedown+other)
    w(ws,32,2,"Unusual / Non-Recurring Items",WHITE,BKF,align=LFT)
    for i,v in enumerate(unusual_act): w(ws,32,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(5): w(ws,32,7+i,f"=Assumptions!{asmp_cols[i]}14",WHITE,BKF,fmt=FMT_NN)

    # Row 33: EBT Incl Unusual
    w(ws,33,2,"EBT Incl. Unusual Items",WHITE,BKF,True,align=LFT)
    ebt_act=[5808,5642,5460,3537]
    for i,v in enumerate(ebt_act): w(ws,33,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for c in range(7,12): w(ws,33,c,f"={get_column_letter(c)}30+{get_column_letter(c)}32",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)

    # Row 34: Tax
    tax_act=[-1398,-1163,-1294,-999]
    w(ws,34,2,"Income Tax Expense",WHITE,BKF,align=LFT)
    for i,v in enumerate(tax_act): w(ws,34,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(5):
        ec=7+i; ac=asmp_cols[i]; cl=get_column_letter(ec)
        w(ws,34,ec,f"=IF({cl}33>0,-{cl}33*Assumptions!{ac}15,0)",WHITE,BKF,fmt=FMT_NN)

    # Row 35: Earnings from Cont Ops
    w(ws,35,2,"Earnings from Continuing Operations",WHITE,BKF,True,align=LFT)
    econt_act=[4410,4479,4166,2538]
    for i,v in enumerate(econt_act): w(ws,35,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for c in range(7,12): w(ws,35,c,f"={get_column_letter(c)}33+{get_column_letter(c)}34",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)

    spacer(ws,36)
    # Row 37: Minority Interest
    mi_act=[-130,-34,-296,-184]
    w(ws,37,2,"Minority Interest",WHITE,BKF,align=LFT)
    for i,v in enumerate(mi_act): w(ws,37,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(5): w(ws,37,7+i,f"=Assumptions!{asmp_cols[i]}16",WHITE,BKF,fmt=FMT_NN)

    # Row 38: Net Income
    w(ws,38,2,"Net Income",WHITE,BKF,True,align=LFT)
    ni_act=[4280,4445,3870,2354]
    for i,v in enumerate(ni_act): w(ws,38,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for c in range(7,12): w(ws,38,c,f"={get_column_letter(c)}35+{get_column_letter(c)}37",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)

    # Row 39: Net Income Margin
    w(ws,39,2,"  % Net Income Margin",WHITE,GYF,italic=True,align=IND)
    for c in range(3,12): w(ws,39,c,f"={get_column_letter(c)}38/{get_column_letter(c)}7",WHITE,GYF,italic=True,fmt=FMT_P)

    spacer(ws,40)
    sect(ws,41,"PER SHARE DATA")

    # Row 42: Diluted Shares
    dshares_act=[2325,2271,2239,2228]
    w(ws,42,2,"Weighted Average Diluted Shares (M)",WHITE,BKF,align=LFT)
    for i,v in enumerate(dshares_act): w(ws,42,3+i,v,WHITE,BLF,fmt=FMT_N)
    for i in range(5): w(ws,42,7+i,f"=Assumptions!{asmp_cols[i]}17",WHITE,BKF,fmt=FMT_N)

    # Row 43: Diluted EPS
    deps_act=[1.84,1.96,1.73,1.06]
    w(ws,43,2,"  Diluted EPS",WHITE,BKF,align=IND)
    for i,v in enumerate(deps_act): w(ws,43,3+i,v,WHITE,BLF,fmt=FMT_EP)
    for c in range(7,12): w(ws,43,c,f"={get_column_letter(c)}38/{get_column_letter(c)}42",WHITE,BKF,fmt=FMT_EP)

    # Row 44: Basic Shares
    bshares_act=[2318,2264,2234,2222]
    w(ws,44,2,"Weighted Average Basic Shares (M)",WHITE,BKF,align=LFT)
    for i,v in enumerate(bshares_act): w(ws,44,3+i,v,WHITE,BLF,fmt=FMT_N)
    for i in range(5): w(ws,44,7+i,f"=Assumptions!{asmp_cols[i]}18",WHITE,BKF,fmt=FMT_N)

    # Row 45: Basic EPS
    beps_act=[1.85,1.96,1.73,1.06]
    w(ws,45,2,"  Basic EPS",WHITE,BKF,align=IND)
    for i,v in enumerate(beps_act): w(ws,45,3+i,v,WHITE,BLF,fmt=FMT_EP)
    for c in range(7,12): w(ws,45,c,f"={get_column_letter(c)}38/{get_column_letter(c)}44",WHITE,BKF,fmt=FMT_EP)

    # Row 46: DPS
    dps_act=[0.91,0.99,1.03,1.03]
    w(ws,46,2,"  Dividends Per Share",WHITE,BKF,align=IND)
    for i,v in enumerate(dps_act): w(ws,46,3+i,v,WHITE,BLF,fmt=FMT_EP)
    for i in range(5): w(ws,46,7+i,f"=Assumptions!{asmp_cols[i]}19",WHITE,BKF,fmt=FMT_EP)

    spacer(ws,47)
    sect(ws,48,"SUPPLEMENTAL METRICS")

    # Row 49: R&D
    rd_act=[58,63,69,74]
    w(ws,49,2,"R&D Expense",WHITE,BKF,align=LFT)
    for i,v in enumerate(rd_act): w(ws,49,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(5): w(ws,49,7+i,f"=Assumptions!{asmp_cols[i]}20",WHITE,BKF,fmt=FMT_NN)

    # Row 50: D&A
    da_act=[651,597,619,674]
    w(ws,50,2,"Depreciation & Amortisation",WHITE,BKF,align=LFT)
    for i,v in enumerate(da_act): w(ws,50,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(5): w(ws,50,7+i,f"=Assumptions!{asmp_cols[i]}21",WHITE,BKF,fmt=FMT_NN)

    # Row 51: EBITDA
    w(ws,51,2,"EBITDA",WHITE,BKF,True,align=LFT)
    ebitda_act=[7038,6963,6497,6274]
    for i,v in enumerate(ebitda_act): w(ws,51,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for c in range(7,12): w(ws,51,c,f"={get_column_letter(c)}22+{get_column_letter(c)}50",WHITE,BKF,True,fmt=FMT_NN,bc=BLUE_H)

    # Row 52: EBITDA Margin
    w(ws,52,2,"  % EBITDA Margin",WHITE,GYF,italic=True,align=IND)
    for c in range(3,12): w(ws,52,c,f"={get_column_letter(c)}51/{get_column_letter(c)}7",WHITE,GYF,italic=True,fmt=FMT_P)

    return ws

# ═══════════════════════════════════════════════════════════════════════════
# BALANCE SHEET TAB
# ═══════════════════════════════════════════════════════════════════════════
def build_bs(wb):
    ws=wb.create_sheet("BS")
    ws.sheet_properties.tabColor=GREEN_H
    col_widths(ws)
    ws.row_dimensions[1].height=3.75
    title(ws,2,"Diageo PLC (DGE) — Balance Sheet (£M)")
    hdrs(ws,3,[""]+ ACT_YRS + EST_YRS)

    spacer(ws,4)
    sect(ws,5,"ASSETS")
    sect(ws,6,"CURRENT ASSETS",11)

    ac=["C","D","E","F"]; ec_cols=["G","H","I","J","K"]

    def act(r,lbl,vals,bold=False,green=False,bc_col=None,ind=False):
        alg=IND if ind else LFT
        w(ws,r,2,lbl,GREEN_TOT if green else WHITE,BKF,bold,align=alg)
        for i,v in enumerate(vals):
            w(ws,r,3+i,v,GREEN_TOT if green else WHITE,BLF,bold,fmt=FMT_NN,bc=bc_col)

    def est_cell(r,c_idx,formula,bold=False,green=False,bc_col=None):
        bg=GREEN_TOT if green else WHITE
        w(ws,r,7+c_idx,formula,bg,BKF,bold,fmt=FMT_NN,bc=bc_col)

    # ── CURRENT ASSETS ─────────────────────────────────────────────────────
    # Row 7: Cash — estimate = green link to CFS ending cash
    act(7,"Cash & Cash Equivalents",[2765,1813,1130,2200])
    for i in range(5): w(ws,7,7+i,f"=CFS!{ec_cols[i]}41",WHITE,GRF,fmt=FMT_NN)

    # Row 8: ST Investments
    act(8,"Short-Term Investments",[0,249,275,447])
    # ST investments: flat at FY25 level (447)
    for i in range(5): w(ws,8,7+i,"=447",WHITE,BKF,fmt=FMT_NN)

    # Row 9: Total Cash & ST Inv
    w(ws,9,2,"Total Cash & Short-Term Investments",WHITE,BKF,True,align=LFT)
    cash_act=[2765,2062,1405,2647]
    for i,v in enumerate(cash_act): w(ws,9,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(5):
        cl=ec_cols[i]; w(ws,9,7+i,f"={cl}7+{cl}8",WHITE,BKF,True,fmt=FMT_NN,bc=GREEN_H)

    # Row 10: Accounts Receivable — driven by DSO
    act(10,"Accounts Receivable",[3549,2534,2674,2789])
    for i,ac2 in enumerate(["C","D","E","F","G"]):
        w(ws,10,7+i,f"=IS!{ec_cols[i]}7*Assumptions!{ac2}24/365",WHITE,BKF,fmt=FMT_NN)

    # Row 11: Other Receivables
    act(11,"Other Receivables",[180,897,843,936])
    for i in range(5): w(ws,11,7+i,"=936",WHITE,BKF,fmt=FMT_NN)

    # Row 12: Total Receivables
    w(ws,12,2,"Total Receivables",WHITE,BKF,True,align=LFT)
    for i,v in enumerate([3729,3431,3517,3725]): w(ws,12,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(5):
        cl=ec_cols[i]; w(ws,12,7+i,f"={cl}10+{cl}11",WHITE,BKF,True,fmt=FMT_NN,bc=GREEN_H)

    # Row 13: Inventory — driven by DIO
    act(13,"Inventory",[8584,9653,9720,10658])
    for i,ac2 in enumerate(["C","D","E","F","G"]):
        cogs_ref=f"-IS!{ec_cols[i]}11"  # COGS is negative on IS, so negate
        w(ws,13,7+i,f"={cogs_ref}*Assumptions!{ac2}25/365",WHITE,BKF,fmt=FMT_NN)

    # Row 14: Prepaid
    act(14,"Prepaid Expenses",[0,288,274,133])
    for i,ac2 in enumerate(["C","D","E","F","G"]):
        w(ws,14,7+i,f"=Assumptions!{ac2}28",WHITE,BKF,fmt=FMT_NN)

    # Row 15: Other Current Assets
    act(15,"Other Current Assets",[572,188,210,334])
    for i,ac2 in enumerate(["C","D","E","F","G"]):
        w(ws,15,7+i,f"=Assumptions!{ac2}27",WHITE,BKF,fmt=FMT_NN)

    # Row 16: Total Current Assets
    w(ws,16,2,"Total Current Assets",WHITE,BKF,True,align=LFT)
    for i,v in enumerate([15650,15622,15126,17497]): w(ws,16,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(5):
        cl=ec_cols[i]; w(ws,16,7+i,f"=SUM({cl}7:{cl}15)",WHITE,BKF,True,fmt=FMT_NN,bc=GREEN_H)

    spacer(ws,17)
    sect(ws,17,"NON-CURRENT ASSETS",11)

    # Row 18: Net PP&E
    act(18,"Net Property, Plant & Equipment",[7076,7738,8509,9528])
    for i,ac2 in enumerate(["C","D","E","F","G"]):
        w(ws,18,7+i,f"=Assumptions!{ac2}29",WHITE,BKF,fmt=FMT_NN)

    # Row 19: LT Investments
    act(19,"Long-Term Investments",[4463,4896,5126,5381])
    for i,ac2 in enumerate(["C","D","E","F","G"]):
        w(ws,19,7+i,f"=Assumptions!{ac2}30",WHITE,BKF,fmt=FMT_NN)

    # Row 20: Goodwill
    act(20,"Goodwill",[2768,2807,2860,2949])
    for i,ac2 in enumerate(["C","D","E","F","G"]):
        w(ws,20,7+i,f"=Assumptions!{ac2}31",WHITE,BKF,fmt=FMT_NN)

    # Row 21: Other Intangibles
    act(21,"Other Intangibles",[11633,11699,11954,11827])
    for i,ac2 in enumerate(["C","D","E","F","G"]):
        w(ws,21,7+i,f"=Assumptions!{ac2}32",WHITE,BKF,fmt=FMT_NN)

    # Row 22: DTA LT
    act(22,"Deferred Tax Assets — Long-Term",[138,178,143,150])
    for i,ac2 in enumerate(["C","D","E","F","G"]):
        w(ws,22,7+i,f"=Assumptions!{ac2}33",WHITE,BKF,fmt=FMT_NN)

    # Row 23: Other LT Assets
    act(23,"Other Long-Term Assets",[2455,1941,1756,1989])
    for i,ac2 in enumerate(["C","D","E","F","G"]):
        w(ws,23,7+i,f"=Assumptions!{ac2}34",WHITE,BKF,fmt=FMT_NN)

    # Row 24: Total Non-Current Assets
    w(ws,24,2,"Total Non-Current Assets",WHITE,BKF,True,align=LFT)
    for i,v in enumerate([28533,29279,30348,31824]): w(ws,24,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(5):
        cl=ec_cols[i]; w(ws,24,7+i,f"=SUM({cl}18:{cl}23)",WHITE,BKF,True,fmt=FMT_NN,bc=GREEN_H)

    spacer(ws,25)
    # Row 26: TOTAL ASSETS
    w(ws,26,2,"TOTAL ASSETS",GREEN_TOT,BKF,True,align=LFT)
    for i,v in enumerate([44183,44883,45474,49322]): w(ws,26,3+i,v,GREEN_TOT,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(5):
        cl=ec_cols[i]; w(ws,26,7+i,f"={cl}16+{cl}24",GREEN_TOT,BKF,True,fmt=FMT_NN,bc=GREEN_H)

    spacer(ws,27)
    sect(ws,28,"LIABILITIES & STOCKHOLDERS' EQUITY")
    sect(ws,29,"CURRENT LIABILITIES",11)

    # Row 30: AP — driven by DPO
    act(30,"Accounts Payable",[7123,3351,3071,3123])
    for i,ac2 in enumerate(["C","D","E","F","G"]):
        cogs_ref=f"-IS!{ec_cols[i]}11"
        w(ws,30,7+i,f"={cogs_ref}*Assumptions!{ac2}26/365",WHITE,BKF,fmt=FMT_NN)

    # Row 31: Accrued Expenses
    act(31,"Accrued Expenses",[0,2644,2646,2957])
    for i,ac2 in enumerate(["C","D","E","F","G"]):
        w(ws,31,7+i,f"=Assumptions!{ac2}35",WHITE,BKF,fmt=FMT_NN)

    # Row 32: ST Borrowings
    act(32,"Short-Term Borrowings",[0,314,514,22])
    for i,ac2 in enumerate(["C","D","E","F","G"]):
        w(ws,32,7+i,f"=Assumptions!{ac2}36",WHITE,BKF,fmt=FMT_NN)

    # Row 33: Current LTD
    act(33,"Current Portion of Long-Term Debt",[1842,1836,2387,2914])
    for i,ac2 in enumerate(["C","D","E","F","G"]):
        w(ws,33,7+i,f"=Assumptions!{ac2}37",WHITE,BKF,fmt=FMT_NN)

    # Row 34: Current Capital Leases
    act(34,"Current Portion of Capital Lease Obligations",[0,94,95,112])
    for i,ac2 in enumerate(["C","D","E","F","G"]):
        w(ws,34,7+i,f"=Assumptions!{ac2}38",WHITE,BKF,fmt=FMT_NN)

    # Row 35: Taxes Payable
    act(35,"Current Income Taxes Payable",[305,170,136,138])
    for i,ac2 in enumerate(["C","D","E","F","G"]):
        w(ws,35,7+i,f"=Assumptions!{ac2}39",WHITE,BKF,fmt=FMT_NN)

    # Row 36: Unearned Revenue
    act(36,"Unearned Revenue — Current",[0,92,84,82])
    for i,ac2 in enumerate(["C","D","E","F","G"]):
        w(ws,36,7+i,f"=Assumptions!{ac2}40",WHITE,BKF,fmt=FMT_NN)

    # Row 37: Other Current Liabilities
    act(37,"Other Current Liabilities",[945,1092,935,1364])
    for i,ac2 in enumerate(["C","D","E","F","G"]):
        w(ws,37,7+i,f"=Assumptions!{ac2}41",WHITE,BKF,fmt=FMT_NN)

    # Row 38: Total Current Liabilities
    w(ws,38,2,"Total Current Liabilities",WHITE,BKF,True,align=LFT)
    for i,v in enumerate([10215,9593,9868,10712]): w(ws,38,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(5):
        cl=ec_cols[i]; w(ws,38,7+i,f"=SUM({cl}30:{cl}37)",WHITE,BKF,True,fmt=FMT_NN,bc=GREEN_H)

    spacer(ws,39)
    sect(ws,39,"NON-CURRENT LIABILITIES",11)

    # Row 40: LT Debt
    act(40,"Long-Term Debt",[17543,19117,18976,21022])
    for i,ac2 in enumerate(["C","D","E","F","G"]):
        w(ws,40,7+i,f"=Assumptions!{ac2}42",WHITE,BKF,fmt=FMT_NN)

    # Row 41: Capital Leases LT
    act(41,"Capital Leases — Long-Term",[0,470,509,541])
    for i,ac2 in enumerate(["C","D","E","F","G"]):
        w(ws,41,7+i,f"=Assumptions!{ac2}43",WHITE,BKF,fmt=FMT_NN)

    # Row 42: Pension
    act(42,"Pension & Other Post-Retirement Benefits",[486,471,429,409])
    for i,ac2 in enumerate(["C","D","E","F","G"]):
        w(ws,42,7+i,f"=Assumptions!{ac2}44",WHITE,BKF,fmt=FMT_NN)

    # Row 43: DTL
    act(43,"Deferred Tax Liability — Non-Current",[2807,2751,2947,2944])
    for i,ac2 in enumerate(["C","D","E","F","G"]):
        w(ws,43,7+i,f"=Assumptions!{ac2}45",WHITE,BKF,fmt=FMT_NN)

    # Row 44: Other NCL
    act(44,"Other Non-Current Liabilities",[1621,772,675,516])
    for i,ac2 in enumerate(["C","D","E","F","G"]):
        w(ws,44,7+i,f"=Assumptions!{ac2}46",WHITE,BKF,fmt=FMT_NN)

    # Row 45: Total Non-Current Liabilities
    w(ws,45,2,"Total Non-Current Liabilities",WHITE,BKF,True,align=LFT)
    for i,v in enumerate([22457,23581,23535,25432]): w(ws,45,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(5):
        cl=ec_cols[i]; w(ws,45,7+i,f"=SUM({cl}40:{cl}44)",WHITE,BKF,True,fmt=FMT_NN,bc=GREEN_H)

    spacer(ws,46)
    # Row 47: Total Liabilities
    w(ws,47,2,"Total Liabilities",WHITE,BKF,True,align=LFT)
    for i,v in enumerate([32672,33174,33403,36144]): w(ws,47,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(5):
        cl=ec_cols[i]; w(ws,47,7+i,f"={cl}38+{cl}45",WHITE,BKF,True,fmt=FMT_NN,bc=GREEN_H)

    spacer(ws,48)
    sect(ws,49,"STOCKHOLDERS' EQUITY")

    # Row 50: Common Stock
    act(50,"Common Stock",[875,898,887,887])
    for i in range(5): w(ws,50,7+i,"=887",WHITE,BKF,fmt=FMT_NN)

    # Row 51: APIC — rolls with stock issuances
    act(51,"Additional Paid-In Capital",[1635,1703,1703,1703])
    # APIC[t] = APIC[t-1] + Stock Issuances
    for i,ac2 in enumerate(["C","D","E","F","G"]):
        prev = "F51" if i==0 else f"{ec_cols[i-1]}51"
        w(ws,51,7+i,f"={prev}+Assumptions!{ac2}57",WHITE,BKF,fmt=FMT_NN)

    # Row 52: Retained Earnings — rolls with NI - Dividends
    act(52,"Retained Earnings",[8490,8876,9783,10274])
    for i,ac2 in enumerate(["C","D","E","F","G"]):
        prev = "F52" if i==0 else f"{ec_cols[i-1]}52"
        ni_ref  = f"IS!{ec_cols[i]}38"
        div_ref = f"IS!{ec_cols[i]}44*Assumptions!{ac2}19"
        w(ws,52,7+i,f"={prev}+{ni_ref}-{div_ref}",WHITE,BKF,fmt=FMT_NN)

    # Row 53: Treasury Stock (flat — no buybacks assumed)
    act(53,"Treasury Stock",[-2223,-2286,-2250,-2228])
    for i in range(5): w(ws,53,7+i,"=-2228",WHITE,BKF,fmt=FMT_NN)

    # Row 54: AOCI (flat at FY25 level)
    act(54,"Comprehensive Income & Other (AOCI)",[658,665,-91,454])
    for i in range(5): w(ws,54,7+i,"=454",WHITE,BKF,fmt=FMT_NN)

    # Row 55: Total Common Equity
    w(ws,55,2,"Total Common Equity",WHITE,BKF,True,align=LFT)
    for i,v in enumerate([9435,9856,10032,11090]): w(ws,55,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(5):
        cl=ec_cols[i]; w(ws,55,7+i,f"=SUM({cl}50:{cl}54)",WHITE,BKF,True,fmt=FMT_NN,bc=GREEN_H)

    # Row 56: Minority Interest in BS (flat)
    act(56,"Minority Interest",[2076,1853,2038,2088])
    for i in range(5): w(ws,56,7+i,"=2088",WHITE,BKF,fmt=FMT_NN)

    # Row 57: Total Equity
    w(ws,57,2,"Total Equity",WHITE,BKF,True,align=LFT)
    for i,v in enumerate([11511,11709,12070,13178]): w(ws,57,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(5):
        cl=ec_cols[i]; w(ws,57,7+i,f"={cl}55+{cl}56",WHITE,BKF,True,fmt=FMT_NN,bc=GREEN_H)

    spacer(ws,58)
    # Row 59: Total L&E
    w(ws,59,2,"TOTAL LIABILITIES & EQUITY",GREEN_TOT,BKF,True,align=LFT)
    for i,v in enumerate([44183,44883,45473,49322]): w(ws,59,3+i,v,GREEN_TOT,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(5):
        cl=ec_cols[i]; w(ws,59,7+i,f"={cl}47+{cl}57",GREEN_TOT,BKF,True,fmt=FMT_NN,bc=GREEN_H)

    # Row 60: Balance Check
    w(ws,60,2,"  Balance Check (Assets − L&E)",WHITE,GYF,italic=True,align=IND)
    for i in range(4): w(ws,60,3+i,f"={get_column_letter(3+i)}26-{get_column_letter(3+i)}59",WHITE,GYF,italic=True,fmt=FMT_NN)
    for i in range(5):
        cl=ec_cols[i]; w(ws,60,7+i,f"={cl}26-{cl}59",WHITE,GYF,italic=True,fmt=FMT_NN)

    return ws

# ═══════════════════════════════════════════════════════════════════════════
# CASH FLOW STATEMENT TAB
# ═══════════════════════════════════════════════════════════════════════════
def build_cfs(wb):
    ws=wb.create_sheet("CFS")
    ws.sheet_properties.tabColor=BROWN
    col_widths(ws)
    ws.row_dimensions[1].height=3.75
    title(ws,2,"Diageo PLC (DGE) — Cash Flow Statement (£M)")
    hdrs(ws,3,[""]+ ACT_YRS + EST_YRS)

    ec_cols=["G","H","I","J","K"]
    asmp=["C","D","E","F","G"]

    spacer(ws,4)
    sect(ws,5,"OPERATING ACTIVITIES")

    # Row 6: Net Income — green link from IS
    w(ws,6,2,"Net Income",WHITE,BKF,align=LFT)
    for i,v in enumerate([4280,4445,3870,2354]): w(ws,6,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(5): w(ws,6,7+i,f"=IS!{ec_cols[i]}38",WHITE,GRF,fmt=FMT_NN)

    # Row 7: D&A
    da_act=[651,597,600,654]
    w(ws,7,2,"  Depreciation & Amortisation",WHITE,BKF,align=IND)
    for i,v in enumerate(da_act): w(ws,7,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(5): w(ws,7,7+i,f"=IS!{ec_cols[i]}50",WHITE,GRF,fmt=FMT_NN)

    # Row 8: Amort of Goodwill & Intangibles
    w(ws,8,2,"  Amortisation of Goodwill & Intangibles",WHITE,BKF,align=IND)
    for i,v in enumerate([0,0,19,20]): w(ws,8,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(5): w(ws,8,7+i,"=20",WHITE,BKF,fmt=FMT_NN)  # flat at FY25

    # Row 9: Amort Deferred Charges
    w(ws,9,2,"  Amortisation of Deferred Charges",WHITE,BKF,align=IND)
    for i,v in enumerate([0,0,58,74]): w(ws,9,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(5): w(ws,9,7+i,"=74",WHITE,BKF,fmt=FMT_NN)

    # Row 10: Asset Writedown
    w(ws,10,2,"  Asset Writedown & Restructuring",WHITE,BKF,align=IND)
    for i,v in enumerate([413,700,-184,590]): w(ws,10,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(5): w(ws,10,7+i,"=0",WHITE,BKF,fmt=FMT_NN)  # normalised to 0

    # Row 11: Equity Income reversal
    w(ws,11,2,"  (Income) Loss on Equity Investments",WHITE,BKF,align=IND)
    for i,v in enumerate([-555,-443,-414,-193]): w(ws,11,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(5): w(ws,11,7+i,f"=-IS!{ec_cols[i]}28",WHITE,GRF,fmt=FMT_NN)

    # Row 12: Other operating
    w(ws,12,2,"  Other Operating Activities",WHITE,BKF,align=IND)
    for i,v in enumerate([665,-249,924,495]): w(ws,12,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(5): w(ws,12,7+i,"=0",WHITE,BKF,fmt=FMT_NN)

    # Working capital changes (derived from BS period changes)
    # Row 13: Change in AR = -(AR[t] - AR[t-1])
    w(ws,13,2,"  Change in Accounts Receivable",WHITE,BKF,align=IND)
    for i,v in enumerate([-502,142,-66,-49]): w(ws,13,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(5):
        cl=ec_cols[i]; prev="F10" if i==0 else f"{ec_cols[i-1]}10"
        w(ws,13,7+i,f"=-(BS!{cl}10-BS!{prev})",WHITE,BKF,fmt=FMT_NN)

    # Row 14: Change in Inventory
    w(ws,14,2,"  Change in Inventories",WHITE,BKF,align=IND)
    for i,v in enumerate([-984,-810,-156,-470]): w(ws,14,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(5):
        cl=ec_cols[i]; prev="F13" if i==0 else f"{ec_cols[i-1]}13"
        w(ws,14,7+i,f"=-(BS!{cl}13-BS!{prev})",WHITE,BKF,fmt=FMT_NN)

    # Row 15: Change in AP
    w(ws,15,2,"  Change in Accounts Payable",WHITE,BKF,align=IND)
    for i,v in enumerate([1245,-746,-546,442]): w(ws,15,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(5):
        cl=ec_cols[i]; prev="F30" if i==0 else f"{ec_cols[i-1]}30"
        w(ws,15,7+i,f"=BS!{cl}30-BS!{prev}",WHITE,BKF,fmt=FMT_NN)

    # Row 16: Net Cash from Operations
    w(ws,16,2,"Net Cash from Operating Activities",WHITE,BKF,True,align=LFT)
    for i,v in enumerate([5213,3636,4105,4297]): w(ws,16,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(5):
        cl=ec_cols[i]; w(ws,16,7+i,f"=SUM({cl}6:{cl}15)",WHITE,BKF,True,fmt=FMT_NN,bc=BROWN)

    spacer(ws,17)
    sect(ws,18,"INVESTING ACTIVITIES")

    # Row 19: Capex
    w(ws,19,2,"Capital Expenditure",WHITE,BKF,align=LFT)
    for i,v in enumerate([-1457,-1417,-1510,-1612]): w(ws,19,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i,ac2 in enumerate(asmp): w(ws,19,7+i,f"=-IS!{ec_cols[i]}7*Assumptions!{ac2}49",WHITE,BKF,fmt=FMT_NN)

    # Row 20: Sale of PP&E
    w(ws,20,2,"Sale of Property, Plant & Equipment",WHITE,BKF,align=LFT)
    for i,v in enumerate([23,16,14,63]): w(ws,20,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i,ac2 in enumerate(asmp): w(ws,20,7+i,f"=Assumptions!{ac2}50",WHITE,BKF,fmt=FMT_NN)

    # Row 21: Cash Acquisitions
    w(ws,21,2,"Cash Acquisitions",WHITE,BKF,align=LFT)
    for i,v in enumerate([-278,-404,-6,-35]): w(ws,21,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i,ac2 in enumerate(asmp): w(ws,21,7+i,f"=Assumptions!{ac2}51",WHITE,BKF,fmt=FMT_NN)

    # Row 22: Divestitures
    w(ws,22,2,"Divestitures",WHITE,BKF,align=LFT)
    for i,v in enumerate([102,559,87,143]): w(ws,22,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i,ac2 in enumerate(asmp): w(ws,22,7+i,f"=Assumptions!{ac2}52",WHITE,BKF,fmt=FMT_NN)

    # Row 23: Investment in Securities
    w(ws,23,2,"Investment in Marketable & Equity Securities",WHITE,BKF,align=LFT)
    for i,v in enumerate([-86,-112,-133,-84]): w(ws,23,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i,ac2 in enumerate(asmp): w(ws,23,7+i,f"=Assumptions!{ac2}53",WHITE,BKF,fmt=FMT_NN)

    # Row 24: Net Loans
    w(ws,24,2,"Net (Increase) Decrease in Loans",WHITE,BKF,align=LFT)
    for i,v in enumerate([-96,-68,-47,-195]): w(ws,24,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i,ac2 in enumerate(asmp): w(ws,24,7+i,f"=Assumptions!{ac2}54",WHITE,BKF,fmt=FMT_NN)

    # Row 25: Net Cash from Investing
    w(ws,25,2,"Net Cash from Investing Activities",WHITE,BKF,True,align=LFT)
    for i,v in enumerate([-1792,-1426,-1595,-1720]): w(ws,25,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(5):
        cl=ec_cols[i]; w(ws,25,7+i,f"=SUM({cl}19:{cl}24)",WHITE,BKF,True,fmt=FMT_NN,bc=BROWN)

    spacer(ws,26)
    sect(ws,27,"FINANCING ACTIVITIES")

    # Row 28: Debt Issued
    w(ws,28,2,"Total Debt Issued",WHITE,BKF,align=LFT)
    for i,v in enumerate([3638,3058,2612,4026]): w(ws,28,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i,ac2 in enumerate(asmp): w(ws,28,7+i,f"=Assumptions!{ac2}55",WHITE,BKF,fmt=FMT_NN)

    # Row 29: Debt Repaid
    w(ws,29,2,"Total Debt Repaid",WHITE,BKF,align=LFT)
    for i,v in enumerate([-2622,-2102,-2160,-3128]): w(ws,29,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i,ac2 in enumerate(asmp): w(ws,29,7+i,f"=Assumptions!{ac2}56",WHITE,BKF,fmt=FMT_NN)

    # Row 30: Stock Issuances
    w(ws,30,2,"Issuance of Common Stock",WHITE,BKF,align=LFT)
    for i,v in enumerate([24,36,21,15]): w(ws,30,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i,ac2 in enumerate(asmp): w(ws,30,7+i,f"=Assumptions!{ac2}57",WHITE,BKF,fmt=FMT_NN)

    # Row 31: Buybacks
    w(ws,31,2,"Repurchase of Common Stock",WHITE,BKF,align=LFT)
    for i,v in enumerate([-2985,-1673,-987,0]): w(ws,31,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i,ac2 in enumerate(asmp): w(ws,31,7+i,f"=Assumptions!{ac2}58",WHITE,BKF,fmt=FMT_NN)

    # Row 32: Common Dividends
    w(ws,32,2,"Common Dividends Paid",WHITE,BKF,align=LFT)
    for i,v in enumerate([-2300,-2065,-2242,-2298]): w(ws,32,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i in range(5): w(ws,32,7+i,f"=-IS!{ec_cols[i]}44*Assumptions!{asmp[i]}19",WHITE,BKF,fmt=FMT_NN)

    # Row 33: Other Financing
    w(ws,33,2,"Other Financing Activities",WHITE,BKF,align=LFT)
    for i,v in enumerate([-128,-295,-350,-109]): w(ws,33,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i,ac2 in enumerate(asmp): w(ws,33,7+i,f"=Assumptions!{ac2}59",WHITE,BKF,fmt=FMT_NN)

    # Row 34: Net Cash from Financing
    w(ws,34,2,"Net Cash from Financing Activities",WHITE,BKF,True,align=LFT)
    for i,v in enumerate([-4373,-3041,-3106,-1494]): w(ws,34,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(5):
        cl=ec_cols[i]; w(ws,34,7+i,f"=SUM({cl}28:{cl}33)",WHITE,BKF,True,fmt=FMT_NN,bc=BROWN)

    spacer(ws,35)
    # Row 36: FX
    w(ws,36,2,"Effect of FX on Cash",WHITE,BKF,align=LFT)
    for i,v in enumerate([-38,-76,-33,-35]): w(ws,36,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i,ac2 in enumerate(asmp): w(ws,36,7+i,f"=Assumptions!{ac2}60",WHITE,BKF,fmt=FMT_NN)

    # Row 37: Misc
    w(ws,37,2,"Miscellaneous Cash Flow Adjustments",WHITE,BKF,align=LFT)
    for i,v in enumerate([0,0,-30,21]): w(ws,37,3+i,v,WHITE,BLF,fmt=FMT_NN)
    for i,ac2 in enumerate(asmp): w(ws,37,7+i,f"=Assumptions!{ac2}61",WHITE,BKF,fmt=FMT_NN)

    # Row 38: Net Change in Cash
    w(ws,38,2,"Net Change in Cash",WHITE,BKF,True,align=LFT)
    for i,v in enumerate([-990,-907,-659,1069]): w(ws,38,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(5):
        cl=ec_cols[i]; w(ws,38,7+i,f"={cl}16+{cl}25+{cl}34+{cl}36+{cl}37",WHITE,BKF,True,fmt=FMT_NN,bc=BROWN)

    # Row 39: Beginning Cash
    w(ws,39,2,"Beginning Cash",WHITE,BKF,align=LFT)
    for i,v in enumerate([3665,2675,1768,1109]): w(ws,39,3+i,v,WHITE,BLF,fmt=FMT_NN)
    # FY26E beginning = FY25 ending cash from BS
    w(ws,39,7,"=BS!F7",WHITE,GRF,fmt=FMT_NN)
    for i in range(1,5): w(ws,39,7+i,f"={ec_cols[i-1]}41",WHITE,BKF,fmt=FMT_NN)

    # Row 40: Beginning Cash for FY25 actual = 1109, matches
    # Row 40 reserved — skip (use 41 for ending)
    spacer(ws,40,h=3)

    # Row 41: Ending Cash — THIS is what BS!Cash links to
    w(ws,41,2,"Ending Cash",WHITE,BKF,True,align=LFT)
    for i,v in enumerate([2675,1768,1109,2200]): w(ws,41,3+i,v,WHITE,BLF,True,fmt=FMT_NN,bc=BLUE_H)
    for i in range(5):
        cl=ec_cols[i]; w(ws,41,7+i,f"={cl}39+{cl}38",WHITE,BKF,True,fmt=FMT_NN,bc=BROWN)

    spacer(ws,42)
    sect(ws,43,"SUPPLEMENTAL")

    # Row 44: FCF
    w(ws,44,2,"  Free Cash Flow (CFO + Capex)",WHITE,GYF,italic=True,align=IND)
    for i,v in enumerate([3756,2219,2595,2685]): w(ws,44,3+i,v,WHITE,BLF,italic=True,fmt=FMT_NN)
    for i in range(5):
        cl=ec_cols[i]; w(ws,44,7+i,f"={cl}16+{cl}19",WHITE,GYF,italic=True,fmt=FMT_NN)

    # Row 45: FCF Margin
    w(ws,45,2,"  % FCF Margin",WHITE,GYF,italic=True,align=IND)
    for c in range(3,12): w(ws,45,c,f"={get_column_letter(c)}44/IS!{get_column_letter(c)}7",WHITE,GYF,italic=True,fmt=FMT_P)

    return ws

# ═══════════════════════════════════════════════════════════════════════════
# DCF TAB
# ═══════════════════════════════════════════════════════════════════════════
def build_dcf(wb):
    ws=wb.create_sheet("DCF")
    ws.sheet_properties.tabColor=RED_BG
    ws.column_dimensions['A'].width=1
    ws.column_dimensions['B'].width=38
    ws.column_dimensions['C'].width=18
    ws.column_dimensions['D'].width=14
    for i in range(7): ws.column_dimensions[get_column_letter(5+i)].width=13
    ws.column_dimensions['L'].width=14
    ws.column_dimensions['M'].width=14

    ws.row_dimensions[1].height=3.75
    title(ws,2,"Diageo PLC (DGE) — DCF Valuation (£M)",13)

    ec_cols=["G","H","I","J","K"]  # IS/CFS estimate columns

    # ── SECTION 1: DCF ASSUMPTIONS ─────────────────────────────────────────
    spacer(ws,3)
    sect(ws,4,"DCF ASSUMPTIONS",13)
    w(ws,5,2,"Assumption",BLUE_H,WHF,True,align=LFT)
    w(ws,5,4,"Value",BLUE_H,WHF,True,align=CTR)
    w(ws,5,5,"Notes",BLUE_H,WHF,True,align=LFT,mc=8)

    dcf_asmp=[
        (6,  "WACC",                          0.090,  FMT_P,  "9.0% — blended cost of capital; Diageo large-cap spirits, ~60% debt"),
        (7,  "Terminal Growth Rate (TGR)",     0.025,  FMT_P,  "2.5% — long-run nominal GDP growth for premium spirits"),
        (8,  "Projection Period (years)",      5,      FMT_N,  "FY26E–FY30E (5 years from consensus estimates)"),
        (9,  "Cash & Equivalents (£M)",        2200,   FMT_NN, "FY25A ending cash; Source: Balance Sheet FY25"),
        (10, "Short-Term Investments (£M)",    447,    FMT_NN, "FY25A; Source: Balance Sheet FY25"),
        (11, "Total Debt (£M)",                24611,  FMT_NN, "FY25A total debt; Source: Balance Sheet FY25"),
        (12, "Net Cash (Cash+STI−Debt) (£M)", "=(D9+D10-D11)", FMT_NN,"Computed from rows above"),
        (13, "Diluted Shares Outstanding (M)", 2228,   FMT_N,  "FY25A; Source: Balance Sheet filing"),
        (14, "Current Share Price (£)",        16.04,  FMT_EP, "As of 30-Jun-2025 per TIKR"),
        (15, "NTM EV/EBITDA (comps ref)",      13.0,   FMT_MX, "Peer comps: spirits sector ~11–15x; Diageo at discount"),
    ]
    for r,lbl,val,fmt,note in dcf_asmp:
        w(ws,r,2,lbl,WHITE,BKF,align=LFT)
        fc=BKF if isinstance(val,str) and val.startswith('=') else BLF
        w(ws,r,4,val,WHITE,fc,fmt=fmt,align=CTR)
        w(ws,r,5,note,WHITE,GYF,italic=True,align=LFT,mc=8)

    # ── SECTION 2: FCF PROJECTIONS ──────────────────────────────────────────
    spacer(ws,15)
    # Section header in col B only (no merge so col headers can share the row below)
    w(ws,16,2,"FREE CASH FLOW PROJECTIONS",SECT,SECT_TXT,True,align=LFT)
    yr_hdrs=["","Base (FY25A)"]+[f"FY{26+i}E" for i in range(5)]+["Terminal"]
    for i,lbl in enumerate(yr_hdrs):
        is_e=i>1
        w(ws,16,2+i,lbl,GREEN_H if is_e else BLUE_H,WHF,True,align=CTR)

    # IS col for base=F, estimates=G..K; Terminal=col J (last estimate)
    base_col="F"; est_map={"G":"G","H":"H","I":"I","J":"J","K":"K"}

    fcf_rows=[
        (17,"Revenue",            "IS!{c}7",   FMT_NN,False),
        (18,"  % Revenue Growth", None,        FMT_P, True),
        (19,"EBIT",               "IS!{c}22",  FMT_NN,False),
        (20,"  % EBIT Margin",    None,        FMT_P, True),
        (21,"Tax on EBIT",        None,        FMT_NN,False),
        (22,"NOPAT",              None,        FMT_NN,False),
        (23,"(+) D&A",            "IS!{c}50",  FMT_NN,False),
        (24,"(+) SBC",            "=0",        FMT_NN,False),  # No SBC line in DGE IS
        (25,"(−) Capex",          "CFS!{c}19", FMT_NN,False),
        (26,"(+/−) Change in WC", None,        FMT_NN,False),
        (27,"Unlevered FCF",      None,        FMT_NN,True),
        (28,"  % uFCF Margin",    None,        FMT_P, True),
    ]

    for r,lbl,ref_tmpl,fmt,bold in fcf_rows:
        is_grey=(lbl.startswith("  ") and "%" in lbl)
        fc=GYF if is_grey else BKF
        w(ws,r,2,lbl,WHITE,fc,bold,italic=is_grey,align=IND if is_grey else LFT)
        # Base year (col D)
        if ref_tmpl and "{c}" in ref_tmpl:
            val=f"={ref_tmpl.format(c=base_col)}"
            fc2=GRF
            w(ws,r,4,val,WHITE,fc2,bold,fmt=fmt,align=CTR)
        elif r==21: w(ws,r,4,f"=IF(IS!F22>0,-IS!F22*0.241,0)",WHITE,BKF,False,fmt=fmt,align=CTR)
        elif r==22: w(ws,r,4,f"=D19+D21",WHITE,BKF,False,fmt=fmt,align=CTR)
        elif r==26: w(ws,r,4,"=0",WHITE,BKF,False,fmt=fmt,align=CTR)
        elif r==27: w(ws,r,4,f"=D22+D23+D24+D25+D26",WHITE,BKF,True,fmt=fmt,align=CTR,bc=RED_BG)
        elif r==18: w(ws,r,4,"",WHITE,GYF,False,italic=True,fmt=fmt,align=CTR)
        elif r==20: w(ws,r,4,f"=D19/D17",WHITE,GYF,False,italic=True,fmt=fmt,align=CTR)
        elif r==28: w(ws,r,4,f"=D27/D17",WHITE,GYF,False,italic=True,fmt=fmt,align=CTR)
        elif r==24: w(ws,r,4,"=0",WHITE,BKF,False,fmt=fmt,align=CTR)

        # Estimate years (cols E..I = 5..9)
        asmp_cols2=["C","D","E","F","G"]
        for i,ec in enumerate(ec_cols):
            col_idx=5+i
            if ref_tmpl and "{c}" in ref_tmpl:
                val=f"={ref_tmpl.format(c=ec)}"
                w(ws,r,col_idx,val,WHITE,GRF,bold,fmt=fmt,align=CTR,bc=RED_BG if r==27 else None)
            elif r==18:
                prev_c=f"D17" if i==0 else f"{get_column_letter(col_idx-1)}17"
                w(ws,r,col_idx,f"={get_column_letter(col_idx)}17/{prev_c}-1",WHITE,GYF,False,italic=True,fmt=fmt,align=CTR)
            elif r==20:
                w(ws,r,col_idx,f"={get_column_letter(col_idx)}19/{get_column_letter(col_idx)}17",WHITE,GYF,False,italic=True,fmt=fmt,align=CTR)
            elif r==21:
                w(ws,r,col_idx,f"=IF(IS!{ec}22>0,-IS!{ec}22*Assumptions!{asmp_cols2[i]}15,0)",WHITE,BKF,False,fmt=fmt,align=CTR)
            elif r==22:
                w(ws,r,col_idx,f"={get_column_letter(col_idx)}19+{get_column_letter(col_idx)}21",WHITE,BKF,False,fmt=fmt,align=CTR)
            elif r==24:
                w(ws,r,col_idx,"=0",WHITE,BKF,False,fmt=fmt,align=CTR)
            elif r==26:
                # Change in WC = -(change in AR + change in Inv - change in AP)
                w(ws,r,col_idx,f"=-(CFS!{ec}13+CFS!{ec}14+CFS!{ec}15)",WHITE,BKF,False,fmt=fmt,align=CTR)
            elif r==27:
                w(ws,r,col_idx,f"=SUM({get_column_letter(col_idx)}22:{get_column_letter(col_idx)}26)",WHITE,BKF,True,fmt=fmt,align=CTR,bc=RED_BG)
            elif r==28:
                w(ws,r,col_idx,f"={get_column_letter(col_idx)}27/{get_column_letter(col_idx)}17",WHITE,GYF,False,italic=True,fmt=fmt,align=CTR)

        # Terminal column (col J = 10)
        if r==17:   w(ws,r,10,f"=I17*(1+D7)",WHITE,BKF,False,fmt=fmt,align=CTR)
        elif r==18: w(ws,r,10,f"=D7",WHITE,GYF,False,italic=True,fmt=FMT_P,align=CTR)
        elif r==19: w(ws,r,10,f"=I19*(1+D7)",WHITE,BKF,False,fmt=fmt,align=CTR)
        elif r==20: w(ws,r,10,f"=J19/J17",WHITE,GYF,False,italic=True,fmt=FMT_P,align=CTR)
        elif r==21: w(ws,r,10,f"=IF(J19>0,-J19*D6,0)",WHITE,BKF,False,fmt=fmt,align=CTR)
        elif r==22: w(ws,r,10,f"=J19+J21",WHITE,BKF,False,fmt=fmt,align=CTR)
        elif r==23: w(ws,r,10,f"=I23*(1+D7)",WHITE,GRF,False,fmt=fmt,align=CTR)
        elif r==24: w(ws,r,10,"=0",WHITE,BKF,False,fmt=fmt,align=CTR)
        elif r==25: w(ws,r,10,f"=I25*(1+D7)",WHITE,GRF,False,fmt=fmt,align=CTR)
        elif r==26: w(ws,r,10,"=0",WHITE,BKF,False,fmt=fmt,align=CTR)
        elif r==27: w(ws,r,10,f"=J22+J23+J24+J25+J26",WHITE,BKF,True,fmt=fmt,align=CTR,bc=RED_BG)
        elif r==28: w(ws,r,10,f"=J27/J17",WHITE,GYF,False,italic=True,fmt=FMT_P,align=CTR)

    # ── SECTION 3: DCF CALCULATION ─────────────────────────────────────────
    spacer(ws,29)
    w(ws,30,2,"DCF CALCULATION",SECT,SECT_TXT,True,align=LFT)
    w(ws,30,3,"Base",BLUE_H,WHF,True,align=CTR)
    for i in range(5): w(ws,30,5+i,f"Yr {i+1}",BLUE_H,WHF,True,align=CTR)
    w(ws,30,10,"Terminal",BLUE_H,WHF,True,align=CTR)

    # Row 31: Period
    w(ws,31,2,"Period",WHITE,GYF,italic=True,align=LFT)
    for i in range(5): w(ws,31,5+i,i+1,WHITE,GYF,italic=True,fmt=FMT_N,align=CTR)
    w(ws,31,10,"∞",WHITE,GYF,italic=True,align=CTR)

    # Row 32: Discount factor
    w(ws,32,2,"Discount Factor",WHITE,GYF,italic=True,align=LFT)
    for i in range(5): w(ws,32,5+i,f"=1/(1+DCF!D6)^{i+1}",WHITE,GYF,italic=True,fmt="0.0000",align=CTR)
    w(ws,32,10,f"=1/(1+DCF!D6)^5",WHITE,GYF,italic=True,fmt="0.0000",align=CTR)

    # Row 33: uFCF
    w(ws,33,2,"Unlevered FCF (£M)",WHITE,BKF,align=LFT)
    for i in range(5): w(ws,33,5+i,f"={get_column_letter(5+i)}27",WHITE,GRF,fmt=FMT_NN,align=CTR)
    # Terminal Value (Gordon Growth)
    w(ws,34,2,"Terminal Value (Gordon Growth) (£M)",WHITE,BKF,align=LFT)
    w(ws,34,10,f"=J27*(1+D7)/(D6-D7)",WHITE,BKF,fmt=FMT_NN,align=CTR)

    # Row 35: PV of each FCF
    w(ws,35,2,"PV of FCF (£M)",WHITE,BKF,align=LFT)
    for i in range(5): w(ws,35,5+i,f"={get_column_letter(5+i)}33*{get_column_letter(5+i)}32",WHITE,BKF,fmt=FMT_NN,align=CTR,bc=RED_BG)
    w(ws,35,10,f"=J34*J32",WHITE,BKF,fmt=FMT_NN,align=CTR,bc=RED_BG)

    # ── SECTION 4: VALUATION BRIDGE ───────────────────────────────────────
    spacer(ws,36)
    w(ws,37,2,"VALUATION BRIDGE",SECT,SECT_TXT,True,align=LFT)

    bridge=[
        (38,"Sum of PV of FCFs (£M)",         f"=SUM(E35:I35)",          FMT_NN,BKF,WHITE),
        (39,"(+) PV of Terminal Value (£M)",   f"=J35",                   FMT_NN,GRF,WHITE),
        (40,"Enterprise Value (£M)",           f"=D38+D39",               FMT_NN,BKF,GREEN_TOT),
        (41,"Enterprise Value (£B)",           f"=D40/1000",              "#,##0.0",BKF,WHITE),
        (42,"  % from Terminal Value",         f"=D39/D40",               FMT_P, GYF,WHITE),
        (43,"(+) Net Cash (£M)",               f"=D12",                   FMT_NN,GRF,WHITE),
        (44,"Equity Value (£M)",               f"=D40+D43",               FMT_NN,BKF,GREEN_TOT),
        (45,"(÷) Diluted Shares (M)",          f"=D13",                   FMT_N, GRF,WHITE),
        (46,"Intrinsic Value Per Share (£)",   f"=D44/D45",               FMT_EP,WHF,RED_BG),
        (47,"Current Share Price (£)",         f"=D14",                   FMT_EP,GRF,WHITE),
        (48,"  Upside / (Downside)",           f"=D46/D47-1",             FMT_P, GYF,WHITE),
    ]
    for r,lbl,formula,fmt,fc,bg in bridge:
        bold=(r in [40,44,46]); italic=(fc==GYF)
        w(ws,r,2,lbl,bg,fc if r!=46 else WHF,bold,italic,align=LFT)
        w(ws,r,4,formula,bg,fc if r!=46 else WHF,bold,italic,fmt=fmt,align=CTR)

    # Cross-checks
    spacer(ws,49)
    cchk=[
        (50,"  TV as % of EV",            f"=D39/D40",        FMT_P),
        (51,"  Implied EV / NTM EBITDA",  f"=D40/IS!G51",     FMT_MX),
        (52,"  Implied EV / NTM Revenue", f"=D40/IS!G7",      FMT_MX),
        (53,"  NTM EBITDA (£M)",          f"=IS!G51",         FMT_NN),
    ]
    for r,lbl,formula,fmt in cchk:
        w(ws,r,2,lbl,WHITE,GYF,italic=True,align=IND)
        w(ws,r,4,formula,WHITE,GYF,italic=True,fmt=fmt,align=CTR)

    # ── SECTION 5: SENSITIVITY ANALYSIS ───────────────────────────────────
    spacer(ws,54)
    w(ws,55,2,"SENSITIVITY ANALYSIS — IMPLIED SHARE PRICE (£)",SECT,SECT_TXT,True,align=LFT,mc=13)

    # Table 1: WACC × TGR
    w(ws,56,3,"WACC \\ TGR",NAVY,WHF,True,align=CTR)
    tgr_vals=[0.015,0.020,0.025,0.030,0.035,0.040]
    wacc_vals=[0.075,0.080,0.090,0.100,0.105]
    base_wacc=0.090; base_tgr=0.025

    for j,tgr in enumerate(tgr_vals):
        cell=ws.cell(row=56,column=4+j)
        cell.value=tgr; cell.number_format=FMT_P
        cell.font=fn(WHF,True); cell.fill=fl(NAVY); cell.alignment=CTR

    for i,wacc in enumerate(wacc_vals):
        row=57+i
        hdr_cell=ws.cell(row=row,column=3)
        hdr_cell.value=wacc; hdr_cell.number_format=FMT_P
        is_base_row=(abs(wacc-base_wacc)<0.001)
        hdr_cell.font=fn(WHF,True); hdr_cell.fill=fl(RED_BG if is_base_row else NAVY); hdr_cell.alignment=CTR
        for j,tgr in enumerate(tgr_vals):
            col=4+j
            is_base=(abs(wacc-base_wacc)<0.001 and abs(tgr-base_tgr)<0.001)
            # Formula: TV/(WACC-TGR) approach
            # Equity Value = (Sum PV FCFs using wacc) + TV/(wacc-tgr)*disc_factor + Net Cash
            # Simplified: adjust D40 formula with local wacc/tgr
            # Use explicit formula referencing the FCF rows
            pv_fcf="+".join([f"E{27-(27-27)+i2*0+27}/((1+{wacc})^{i2+1})" for i2 in range(5)])
            # Actually build a clean formula
            fcf_pv="+".join([f"{get_column_letter(5+i2)}27/((1+{wacc})^{i2+1})" for i2 in range(5)])
            tv_formula=f"J27*(1+{tgr})/({wacc}-{tgr})/(1+{wacc})^5"
            eq_val=f"(({fcf_pv}+{tv_formula})+D12)/D13"
            cell=ws.cell(row=row,column=col)
            cell.value=f"={eq_val}"
            cell.number_format=FMT_EP
            cell.font=fn(BKF,is_base); cell.alignment=CTR
            cell.fill=fl(YELLOW if is_base else WHITE)
    w(ws,62,3,"Base case: WACC=9.0%, TGR=2.5%",WHITE,GYF,italic=True,align=LFT,mc=9)
    ws.row_dimensions[62].height=12

    # Table 2: WACC × Exit Multiple (EV/NTM EBITDA)
    mult_vals=[9.0,10.0,11.0,13.0,14.0,15.0]
    base_mult=13.0
    spacer(ws,63)
    w(ws,64,3,"WACC \\ EV/EBITDA",NAVY,WHF,True,align=CTR)
    for j,m in enumerate(mult_vals):
        cell=ws.cell(row=64,column=4+j)
        cell.value=m; cell.number_format=FMT_MX
        cell.font=fn(WHF,True); cell.fill=fl(NAVY); cell.alignment=CTR

    for i,wacc in enumerate(wacc_vals):
        row=65+i
        hdr_cell=ws.cell(row=row,column=3)
        hdr_cell.value=wacc; hdr_cell.number_format=FMT_P
        is_base_row=(abs(wacc-base_wacc)<0.001)
        hdr_cell.font=fn(WHF,True); hdr_cell.fill=fl(RED_BG if is_base_row else NAVY); hdr_cell.alignment=CTR
        for j,m in enumerate(mult_vals):
            col=4+j
            is_base=(abs(wacc-base_wacc)<0.001 and abs(m-base_mult)<0.01)
            fcf_pv="+".join([f"{get_column_letter(5+i2)}27/((1+{wacc})^{i2+1})" for i2 in range(5)])
            tv_exit=f"IS!K51*{m}/(1+{wacc})^5"
            eq_val=f"(({fcf_pv}+{tv_exit})+D12)/D13"
            cell=ws.cell(row=row,column=col)
            cell.value=f"={eq_val}"
            cell.number_format=FMT_EP
            cell.font=fn(BKF,is_base); cell.alignment=CTR
            cell.fill=fl(YELLOW if is_base else WHITE)
    w(ws,70,3,"Base case: WACC=9.0%, Exit EV/EBITDA=13.0x",WHITE,GYF,italic=True,align=LFT,mc=9)
    ws.row_dimensions[70].height=12

    # ── SECTION 6: EV/NTM REVENUE MULTIPLES ───────────────────────────────
    spacer(ws,71)
    w(ws,72,2,"EV / NTM REVENUE MULTIPLES",SECT,SECT_TXT,True,align=LFT,mc=13)
    # NTM Revenue inputs
    w(ws,73,2,"NTM Revenue (£M)",WHITE,BKF,align=LFT)
    w(ws,73,4,"=IS!G7",WHITE,GRF,fmt=FMT_NN,align=CTR)
    w(ws,74,2,"Net Cash (£M)",WHITE,BKF,align=LFT)
    w(ws,74,4,"=D12",WHITE,GRF,fmt=FMT_NN,align=CTR)
    w(ws,75,2,"Diluted Shares (M)",WHITE,BKF,align=LFT)
    w(ws,75,4,"=D13",WHITE,GRF,fmt=FMT_NN,align=CTR)
    w(ws,76,2,"Peer context: premium spirits peers (DEO comps) trade 2.5x–4.5x NTM Revenue; DGE at discount given organic growth concerns",WHITE,GYF,italic=True,align=LFT,mc=12)
    ws.row_dimensions[76].height=14

    mult_rev=[2.0,2.5,3.0,3.5,4.0,4.5,5.0]
    base_rev=3.0
    w(ws,77,3,"Multiple →",NAVY,WHF,True,align=CTR)
    for j,m in enumerate(mult_rev):
        cell=ws.cell(row=77,column=4+j)
        cell.value=m; cell.number_format=FMT_MX
        cell.font=fn(WHF,True); cell.fill=fl(NAVY); cell.alignment=CTR
    # Row 78: Implied EV
    w(ws,78,2,"Implied EV (£M)",WHITE,BKF,align=LFT)
    for j,m in enumerate(mult_rev):
        is_base=abs(m-base_rev)<0.01
        w(ws,78,4+j,f"=D73*{m}",LT_BLUE,BKF,fmt=FMT_NN,align=CTR)
        ws.cell(row=78,column=4+j).fill=fl(YELLOW if is_base else LT_BLUE)
    # Row 79: Equity Value
    w(ws,79,2,"Equity Value (£M)",WHITE,BKF,align=LFT)
    for j,m in enumerate(mult_rev):
        is_base=abs(m-base_rev)<0.01
        w(ws,79,4+j,f"={get_column_letter(4+j)}78+D74",WHITE,BKF,fmt=FMT_NN,align=CTR)
        ws.cell(row=79,column=4+j).fill=fl(YELLOW if is_base else WHITE)
    # Row 80: Implied Share Price
    w(ws,80,2,"Implied Share Price (£)",WHITE,WHF,True,align=LFT)
    for j,m in enumerate(mult_rev):
        is_base=abs(m-base_rev)<0.01
        w(ws,80,4+j,f"={get_column_letter(4+j)}79/D75",GREEN_H,WHF,True,fmt=FMT_EP,align=CTR)
        ws.cell(row=80,column=4+j).fill=fl(YELLOW if is_base else GREEN_H)
    w(ws,81,3,"Base case: 3.0x NTM Revenue. NTM period = FY26E ending June 2026.",WHITE,GYF,italic=True,align=LFT,mc=10)
    ws.row_dimensions[81].height=12

    return ws

# ═══════════════════════════════════════════════════════════════════════════
# MAIN
# ═══════════════════════════════════════════════════════════════════════════
def main():
    wb=Workbook()
    wb.remove(wb.active)
    # Build in order: Assumptions first (referenced by IS/BS/CFS), then the rest
    build_assumptions(wb)
    build_is(wb)
    build_bs(wb)
    build_cfs(wb)
    build_dcf(wb)

    # Reorder sheets: IS, BS, CFS, Assumptions, DCF
    # Reorder: move each sheet to position i
    sheet_order=["IS","BS","CFS","Assumptions","DCF"]
    for target_idx, name in enumerate(sheet_order):
        current_idx = wb.sheetnames.index(name)
        if current_idx != target_idx:
            wb.move_sheet(name, offset=target_idx - current_idx)

    OUT.parent.mkdir(parents=True,exist_ok=True)
    wb.save(OUT)
    print(f"Saved → {OUT}")

if __name__=="__main__":
    main()
